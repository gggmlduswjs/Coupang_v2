"""광고 추천 엑셀 리포트 생성 — 키워드 입찰가 포함"""
import json
import os
import re
import sys

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.config import AnalysisConfig
from core.database import CoupangDB


# ═══════════════════════════════════════
# 키워드 추출 + 입찰가 산출
# ═══════════════════════════════════════

# 과목/분야 키워드
SUBJECTS = [
    "국어", "수학", "영어", "과학", "사회", "역사", "도덕", "윤리",
    "물리학", "물리", "화학", "생명과학", "생명", "지구과학", "지구",
    "통합과학", "통합사회", "한국사", "세계사", "세계지리", "한국지리",
    "경제", "정치", "법과정치", "사회문화", "생활과윤리", "윤리와사상",
    "확률과통계", "미적분", "기하", "대수",
    "독서", "문학", "화법", "작문", "언어", "문법", "듣기",
    "컴활", "컴퓨터활용능력", "정보처리", "정보처리기사", "정보처리기능사",
    "바리스타", "제과", "제빵", "웹디자인", "GTQ", "포토샵", "일러스트",
]

# 학년/레벨 키워드
LEVELS = [
    "초등", "중등", "중학", "고등", "고1", "고2", "고3", "수능",
    "1학년", "2학년", "3학년", "1-1", "1-2", "2-1", "2-2", "3-1", "3-2",
    "4-1", "4-2", "5-1", "5-2", "6-1", "6-2",
    "필기", "실기", "1급", "2급",
]


def _extract_ad_keywords(product_name: str, main_keyword: str) -> list[dict]:
    """상품명에서 광고 입찰 키워드 후보를 추출.

    반환: [{"keyword": str, "type": "대표|세부|롱테일", "priority": int}, ...]
    """
    name = product_name or ""
    mk = main_keyword or ""
    results = []
    seen = set()

    def _add(kw, kw_type, priority):
        kw = kw.strip()
        if kw and kw not in seen and len(kw) >= 2:
            seen.add(kw)
            results.append({"keyword": kw, "type": kw_type, "priority": priority})

    def _already_in_main(token: str) -> bool:
        """토큰이 메인 키워드에 이미 포함되어 있는지 확인."""
        return token in mk

    # 1. 대표 키워드 (매칭된 메인 키워드)
    _add(main_keyword, "대표", 1)

    # 2. 상품명에서 과목 추출 (메인 키워드에 없는 것만)
    found_subjects = []
    for subj in SUBJECTS:
        if subj in name:
            found_subjects.append(subj)
            if not _already_in_main(subj):
                _add(f"{main_keyword} {subj}", "세부", 2)

    # 3. 학년/레벨 추출
    found_levels = []
    for level in LEVELS:
        if level in name:
            found_levels.append(level)

    # 4. 브랜드+레벨 세부 키워드 (메인 키워드에 없는 것만)
    for level in found_levels[:2]:
        if not _already_in_main(level):
            _add(f"{main_keyword} {level}", "세부", 3)

    # 5. 롱테일: 브랜드+과목+레벨 (중복 토큰 제거)
    for subj in found_subjects[:2]:
        for level in found_levels[:2]:
            parts = [main_keyword]
            if not _already_in_main(subj):
                parts.append(subj)
            if not _already_in_main(level):
                parts.append(level)
            if len(parts) >= 2:  # 메인 외에 최소 1개 추가돼야 의미 있음
                _add(" ".join(parts), "롱테일", 4)

    # 6. 과목+레벨 단독 조합 (검색 볼륨용)
    for subj in found_subjects[:2]:
        for level in found_levels[:1]:
            _add(f"{subj} {level} 문제집", "롱테일", 5)

    # 7. 상품명에서 핵심 2~3어절 추출 (세트, 기출, 모의고사 등)
    specials = ["기출문제집", "기출문제", "모의고사", "기본서", "문제집", "세트"]
    for sp in specials:
        if sp in name and not _already_in_main(sp):
            _add(f"{main_keyword} {sp}", "세부", 3)
            break

    return results


def _calc_bid_price(price: int | None, competition: str, organic_rank: int | None,
                    kw_type: str, ad_count: int) -> dict:
    """입찰가 추천. 반환: {"bid": int, "min_bid": int, "max_bid": int, "strategy": str}

    산출 기준:
    - 마진율 22% 가정 (도서 리셀러)
    - 전환율 4% 가정 (클릭 25회당 1건 판매)
    - 목표 ROAS 300% 이상
    """
    if not price or price <= 0:
        return {"bid": 100, "min_bid": 70, "max_bid": 150, "strategy": "가격정보 없음 — 최소 입찰"}

    # 수익 기반 최대 CPC
    margin = price * 0.22
    max_cpc = margin / 25  # 전환율 4%

    # 경쟁도 보정
    comp_mult = {"LOW": 0.5, "MID": 0.7, "HIGH": 0.9}.get(competition, 0.7)

    # 키워드 유형 보정 (롱테일은 CPC 낮아도 됨)
    type_mult = {"대표": 1.0, "세부": 0.75, "롱테일": 0.55}.get(kw_type, 0.8)

    # 순위 기반 전략
    if organic_rank and organic_rank <= 10:
        strategy = "방어 — 이미 상위, 최소 입찰로 유지"
        rank_mult = 0.5
    elif organic_rank and organic_rank <= 30:
        strategy = "강화 — 상위권 진입 가능, 중간 입찰"
        rank_mult = 0.75
    elif organic_rank and organic_rank <= 60:
        strategy = "공격 — 노출은 되나 중하위, 적극 입찰"
        rank_mult = 0.9
    else:
        strategy = "진입 — 미노출/하위, 적극 입찰 필요"
        rank_mult = 1.0

    # 광고 포화도 보정
    if ad_count > 80:
        ad_mult = 1.15  # 광고 많은 키워드 = 단가 높음
    elif ad_count > 40:
        ad_mult = 1.0
    else:
        ad_mult = 0.85  # 광고 적은 = 싸게 가능

    bid = max_cpc * comp_mult * type_mult * rank_mult * ad_mult
    bid = max(70, round(bid / 10) * 10)  # 최소 70원, 10원 단위

    min_bid = max(70, round(bid * 0.7 / 10) * 10)
    max_bid = max(min_bid + 30, round(bid * 1.4 / 10) * 10)

    # 절대 상한: 마진의 50% 초과 금지
    absolute_max = max(70, round(margin * 0.5 / 10) * 10)
    bid = min(bid, absolute_max)
    max_bid = min(max_bid, absolute_max)

    return {"bid": int(bid), "min_bid": int(min_bid), "max_bid": int(max_bid), "strategy": strategy}


# ═══════════════════════════════════════
# 리포트 생성
# ═══════════════════════════════════════

def _load_api_products(account_code: str, base_dir: str = ".") -> list[dict]:
    """API JSON 파일에서 상품 로드. DB 호환 dict 리스트 반환."""
    # 계정코드에서 파일명 유추 (007-ez → 007ez)
    code_clean = account_code.replace("-", "")
    json_path = os.path.join(base_dir, "data", f"{code_clean}_api_products.json")
    if not os.path.exists(json_path):
        print(f"\n  API 데이터 파일 없음: {json_path}")
        print(f"  먼저 API로 상품을 가져오세요.")
        return []

    with open(json_path, "r", encoding="utf-8") as f:
        api_products = json.load(f)

    # DB 호환 형식으로 변환
    products = []
    for ap in api_products:
        products.append({
            "id": ap.get("sellerProductId"),
            "seller_product_id": str(ap.get("sellerProductId", "")),
            "product_name": ap.get("sellerProductName", ""),
            "category": str(ap.get("displayCategoryCode", "")),
            "brand": ap.get("brand", ""),
            "barcode": "",
            "wing_product_id": str(ap.get("productId", "")),
        })
    return products


def generate_ad_report(account_code: str, config: AnalysisConfig = None,
                       output: str = "", source: str = "db") -> str:
    """광고 추천 리포트 생성. 반환값: 저장된 파일 경로.

    source: "db" = DB inventory_products, "api" = API JSON 파일
    """
    config = config or AnalysisConfig()
    db = CoupangDB(config)
    conn = db.conn

    # ── 상품 로드 ──
    if source == "api":
        products = _load_api_products(account_code, config.base_dir)
        if not products:
            db.close()
            return ""
        account_name = account_code.upper()
    else:
        account = db.get_account_by_code(account_code)
        if not account:
            print(f"\n  계정을 찾을 수 없습니다: {account_code}")
            db.close()
            return ""
        products = conn.execute("""
            SELECT id, seller_product_id, product_name, category, brand, barcode, wing_product_id
            FROM inventory_products WHERE account_id = ? AND status = '판매중'
        """, (account.id,)).fetchall()
        account_name = account.account_name

    if not products:
        print(f"\n  '{account_code}' 판매중 상품이 없습니다.")
        db.close()
        return ""

    keywords = [r["keyword"] for r in conn.execute("SELECT keyword FROM keywords").fetchall()]
    if not keywords:
        print("\n  수집된 키워드가 없습니다. 먼저 collect를 실행하세요.")
        db.close()
        return ""

    src_label = "API" if source == "api" else "DB"
    print(f"\n  계정: {account_code} ({account_name}) [{src_label}]")
    print(f"  판매중 상품: {len(products)}개 / 수집 키워드: {len(keywords)}개")

    # ── 키워드별 광고 수 사전 계산 ──
    kw_ad_counts = {}
    for kw in keywords:
        df = db.get_analysis_dataframe(kw)
        if not df.empty:
            kw_ad_counts[kw] = int((df["ad_type"] == "AD").sum())

    # ── 광고 후보 데이터 수집 ──
    ad_candidates = []
    bid_rows = []  # 입찰 키워드 시트용

    for kw in keywords:
        df = db.get_analysis_dataframe(kw)
        if df.empty:
            continue
        organic = df[df["ad_type"] == "자연검색"]
        ads = df[df["ad_type"] == "AD"]
        matched = [p for p in products if kw.lower() in (p["product_name"] or "").lower()]
        if not matched:
            continue

        top10 = organic.nsmallest(10, "organic_rank") if len(organic) >= 10 else organic
        top10_avg_reviews = top10["review_count"].mean()
        top10_avg_price = top10["sale_price"].mean()
        ad_count = len(ads)

        for p in matched:
            pname = (p["product_name"] or "").lower()
            pwords = set(re.findall(r"[가-힣a-zA-Z0-9]+", pname))

            best_match = None
            best_overlap = 0
            for _, row in organic.iterrows():
                sname = (row.get("product_name") or "").lower()
                swords = set(re.findall(r"[가-힣a-zA-Z0-9]+", sname))
                overlap = len(pwords & swords) / max(len(pwords), 1)
                if overlap > best_overlap and overlap > 0.5:
                    best_overlap = overlap
                    best_match = row

            ad_match = None
            for _, row in ads.iterrows():
                sname = (row.get("product_name") or "").lower()
                swords = set(re.findall(r"[가-힣a-zA-Z0-9]+", sname))
                overlap = len(pwords & swords) / max(len(pwords), 1)
                if overlap > 0.5:
                    ad_match = row
                    break

            in_organic = best_match is not None
            org_rank = int(best_match["organic_rank"]) if in_organic and pd.notna(best_match.get("organic_rank")) else None
            s_price = int(best_match["sale_price"]) if in_organic and pd.notna(best_match.get("sale_price")) else None
            s_reviews = int(best_match["review_count"]) if in_organic and pd.notna(best_match.get("review_count")) else None
            s_rating = float(best_match["rating"]) if in_organic and pd.notna(best_match.get("rating")) else None
            already_ad = ad_match is not None
            competition = "LOW" if top10_avg_reviews < 50 else "MID" if top10_avg_reviews < 300 else "HIGH"

            score = 0
            if in_organic:
                score += 30
                if org_rank and org_rank <= 30:
                    score += 20
                elif org_rank and org_rank <= 60:
                    score += 10
            if competition == "LOW":
                score += 25
            elif competition == "MID":
                score += 10
            if s_reviews and s_reviews > 10:
                score += 15
            elif s_reviews and s_reviews > 0:
                score += 5
            if ad_count > 20:
                score += 10

            # ── 키워드 추출 + 입찰가 산출 ──
            extracted_kws = _extract_ad_keywords(p["product_name"], kw)
            bid_info = _calc_bid_price(s_price, competition, org_rank, "대표", ad_count)

            # 키워드별 입찰가 계산
            product_bid_rows = []
            for ekw in extracted_kws:
                kw_bid = _calc_bid_price(s_price, competition, org_rank, ekw["type"], ad_count)
                product_bid_rows.append({
                    "seller_product_id": p["seller_product_id"],
                    "product_name": p["product_name"],
                    "search_price": s_price,
                    "organic_rank": org_rank,
                    "ad_keyword": ekw["keyword"],
                    "kw_type": ekw["type"],
                    "kw_priority": ekw["priority"],
                    "bid": kw_bid["bid"],
                    "min_bid": kw_bid["min_bid"],
                    "max_bid": kw_bid["max_bid"],
                    "strategy": kw_bid["strategy"],
                    "competition": competition,
                    "ad_score": score,
                    "wing_product_id": p["wing_product_id"] or "",
                })
            bid_rows.extend(product_bid_rows)

            # 대표 키워드 목록 (최대 5개)
            top_kws = [ek["keyword"] for ek in extracted_kws[:5]]

            ad_candidates.append({
                "inv_id": p["id"],
                "seller_product_id": p["seller_product_id"],
                "product_name": p["product_name"],
                "brand": p["brand"] or "",
                "category": p["category"] or "",
                "wing_product_id": p["wing_product_id"] or "",
                "keyword": kw,
                "in_organic": "노출" if in_organic else "미노출",
                "organic_rank": org_rank,
                "search_price": s_price,
                "search_reviews": s_reviews,
                "search_rating": s_rating,
                "already_ad": "광고중" if already_ad else "",
                "competition": competition,
                "top10_avg_reviews": round(top10_avg_reviews),
                "top10_avg_price": round(top10_avg_price),
                "ad_score": score,
                "bid_keywords": " / ".join(top_kws),
                "recommended_bid": bid_info["bid"],
                "bid_range": f"{bid_info['min_bid']}~{bid_info['max_bid']}",
                "bid_strategy": bid_info["strategy"],
            })

    db.close()
    ad_candidates.sort(key=lambda x: (-x["ad_score"], x["organic_rank"] or 999))
    bid_rows.sort(key=lambda x: (-x["ad_score"], x["kw_priority"]))

    # ══════════════════════════════════
    # Excel 생성
    # ══════════════════════════════════
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill2 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_fill3 = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    low_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    mid_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    high_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    top_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    score_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    longtail_fill = PatternFill(start_color="E8E0F0", end_color="E8E0F0", fill_type="solid")
    detail_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    body_font = Font(name="맑은 고딕", size=10)

    def style_header(ws, row, fill=header_fill):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def style_body(ws, start_row, end_row):
        for r in range(start_row, end_row + 1):
            for cell in ws[r]:
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    def auto_width(ws, max_w=40):
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val) * 1.2 + 2, max_w))
            ws.column_dimensions[col_letter].width = max(max_len, 8)

    wb = Workbook()

    # ═══ Sheet 1: 요약 ═══
    ws1 = wb.active
    ws1.title = "광고추천_요약"

    kw_summary = {}
    for c in ad_candidates:
        kw = c["keyword"]
        if kw not in kw_summary:
            kw_summary[kw] = {
                "total": 0, "in_search": 0, "recommend": 0, "top30": 0,
                "competition": c["competition"], "top10_reviews": c["top10_avg_reviews"],
                "top10_price": c["top10_avg_price"],
                "avg_bid": [],
            }
        kw_summary[kw]["total"] += 1
        kw_summary[kw]["avg_bid"].append(c["recommended_bid"])
        if c["in_organic"] == "노출":
            kw_summary[kw]["in_search"] += 1
        if c["ad_score"] >= 50:
            kw_summary[kw]["recommend"] += 1
        if c["organic_rank"] and c["organic_rank"] <= 30:
            kw_summary[kw]["top30"] += 1

    comp_order = {"LOW": 0, "MID": 1, "HIGH": 2}
    sorted_kws = sorted(kw_summary.items(), key=lambda x: comp_order.get(x[1]["competition"], 9))

    ws1.append([f"{account_code.upper()} 광고 추천 분석"])
    ws1["A1"].font = Font(name="맑은 고딕", bold=True, size=14)
    ws1.merge_cells("A1:J1")
    ws1.append([])
    ws1.append(["키워드", "경쟁도", "Top10리뷰", "Top10가격", "매칭", "노출",
                "추천", "Top30", "평균입찰가", "광고수"])
    style_header(ws1, 3)

    for kw, info in sorted_kws:
        avg_bid = round(np.mean(info["avg_bid"])) if info["avg_bid"] else 0
        ws1.append([kw, info["competition"], info["top10_reviews"],
                    info["top10_price"], info["total"], info["in_search"],
                    info["recommend"], info["top30"], avg_bid,
                    kw_ad_counts.get(kw, 0)])

    end_row = 3 + len(sorted_kws)
    style_body(ws1, 4, end_row)
    for r in range(4, end_row + 1):
        comp_val = ws1.cell(r, 2).value
        fill = low_fill if comp_val == "LOW" else mid_fill if comp_val == "MID" else high_fill
        for c in range(1, 11):
            ws1.cell(r, c).fill = fill
        ws1.cell(r, 4).number_format = "#,##0"
        ws1.cell(r, 9).number_format = "#,##0"

    ws1.append([])
    ws1.append(["경쟁도: LOW=Top10 리뷰 50개 미만 / MID=300 미만 / HIGH=300+"])
    ws1.append(["입찰가 산출: 마진22% × 전환율4% 기준, 경쟁도/순위/키워드유형별 보정"])
    ws1.append(["키워드유형: 대표(브랜드명) > 세부(브랜드+과목) > 롱테일(브랜드+과목+학년)"])
    auto_width(ws1)

    # ═══ Sheet 2: 전체 광고후보 (키워드+입찰가 포함) ═══
    ws2 = wb.create_sheet("전체_광고후보")
    headers = ["점수", "키워드", "상품명", "브랜드", "경쟁도", "노출",
               "자연순위", "가격", "리뷰", "경쟁사광고",
               "입찰키워드", "추천입찰가", "입찰범위", "입찰전략",
               "셀러상품ID", "WING ID"]
    ws2.append(headers)
    style_header(ws2, 1)

    for c in ad_candidates:
        ws2.append([
            c["ad_score"], c["keyword"], c["product_name"], c["brand"],
            c["competition"], c["in_organic"], c["organic_rank"],
            c["search_price"], c["search_reviews"], c["already_ad"],
            c["bid_keywords"], c["recommended_bid"], c["bid_range"],
            c["bid_strategy"],
            c["seller_product_id"], c["wing_product_id"],
        ])

    end_row2 = 1 + len(ad_candidates)
    style_body(ws2, 2, end_row2)
    for r in range(2, end_row2 + 1):
        score = ws2.cell(r, 1).value or 0
        if score >= 90:
            ws2.cell(r, 1).fill = score_fill
            ws2.cell(r, 1).font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        elif score >= 70:
            ws2.cell(r, 1).fill = low_fill
        comp_val = ws2.cell(r, 5).value
        fill = low_fill if comp_val == "LOW" else mid_fill if comp_val == "MID" else high_fill
        ws2.cell(r, 5).fill = fill
        rank_val = ws2.cell(r, 7).value
        if rank_val and rank_val <= 10:
            ws2.cell(r, 7).fill = top_fill
            ws2.cell(r, 7).font = Font(name="맑은 고딕", size=10, bold=True)
        if ws2.cell(r, 8).value:
            ws2.cell(r, 8).number_format = "#,##0"
        ws2.cell(r, 12).number_format = "#,##0"
        ws2.cell(r, 7).alignment = Alignment(horizontal="center")
        ws2.cell(r, 1).alignment = Alignment(horizontal="center")
        ws2.cell(r, 12).alignment = Alignment(horizontal="center")

    auto_width(ws2, 50)
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["K"].width = 40
    ws2.column_dimensions["N"].width = 30
    ws2.auto_filter.ref = f"A1:P{end_row2}"

    # ═══ Sheet 3: 입찰키워드 상세 (WING 등록용) ═══
    ws3 = wb.create_sheet("입찰키워드_상세")
    headers3 = ["상품명", "셀러상품ID", "입찰키워드", "키워드유형",
                "추천입찰가", "최소입찰가", "최대입찰가",
                "경쟁도", "자연순위", "가격", "입찰전략", "WING ID"]
    ws3.append(headers3)
    style_header(ws3, 1, purple_fill)

    for br in bid_rows:
        ws3.append([
            br["product_name"], br["seller_product_id"],
            br["ad_keyword"], br["kw_type"],
            br["bid"], br["min_bid"], br["max_bid"],
            br["competition"], br["organic_rank"], br["search_price"],
            br["strategy"], br["wing_product_id"],
        ])

    end_row3 = 1 + len(bid_rows)
    style_body(ws3, 2, end_row3)
    for r in range(2, end_row3 + 1):
        # 키워드유형별 색상
        kw_type_val = ws3.cell(r, 4).value
        if kw_type_val == "롱테일":
            ws3.cell(r, 4).fill = longtail_fill
        elif kw_type_val == "세부":
            ws3.cell(r, 4).fill = detail_fill
        elif kw_type_val == "대표":
            ws3.cell(r, 4).fill = top_fill
        # 입찰가 포맷
        for col in [5, 6, 7]:
            ws3.cell(r, col).number_format = "#,##0"
            ws3.cell(r, col).alignment = Alignment(horizontal="center")
        if ws3.cell(r, 10).value:
            ws3.cell(r, 10).number_format = "#,##0"
        ws3.cell(r, 9).alignment = Alignment(horizontal="center")
        # 경쟁도 색상
        comp_val = ws3.cell(r, 8).value
        fill = low_fill if comp_val == "LOW" else mid_fill if comp_val == "MID" else high_fill
        ws3.cell(r, 8).fill = fill

    auto_width(ws3, 50)
    ws3.column_dimensions["A"].width = 50
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["K"].width = 30
    ws3.auto_filter.ref = f"A1:L{end_row3}"

    # ═══ Sheet 4: 키워드별 TOP 추천 ═══
    ws4 = wb.create_sheet("키워드별_TOP추천")
    row_num = 1

    for kw, info in sorted_kws:
        kw_items = [c for c in ad_candidates if c["keyword"] == kw]
        top_items = sorted(kw_items, key=lambda x: (-x["ad_score"], x.get("organic_rank") or 999))[:15]

        title = (f'[{kw}] 경쟁: {info["competition"]} | Top10 리뷰 {info["top10_reviews"]}개 '
                 f'| 매칭 {info["total"]}개')
        ws4.cell(row_num, 1, title)
        ws4.cell(row_num, 1).font = Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
        ws4.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=12)
        row_num += 1

        sub_headers = ["점수", "상품명", "브랜드", "순위", "가격", "리뷰",
                       "입찰키워드", "추천입찰가", "입찰범위", "전략", "경쟁사광고", "셀러상품ID"]
        hfill = header_fill2 if info["competition"] == "LOW" else header_fill3 if info["competition"] == "MID" else header_fill
        for ci, h in enumerate(sub_headers, 1):
            cell = ws4.cell(row_num, ci, h)
            cell.font = header_font
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row_num += 1

        for item in top_items:
            ws4.cell(row_num, 1, item["ad_score"])
            ws4.cell(row_num, 2, item["product_name"])
            ws4.cell(row_num, 3, item["brand"])
            ws4.cell(row_num, 4, item["organic_rank"])
            ws4.cell(row_num, 5, item["search_price"])
            ws4.cell(row_num, 6, item["search_reviews"])
            ws4.cell(row_num, 7, item["bid_keywords"])
            ws4.cell(row_num, 8, item["recommended_bid"])
            ws4.cell(row_num, 9, item["bid_range"])
            ws4.cell(row_num, 10, item["bid_strategy"])
            ws4.cell(row_num, 11, item["already_ad"])
            ws4.cell(row_num, 12, item["seller_product_id"])
            for ci in range(1, 13):
                ws4.cell(row_num, ci).font = body_font
                ws4.cell(row_num, ci).border = thin_border
            ws4.cell(row_num, 1).alignment = Alignment(horizontal="center")
            ws4.cell(row_num, 4).alignment = Alignment(horizontal="center")
            ws4.cell(row_num, 8).alignment = Alignment(horizontal="center")
            if item["search_price"]:
                ws4.cell(row_num, 5).number_format = "#,##0"
            ws4.cell(row_num, 8).number_format = "#,##0"
            if item["organic_rank"] and item["organic_rank"] <= 10:
                ws4.cell(row_num, 4).fill = top_fill
                ws4.cell(row_num, 4).font = Font(name="맑은 고딕", size=10, bold=True)
            row_num += 1

        row_num += 1

    auto_width(ws4, 50)
    ws4.column_dimensions["B"].width = 50
    ws4.column_dimensions["G"].width = 35
    ws4.column_dimensions["J"].width = 28

    # ═══ Sheet 5: 즉시 광고 TOP 50 ═══
    ws5 = wb.create_sheet("즉시광고_TOP50")
    headers5 = ["순번", "점수", "키워드", "상품명", "경쟁도",
                "순위", "가격", "리뷰",
                "입찰키워드", "추천입찰가", "입찰범위", "전략",
                "셀러상품ID", "WING ID"]
    ws5.append(headers5)
    style_header(ws5, 1, red_fill)

    seen_names = set()
    top50 = []
    for c in ad_candidates:
        short_name = c["product_name"][:40]
        if short_name not in seen_names and c["ad_score"] >= 65:
            seen_names.add(short_name)
            top50.append(c)
            if len(top50) >= 50:
                break

    for i, c in enumerate(top50, 1):
        ws5.append([
            i, c["ad_score"], c["keyword"], c["product_name"], c["competition"],
            c["organic_rank"], c["search_price"], c["search_reviews"],
            c["bid_keywords"], c["recommended_bid"], c["bid_range"], c["bid_strategy"],
            c["seller_product_id"], c["wing_product_id"],
        ])

    end_row5 = 1 + len(top50)
    style_body(ws5, 2, end_row5)
    for r in range(2, end_row5 + 1):
        ws5.cell(r, 1).alignment = Alignment(horizontal="center")
        ws5.cell(r, 2).alignment = Alignment(horizontal="center")
        ws5.cell(r, 10).alignment = Alignment(horizontal="center")
        if ws5.cell(r, 7).value:
            ws5.cell(r, 7).number_format = "#,##0"
        ws5.cell(r, 10).number_format = "#,##0"
        rank_val = ws5.cell(r, 6).value
        if rank_val and rank_val <= 10:
            ws5.cell(r, 6).fill = top_fill
            ws5.cell(r, 6).font = Font(name="맑은 고딕", size=10, bold=True)

    auto_width(ws5, 50)
    ws5.column_dimensions["D"].width = 50
    ws5.column_dimensions["I"].width = 35
    ws5.column_dimensions["L"].width = 28

    # ═══ Sheet 6: WING 키워드등록용 (상품별 시트) ═══
    # 상품별로 키워드+입찰가를 WING 양식대로 정리
    # 즉시광고 TOP50 상품 기준으로 생성
    wing_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    wing_header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    wing_kw_font = Font(name="맑은 고딕", size=10)
    wing_help_font = Font(name="맑은 고딕", size=9, color="888888")
    wing_sep_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    ws6 = wb.create_sheet("WING키워드등록")

    # bid_rows를 상품별로 그룹핑 (즉시광고 TOP50에 포함된 것만)
    top50_pids = set(c["seller_product_id"] for c in top50)
    wing_row = 1

    # 전체 요약 헤더
    ws6.cell(wing_row, 1, "WING 수동 키워드 등록용").font = Font(
        name="맑은 고딕", bold=True, size=13, color="1F4E79")
    ws6.merge_cells(start_row=wing_row, start_column=1, end_row=wing_row, end_column=4)
    wing_row += 1
    ws6.cell(wing_row, 1, "아래 키워드/입찰가를 WING 광고 등록 시 복사하여 사용").font = wing_help_font
    wing_row += 2

    for ci, cand in enumerate(top50, 1):
        pid = cand["seller_product_id"]
        product_bids = [br for br in bid_rows if br["seller_product_id"] == pid
                        and br["ad_keyword"] == br["ad_keyword"]]  # 해당 상품의 키워드들
        # 같은 상품이 여러 키워드로 매칭될 수 있으므로 중복 제거
        seen_kws = set()
        unique_bids = []
        for br in bid_rows:
            if br["seller_product_id"] == pid and br["ad_keyword"] not in seen_kws:
                seen_kws.add(br["ad_keyword"])
                unique_bids.append(br)

        if not unique_bids:
            continue

        # 상품명 헤더 (구분선)
        short_name = (cand["product_name"] or "")[:60]
        ws6.cell(wing_row, 1, f"#{ci} {short_name}").font = Font(
            name="맑은 고딕", bold=True, size=10, color="1F4E79")
        ws6.cell(wing_row, 3, f"셀러상품ID: {pid}").font = wing_help_font
        ws6.merge_cells(start_row=wing_row, start_column=1, end_row=wing_row, end_column=2)
        for col in range(1, 5):
            ws6.cell(wing_row, col).fill = wing_sep_fill
            ws6.cell(wing_row, col).border = thin_border
        wing_row += 1

        # 키워드 | 입찰가 헤더 (WING 양식 동일)
        ws6.cell(wing_row, 1, "키워드").font = wing_header_font
        ws6.cell(wing_row, 1).fill = wing_fill
        ws6.cell(wing_row, 1).alignment = Alignment(horizontal="center")
        ws6.cell(wing_row, 1).border = thin_border
        ws6.cell(wing_row, 2, "입찰가").font = wing_header_font
        ws6.cell(wing_row, 2).fill = wing_fill
        ws6.cell(wing_row, 2).alignment = Alignment(horizontal="center")
        ws6.cell(wing_row, 2).border = thin_border
        ws6.cell(wing_row, 4, "도움말:\n좌측 키워드, 입찰가 열을 입력해주세요 ").font = wing_help_font
        wing_row += 1

        # 키워드 행들 (우선순위순)
        unique_bids.sort(key=lambda x: x["kw_priority"])
        for br in unique_bids:
            ws6.cell(wing_row, 1, br["ad_keyword"]).font = wing_kw_font
            ws6.cell(wing_row, 1).border = thin_border
            ws6.cell(wing_row, 2, br["bid"]).font = wing_kw_font
            ws6.cell(wing_row, 2).alignment = Alignment(horizontal="center")
            ws6.cell(wing_row, 2).border = thin_border
            ws6.cell(wing_row, 2).number_format = "#,##0"
            # 키워드 유형별 배경색
            if br["kw_type"] == "대표":
                ws6.cell(wing_row, 1).fill = top_fill
            elif br["kw_type"] == "세부":
                ws6.cell(wing_row, 1).fill = detail_fill
            elif br["kw_type"] == "롱테일":
                ws6.cell(wing_row, 1).fill = longtail_fill
            wing_row += 1

        wing_row += 1  # 상품 간 빈 줄

    ws6.column_dimensions["A"].width = 35
    ws6.column_dimensions["B"].width = 12
    ws6.column_dimensions["C"].width = 25
    ws6.column_dimensions["D"].width = 40

    # ── 저장 ──
    suffix = "_API광고추천" if source == "api" else "_광고추천"
    output_path = output or os.path.join(config.reports_dir, f"{account_code}{suffix}.xlsx")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    total_kws = len(set(br["ad_keyword"] for br in bid_rows))
    avg_bid = round(np.mean([c["recommended_bid"] for c in ad_candidates])) if ad_candidates else 0
    print(f"\n  저장 완료: {output_path}")
    print(f"  전체 후보: {len(ad_candidates)}개 / 입찰키워드: {total_kws}개")
    print(f"  평균 추천입찰가: {avg_bid}원")
    print(f"  즉시 추천 (65점+): {len(top50)}개")
    print(f"  시트: 요약 / 전체후보 / 입찰키워드상세 / 키워드별TOP / 즉시TOP50 / WING키워드등록")
    return output_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="광고 추천 리포트")
    parser.add_argument("--account", "-a", default="007-ez", help="계정 코드")
    parser.add_argument("--output", "-o", default="", help="저장 경로")
    parser.add_argument("--source", "-s", choices=["db", "api"], default="db",
                        help="상품 데이터 소스 (db=DB재고, api=API JSON)")
    a = parser.parse_args()
    generate_ad_report(a.account, output=a.output, source=a.source)

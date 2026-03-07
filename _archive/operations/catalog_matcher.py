"""카탈로그 매칭 반자동화 — 우선순위 정리 + 소량 후보 검색 + 수동 매칭용 Excel"""

import os
import re
import random
import time
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from urllib.parse import quote

from bs4 import BeautifulSoup

from core.config import AnalysisConfig
from core.database import CoupangDB
from core.models import CatalogMatch
from analysis.collector import extract_product_cards, _launch_chrome_debug


# ──────────────────────────────────────────────
# 상품명 정규화 / 유사도
# ──────────────────────────────────────────────

def normalize_product_name(name: str) -> str:
    """상품명 정규화: 괄호/특수문자 제거, 공백 정리"""
    if not name:
        return ""
    name = re.sub(r'[\(\)\[\]【】「」\{\}]', ' ', name)
    name = re.sub(r'[^\w\s가-힣a-zA-Z0-9]', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def extract_search_keywords(name: str, max_tokens: int = 4) -> str:
    """상품명에서 핵심 2~4어절 검색 키워드 추출"""
    normalized = normalize_product_name(name)
    tokens = normalized.split()

    stopwords = {"세트", "개입", "박스", "무료배송", "국내", "수입", "정품",
                 "특가", "할인", "대용량", "미니", "소형", "대형", "중형",
                 "1개", "2개", "3개", "5개", "10개", "묶음", "단품"}
    filtered = [t for t in tokens if t not in stopwords and len(t) > 1]

    if not filtered:
        filtered = tokens

    selected = filtered[:min(max_tokens, len(filtered))]
    if len(selected) < 2 and len(tokens) >= 2:
        selected = tokens[:2]

    return " ".join(selected)


def calculate_name_similarity(name1: str, name2: str) -> float:
    """상품명 유사도 (0~1): SequenceMatcher + Jaccard 한국어 토큰 평균"""
    n1 = normalize_product_name(name1).lower()
    n2 = normalize_product_name(name2).lower()

    if not n1 or not n2:
        return 0.0

    seq_ratio = SequenceMatcher(None, n1, n2).ratio()

    tokens1 = set(n1.split())
    tokens2 = set(n2.split())
    if tokens1 or tokens2:
        jaccard = len(tokens1 & tokens2) / len(tokens1 | tokens2)
    else:
        jaccard = 0.0

    return (seq_ratio + jaccard) / 2.0


def calculate_price_similarity(price1, price2) -> float:
    """가격 유사도 (0~1)"""
    if price1 is None or price2 is None or price1 == 0 or price2 == 0:
        return 0.0
    diff_ratio = abs(price1 - price2) / max(price1, price2)
    return max(0, 1.0 - diff_ratio)


def calculate_match_score(my_product, candidate: dict, config: AnalysisConfig) -> CatalogMatch:
    """매칭 점수 산출 (0~100)"""
    name_sim = calculate_name_similarity(
        my_product.product_name, candidate.get("상품명", "")
    )
    name_score = name_sim * config.catalog_name_weight

    price_sim = calculate_price_similarity(
        my_product.sale_price, candidate.get("판매가")
    )
    price_score = price_sim * config.catalog_price_weight

    cat_score = 0.0
    if my_product.category and candidate.get("카테고리"):
        if my_product.category in candidate["카테고리"] or candidate["카테고리"] in my_product.category:
            cat_score = config.catalog_category_weight

    review_bonus = 0.0
    review_count = candidate.get("리뷰수") or 0
    if isinstance(review_count, str):
        try:
            review_count = int(re.sub(r'[^\d]', '', review_count))
        except ValueError:
            review_count = 0
    if review_count > 0:
        review_bonus = min(review_count / 100, 1.0) * config.catalog_review_weight

    total_score = name_score + price_score + cat_score + review_bonus

    if total_score >= 75:
        confidence = "높음"
    elif total_score >= 55:
        confidence = "보통"
    else:
        confidence = "낮음"

    return CatalogMatch(
        inventory_product_id=my_product.id,
        account_id=my_product.account_id,
        candidate_product_id=str(candidate.get("productId", "")),
        candidate_vendor_item_id=str(candidate.get("vendorItemId", "")),
        candidate_name=candidate.get("상품명", ""),
        candidate_price=candidate.get("판매가") if isinstance(candidate.get("판매가"), int) else None,
        candidate_review_count=review_count if review_count else None,
        candidate_rating=candidate.get("평점") if isinstance(candidate.get("평점"), (int, float)) else None,
        candidate_url=candidate.get("URL", ""),
        candidate_category=candidate.get("카테고리", ""),
        name_score=round(name_score, 1),
        price_score=round(price_score, 1),
        category_score=round(cat_score, 1),
        review_bonus=round(review_bonus, 1),
        total_score=round(total_score, 1),
        confidence=confidence,
    )


# ──────────────────────────────────────────────
# 1단계: prepare — 우선순위 정리 Excel (웹 검색 없음)
# ──────────────────────────────────────────────

def prepare_catalog_worksheet(account_code: str, config: AnalysisConfig = None) -> str:
    """재고 상품을 우선순위별로 정리한 '수동 매칭용 Excel' 생성.

    웹 검색 없이 DB의 재고 데이터만 사용.
    - 브랜드/카테고리별 그룹핑
    - WING 검색용 키워드 자동 추출
    - 가격대별 우선순위 정렬
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    config = config or AnalysisConfig()
    db = CoupangDB(config)

    account = db.get_account_by_code(account_code)
    if not account:
        print(f"\n계정을 찾을 수 없습니다: {account_code}")
        db.close()
        return ""

    products = db.list_inventory(account.id, status="판매중", limit=10000)
    if not products:
        print(f"\n재고 상품이 없습니다. 먼저 inv import를 실행하세요.")
        db.close()
        return ""

    # 이미 매칭된 상품 제외
    already_matched = db.get_matched_inventory_ids(account.id)
    db.close()

    unmatched = [p for p in products if p.id not in already_matched]

    print(f"\n[카탈로그 준비] 계정: {account_code}")
    print(f"  전체 재고: {len(products)}개")
    print(f"  이미 매칭: {len(already_matched)}개")
    print(f"  작업 대상: {len(unmatched)}개")

    if not unmatched:
        print(f"\n  모든 상품이 이미 매칭되었습니다.")
        return ""

    # ── 스타일 ──
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    priority_high = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    priority_mid = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # ── 시트 1: 전체 작업 목록 (우선순위순) ──
    ws_main = wb.active
    ws_main.title = "매칭 작업 목록"
    headers = [
        "우선순위", "셀러상품ID", "상품명", "판매가", "브랜드",
        "카테고리", "바코드", "WING 검색어", "매칭 완료", "메모"
    ]
    ws_main.append(headers)
    style_header(ws_main)

    # 우선순위: 가격 높은 순 (매출 잠재력)
    sorted_products = sorted(unmatched, key=lambda p: p.sale_price or 0, reverse=True)

    for rank, p in enumerate(sorted_products, 1):
        search_kw = extract_search_keywords(p.product_name)
        ws_main.append([
            rank,
            p.seller_product_id,
            p.product_name,
            p.sale_price,
            p.brand,
            p.category,
            p.barcode,
            search_kw,
            "",  # 매칭 완료 체크용 빈칸
            "",  # 메모
        ])

        # 상위 20%는 녹색, 상위 50%는 노란색
        row_idx = ws_main.max_row
        if rank <= len(sorted_products) * 0.2:
            for cell in ws_main[row_idx]:
                cell.fill = priority_high
        elif rank <= len(sorted_products) * 0.5:
            for cell in ws_main[row_idx]:
                cell.fill = priority_mid

    widths = [8, 15, 45, 12, 15, 20, 15, 30, 10, 20]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
        ws_main.column_dimensions[col_letter].width = w

    # ── 시트 2: 브랜드별 그룹 ──
    ws_brand = wb.create_sheet("브랜드별")
    ws_brand.append(["브랜드", "상품 수", "평균 가격", "상품명 예시"])
    style_header(ws_brand)

    brand_groups = defaultdict(list)
    for p in unmatched:
        brand = p.brand or "(브랜드 없음)"
        brand_groups[brand].append(p)

    # 상품 수 많은 순
    for brand, items in sorted(brand_groups.items(), key=lambda x: len(x[1]), reverse=True):
        avg_price = sum(p.sale_price or 0 for p in items) / len(items)
        example = items[0].product_name[:50] if items else ""
        ws_brand.append([brand, len(items), round(avg_price), example])

    ws_brand.column_dimensions["A"].width = 20
    ws_brand.column_dimensions["B"].width = 10
    ws_brand.column_dimensions["C"].width = 12
    ws_brand.column_dimensions["D"].width = 50

    # ── 시트 3: 카테고리별 그룹 ──
    ws_cat = wb.create_sheet("카테고리별")
    ws_cat.append(["카테고리", "상품 수", "평균 가격", "상품명 예시"])
    style_header(ws_cat)

    cat_groups = defaultdict(list)
    for p in unmatched:
        cat = p.category or "(카테고리 없음)"
        cat_groups[cat].append(p)

    for cat, items in sorted(cat_groups.items(), key=lambda x: len(x[1]), reverse=True):
        avg_price = sum(p.sale_price or 0 for p in items) / len(items)
        example = items[0].product_name[:50] if items else ""
        ws_cat.append([cat, len(items), round(avg_price), example])

    ws_cat.column_dimensions["A"].width = 30
    ws_cat.column_dimensions["B"].width = 10
    ws_cat.column_dimensions["C"].width = 12
    ws_cat.column_dimensions["D"].width = 50

    # ── 시트 4: 바코드 있는 상품 (정확 매칭 가능) ──
    ws_barcode = wb.create_sheet("바코드 보유")
    ws_barcode.append(["셀러상품ID", "상품명", "판매가", "바코드", "브랜드", "WING 검색어"])
    style_header(ws_barcode)

    barcode_products = [p for p in unmatched if p.barcode and p.barcode.strip()]
    for p in barcode_products:
        search_kw = extract_search_keywords(p.product_name)
        ws_barcode.append([
            p.seller_product_id, p.product_name, p.sale_price,
            p.barcode, p.brand, search_kw,
        ])

    ws_barcode.column_dimensions["A"].width = 15
    ws_barcode.column_dimensions["B"].width = 45
    ws_barcode.column_dimensions["C"].width = 12
    ws_barcode.column_dimensions["D"].width = 18
    ws_barcode.column_dimensions["E"].width = 15
    ws_barcode.column_dimensions["F"].width = 30

    # ── 시트 5: 요약 ──
    ws_summary = wb.create_sheet("요약")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15
    summary_rows = [
        ["항목", "값"],
        ["계정", f"{account_code} ({account.account_name})"],
        ["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["전체 재고", len(products)],
        ["이미 매칭", len(already_matched)],
        ["작업 대상", len(unmatched)],
        ["바코드 보유", len(barcode_products)],
        ["브랜드 수", len(brand_groups)],
        ["카테고리 수", len(cat_groups)],
        ["", ""],
        ["사용법", ""],
        ["1", "바코드 보유 시트부터 WING에서 바코드로 검색"],
        ["2", "작업 목록의 'WING 검색어'로 쿠팡 검색"],
        ["3", "동일 상품 찾으면 WING에서 카탈로그 매칭 제출"],
        ["4", "매칭 완료 열에 O 표시"],
    ]
    for row_data in summary_rows:
        ws_summary.append(row_data)
    style_header(ws_summary)

    # 저장
    config.ensure_dirs()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"catalog_prepare_{account_code}_{timestamp}.xlsx"
    filepath = os.path.join(config.reports_dir, filename)
    wb.save(filepath)

    print(f"\n  Excel 생성: {filepath}")
    print(f"  작업 대상: {len(unmatched)}개")
    print(f"  바코드 보유: {len(barcode_products)}개 (우선 매칭 권장)")
    print(f"  브랜드 {len(brand_groups)}개 / 카테고리 {len(cat_groups)}개")
    print(f"\n  사용법:")
    print(f"    1. '바코드 보유' 시트 → WING에서 바코드 검색으로 정확 매칭")
    print(f"    2. '매칭 작업 목록' 시트 → 우선순위 높은 것부터 WING 검색어로 수동 매칭")

    return filepath


# ──────────────────────────────────────────────
# 2단계: match — 소량 후보 검색 (하루 50개 권장)
# ──────────────────────────────────────────────

DAILY_SAFE_LIMIT = 50  # CAPTCHA 방지 권장 한도

def _search_coupang_web(page, keyword: str, config: AnalysisConfig) -> list[dict]:
    """쿠팡 웹 검색 1페이지 → 상품 카드 리스트"""
    url = f"https://www.coupang.com/np/search?component=&q={quote(keyword)}&page=1"

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
    except Exception as e:
        print(f"    페이지 로딩 타임아웃: {e}")

    try:
        page.wait_for_selector("li[class*='ProductUnit']", timeout=15000)
    except Exception:
        pass

    for _ in range(4):
        page.mouse.wheel(0, random.randint(300, 700))
        time.sleep(random.uniform(0.3, 0.8))

    time.sleep(random.uniform(0.5, 1.0))
    content = page.content()

    if "Access Denied" in content or "captcha" in content.lower() or "로봇이 아닙니다" in content:
        print("    !! CAPTCHA 감지 — 브라우저에서 수동 해결 후 Enter")
        try:
            input()
            content = page.content()
        except EOFError:
            return []

    soup = BeautifulSoup(content, "html.parser")
    return extract_product_cards(soup)


def batch_match(account_code: str, config: AnalysisConfig = None, limit: int = 50) -> dict:
    """소량 배치 후보 검색. 기본 50개, CAPTCHA 방지 딜레이 5~10초.

    검색 결과는 DB에 저장 → report로 Excel 확인 → 사람이 WING에서 수동 제출.
    """
    from playwright.sync_api import sync_playwright

    config = config or AnalysisConfig()
    db = CoupangDB(config)

    account = db.get_account_by_code(account_code)
    if not account:
        print(f"\n계정을 찾을 수 없습니다: {account_code}")
        db.close()
        return {}

    all_inventory = db.list_inventory(account.id, status="판매중", limit=10000)
    if not all_inventory:
        print(f"\n재고 상품이 없습니다. 먼저 inv import를 실행하세요.")
        db.close()
        return {}

    already_matched = db.get_matched_inventory_ids(account.id)
    to_process = [p for p in all_inventory if p.id not in already_matched]

    # 가격 높은 순으로 우선 처리
    to_process.sort(key=lambda p: p.sale_price or 0, reverse=True)

    actual_limit = min(limit, len(to_process))
    if actual_limit == 0:
        print(f"\n매칭할 상품이 없습니다. (이미 {len(already_matched)}개 매칭 완료)")
        db.close()
        return db.get_catalog_match_summary(account.id)

    to_process = to_process[:actual_limit]

    if actual_limit > DAILY_SAFE_LIMIT:
        print(f"\n  주의: {actual_limit}개는 CAPTCHA 위험이 있습니다.")
        print(f"  권장: --limit {DAILY_SAFE_LIMIT} (하루 {DAILY_SAFE_LIMIT}개씩)")

    print(f"\n[후보 검색] 계정: {account_code}")
    print(f"  전체 재고: {len(all_inventory)}개")
    print(f"  이미 검색: {len(already_matched)}개")
    print(f"  이번 검색: {actual_limit}개")
    print(f"  딜레이: {config.delay_min}~{config.delay_max}초")
    print(f"  예상 소요: {actual_limit * (config.delay_min + config.delay_max) / 2 / 60:.0f}분")
    print(f"{'='*60}")

    cdp_url = f"http://127.0.0.1:{config.chrome_debug_port}"
    processed = 0
    matched = 0

    with sync_playwright() as pw:
        browser = None
        try:
            browser = pw.chromium.connect_over_cdp(cdp_url, timeout=5000)
            print("  Chrome 연결됨")
        except Exception:
            print("  Chrome 디버깅 모드 연결 시도...")
            _launch_chrome_debug(config.chrome_debug_port)
            time.sleep(2)
            try:
                browser = pw.chromium.connect_over_cdp(cdp_url, timeout=10000)
                print("  Chrome 연결 성공")
            except Exception as e:
                print(f"  Chrome 연결 실패: {e}")
                print(f'  chrome.exe --remote-debugging-port={config.chrome_debug_port}')
                db.close()
                return {}

        context = browser.contexts[0]
        page = context.new_page()

        db.save_matching_progress(account.id, actual_limit, 0, 0)

        try:
            for i, product in enumerate(to_process, 1):
                search_kw = extract_search_keywords(product.product_name)
                print(f"\n  [{i}/{actual_limit}] {product.product_name[:50]}")
                print(f"    검색: '{search_kw}'")

                candidates = _search_coupang_web(page, search_kw, config)

                if not candidates:
                    print(f"    검색 결과 없음")
                    processed += 1
                    if processed % 10 == 0:
                        db.save_matching_progress(account.id, actual_limit, processed, matched)
                    delay = random.uniform(config.delay_min, config.delay_max)
                    time.sleep(delay)
                    continue

                scored = []
                for card in candidates[:20]:
                    match = calculate_match_score(product, card, config)
                    scored.append(match)

                scored.sort(key=lambda m: m.total_score, reverse=True)
                top_matches = scored[:5]

                for rank, m in enumerate(top_matches, 1):
                    m.rank = rank
                    db.insert_catalog_match(m)

                if top_matches and top_matches[0].total_score > 0:
                    matched += 1
                    best = top_matches[0]
                    print(f"    최고 후보: {best.candidate_name[:40]} "
                          f"({best.total_score}점, {best.confidence})")
                else:
                    print(f"    적합한 후보 없음")

                processed += 1

                if processed % 10 == 0:
                    db.save_matching_progress(account.id, actual_limit, processed, matched)
                    print(f"\n    -- 진행: {processed}/{actual_limit} (후보 발견: {matched}) --")

                delay = random.uniform(config.delay_min, config.delay_max)
                time.sleep(delay)

        except KeyboardInterrupt:
            print(f"\n\n  중단됨 (Ctrl+C). 진행 저장 중...")
            db.save_matching_progress(account.id, actual_limit, processed, matched, status="중단")
        finally:
            page.close()
            browser.close()

    db.save_matching_progress(account.id, actual_limit, processed, matched, status="완료")
    summary = db.get_catalog_match_summary(account.id)
    db.close()

    print(f"\n{'='*60}")
    print(f"  후보 검색 완료")
    print(f"  처리: {processed}/{actual_limit}")
    print(f"  후보 발견: {matched}개")
    print(f"  높은 유사도: {summary.get('high_confidence', 0)}개")
    print(f"  평균 점수: {summary.get('avg_score', 0)}점")
    print(f"{'='*60}")
    print(f"\n  다음 단계:")
    print(f"    python main.py catalog report -a {account_code}  # Excel 확인")
    print(f"    → Excel 보고 WING에서 수동 카탈로그 매칭")

    return summary


# ──────────────────────────────────────────────
# 결과 확인
# ──────────────────────────────────────────────

def review_matches(account_code: str, config: AnalysisConfig = None):
    """매칭 후보 결과 콘솔 출력"""
    config = config or AnalysisConfig()
    db = CoupangDB(config)

    account = db.get_account_by_code(account_code)
    if not account:
        print(f"\n계정을 찾을 수 없습니다: {account_code}")
        db.close()
        return

    summary = db.get_catalog_match_summary(account.id)
    matches = db.get_best_matches(account.id)

    print(f"\n{'='*70}")
    print(f"  카탈로그 매칭 후보 현황: {account_code}")
    print(f"{'='*70}")
    print(f"  전체 재고:     {summary['total_inventory']}개")
    print(f"  후보 검색 완료: {summary['matched']}개")
    print(f"  미검색:        {summary['unmatched']}개")
    print(f"  높은 유사도:   {summary['high_confidence']}개 (75점+)")
    print(f"  보통 유사도:   {summary['medium_confidence']}개 (55-74점)")
    print(f"  평균 점수:     {summary['avg_score']}점")
    print(f"{'='*70}")

    if not matches:
        print("\n  검색된 후보가 없습니다.")
        print(f"  먼저 실행: python main.py catalog match -a {account_code}")
        db.close()
        return

    print(f"\n  {'점수':>5} {'유사도':5} {'내 상품':<30} {'후보 상품':<30} {'리뷰':>5}")
    print(f"  {'-'*80}")

    for m in matches[:20]:
        my_name = (m["my_product_name"] or "")[:28]
        cand_name = (m["candidate_name"] or "")[:28]
        review = m["candidate_review_count"] or 0
        print(f"  {m['total_score']:5.1f} {m['confidence']:5} "
              f"{my_name:<30} {cand_name:<30} {review:>5}")

    if len(matches) > 20:
        print(f"\n  ... 외 {len(matches) - 20}개")

    print(f"\n  Excel 상세: python main.py catalog report -a {account_code}")
    db.close()


# ──────────────────────────────────────────────
# Excel 리포트 (수동 매칭 액션용)
# ──────────────────────────────────────────────

def generate_catalog_report(account_code: str, config: AnalysisConfig = None) -> str:
    """후보 검색 결과 + 수동 매칭 체크리스트 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    config = config or AnalysisConfig()
    db = CoupangDB(config)

    account = db.get_account_by_code(account_code)
    if not account:
        print(f"\n계정을 찾을 수 없습니다: {account_code}")
        db.close()
        return ""

    summary = db.get_catalog_match_summary(account.id)
    all_matches = db.get_best_matches(account.id)
    high_matches = db.get_best_matches(account.id, min_score=75)

    if not all_matches:
        print("\n검색된 후보가 없습니다.")
        print(f"먼저 실행: python main.py catalog match -a {account_code}")
        db.close()
        return ""

    wb = openpyxl.Workbook()

    # 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # ── 시트 1: 수동 매칭 체크리스트 (핵심 시트) ──
    ws_action = wb.active
    ws_action.title = "매칭 체크리스트"
    action_headers = [
        "셀러상품ID", "내 상품명", "내 가격",
        "추천 후보", "후보 가격", "후보 리뷰수", "유사도 점수",
        "후보 URL", "WING 검색어",
        "매칭 완료", "메모"
    ]
    ws_action.append(action_headers)
    style_header(ws_action)

    for m in all_matches:
        search_kw = extract_search_keywords(m["my_product_name"])
        ws_action.append([
            m["seller_product_id"],
            m["my_product_name"],
            m["my_price"],
            m["candidate_name"],
            m["candidate_price"],
            m["candidate_review_count"],
            m["total_score"],
            m["candidate_url"],
            search_kw,
            "",  # 매칭 완료 체크
            "",  # 메모
        ])

        row_idx = ws_action.max_row
        if m["confidence"] == "높음":
            for cell in ws_action[row_idx]:
                cell.fill = high_fill
        elif m["confidence"] == "보통":
            for cell in ws_action[row_idx]:
                cell.fill = medium_fill

    action_widths = [15, 35, 10, 35, 10, 10, 8, 50, 25, 10, 20]
    for i, w in enumerate(action_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
        ws_action.column_dimensions[col_letter].width = w

    # ── 시트 2: 높은 유사도 (75점+, 우선 매칭) ──
    ws_high = wb.create_sheet("우선 매칭 (75+)")
    ws_high.append(action_headers)
    style_header(ws_high)

    for m in high_matches:
        search_kw = extract_search_keywords(m["my_product_name"])
        ws_high.append([
            m["seller_product_id"], m["my_product_name"], m["my_price"],
            m["candidate_name"], m["candidate_price"], m["candidate_review_count"],
            m["total_score"], m["candidate_url"], search_kw, "", "",
        ])
        row_idx = ws_high.max_row
        for cell in ws_high[row_idx]:
            cell.fill = high_fill

    for i, w in enumerate(action_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
        ws_high.column_dimensions[col_letter].width = w

    # ── 시트 3: 미검색 상품 ──
    ws_unmatched = wb.create_sheet("미검색 상품")
    ws_unmatched.append(["셀러상품ID", "상품명", "가격", "브랜드", "바코드", "WING 검색어"])
    style_header(ws_unmatched)

    matched_inv_ids = {m["inventory_product_id"] for m in all_matches}
    db2 = CoupangDB(config)
    all_inv = db2.list_inventory(account.id, limit=10000)
    db2.close()

    for p in all_inv:
        if p.id not in matched_inv_ids:
            search_kw = extract_search_keywords(p.product_name)
            ws_unmatched.append([
                p.seller_product_id, p.product_name, p.sale_price,
                p.brand, p.barcode, search_kw,
            ])

    ws_unmatched.column_dimensions["A"].width = 15
    ws_unmatched.column_dimensions["B"].width = 40
    ws_unmatched.column_dimensions["C"].width = 10
    ws_unmatched.column_dimensions["D"].width = 15
    ws_unmatched.column_dimensions["E"].width = 18
    ws_unmatched.column_dimensions["F"].width = 30

    # ── 시트 4: 요약 + 통계 ──
    ws_summary = wb.create_sheet("요약")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    summary_data = [
        ["항목", "값"],
        ["계정", f"{account_code} ({account.account_name})"],
        ["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["전체 재고", summary["total_inventory"]],
        ["후보 검색 완료", summary["matched"]],
        ["미검색", summary["unmatched"]],
        ["높은 유사도 (75+)", summary["high_confidence"]],
        ["보통 유사도 (55-74)", summary["medium_confidence"]],
        ["평균 점수", summary["avg_score"]],
        ["", ""],
        ["작업 순서", ""],
        ["1", "'우선 매칭' 시트의 녹색 항목부터 처리"],
        ["2", "후보 URL 클릭 → 동일 상품인지 확인"],
        ["3", "동일하면 WING에서 카탈로그 매칭 제출"],
        ["4", "'매칭 완료' 열에 O 표시"],
    ]
    for row_data in summary_data:
        ws_summary.append(row_data)
    style_header(ws_summary)

    # 점수 분포 차트
    ws_summary.append([])
    ws_summary.append(["점수 구간", "상품 수"])
    score_bins = {"90-100": 0, "80-89": 0, "70-79": 0, "60-69": 0,
                  "50-59": 0, "40-49": 0, "30-39": 0, "0-29": 0}
    for m in all_matches:
        s = m["total_score"]
        if s >= 90: score_bins["90-100"] += 1
        elif s >= 80: score_bins["80-89"] += 1
        elif s >= 70: score_bins["70-79"] += 1
        elif s >= 60: score_bins["60-69"] += 1
        elif s >= 50: score_bins["50-59"] += 1
        elif s >= 40: score_bins["40-49"] += 1
        elif s >= 30: score_bins["30-39"] += 1
        else: score_bins["0-29"] += 1

    chart_start = ws_summary.max_row
    for label, count in score_bins.items():
        ws_summary.append([label, count])

    chart = BarChart()
    chart.title = "유사도 점수 분포"
    chart.y_axis.title = "상품 수"
    chart.x_axis.title = "점수 구간"
    chart.style = 10
    data = Reference(ws_summary, min_col=2, min_row=chart_start, max_row=chart_start + len(score_bins))
    cats = Reference(ws_summary, min_col=1, min_row=chart_start + 1, max_row=chart_start + len(score_bins))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws_summary.add_chart(chart, "D17")

    # 저장
    config.ensure_dirs()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"catalog_report_{account_code}_{timestamp}.xlsx"
    filepath = os.path.join(config.reports_dir, filename)
    wb.save(filepath)

    print(f"\n리포트 생성: {filepath}")
    print(f"  전체 후보: {len(all_matches)}개")
    print(f"  우선 매칭 (75+): {len(high_matches)}개")
    print(f"\n  '매칭 체크리스트' 시트에서 후보 URL 확인 → WING 수동 매칭")

    return filepath

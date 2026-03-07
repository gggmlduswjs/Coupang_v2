"""Wing 일괄등록 Excel 파일 관리 모듈

Coupong/엑셀/ 폴더의 업로드용 Excel 파일들을 읽기/필터/검색/비교/검증한다.

Excel 형식 (Wing 일괄등록 Ver.4.6):
  Row 1: 섹션 헤더 (병합셀)
  Row 2: 컬럼 헤더
  Row 3: 필수/선택 표시
  Row 4: 입력 안내문
  Row 5+: 실제 데이터

성능 주의: read_only 모드에서는 반드시 iter_rows()를 사용할 것.
           ws.cell() 랜덤 접근은 250배 느림.
"""

import json
import os
import re
from collections import Counter
from pathlib import Path
from typing import Optional

import openpyxl

from core.constants import ORIGINAL_PRICE_RATIO

# ─── 컬럼 매핑 (0-indexed for iter_rows tuple) ──────────
# Excel col → tuple index (col - 1)
COL_CATEGORY = 1       # 카테고리
COL_NAME = 2           # 등록상품명
COL_START_DATE = 3     # 판매시작일
COL_END_DATE = 4       # 판매종료일
COL_STATUS = 5         # 상품상태
COL_BRAND = 7          # 브랜드
COL_MAKER = 8          # 제조사
COL_SEARCH = 9         # 검색어
COL_PRICE = 62         # 판매가격
COL_COMMISSION = 63    # 판매대행수수료
COL_DISCOUNT_REF = 64  # 할인율기준가
COL_STOCK = 65         # 재고수량
COL_LEAD_TIME = 66     # 출고리드타임
COL_SELLER_CODE = 73   # 업체상품코드
COL_BARCODE = 75       # 바코드

# iter_rows 에서 사용할 인덱스 (0-based)
_I_CAT = COL_CATEGORY - 1
_I_NAME = COL_NAME - 1
_I_START = COL_START_DATE - 1
_I_END = COL_END_DATE - 1
_I_STATUS = COL_STATUS - 1
_I_BRAND = COL_BRAND - 1
_I_MAKER = COL_MAKER - 1
_I_SEARCH = COL_SEARCH - 1
_I_PRICE = COL_PRICE - 1
_I_DISC_REF = COL_DISCOUNT_REF - 1
_I_STOCK = COL_STOCK - 1
_I_LEAD = COL_LEAD_TIME - 1
_I_SCODE = COL_SELLER_CODE - 1
_I_BARCODE = COL_BARCODE - 1

MAX_COL = COL_BARCODE  # iter_rows max_col
DATA_START_ROW = 5     # 데이터 시작 행

# Wing 템플릿 파일 경로
_WING_TEMPLATE_PATH = Path(__file__).parent / "data" / "wing_template_ok.json"

# Row 1 병합 셀 범위 (섹션 헤더)
_WING_MERGES = [
    "A1:I1", "J1:U1", "V1:BI1", "BJ1:CJ1",
    "CK1:CY1", "CZ1:DE1", "DG1:DM1",
]


# ─── Wing 템플릿 ──────────────────────────────────────────

def _load_wing_template() -> dict:
    """Wing 업로드 템플릿 Row 1~4 구조를 로드."""
    with open(_WING_TEMPLATE_PATH, encoding="utf-8") as f:
        tpl = json.load(f)
    return {
        int(r): {int(c): v for c, v in cols.items()}
        for r, cols in tpl.items()
        if r.isdigit()
    }


def create_wing_workbook() -> openpyxl.Workbook:
    """올바른 Wing 템플릿 구조(Row 1~4 + 병합셀)가 적용된 새 Workbook 생성."""
    tpl = _load_wing_template()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    for r in range(1, 5):
        if r in tpl:
            for c, v in tpl[r].items():
                ws.cell(r, c).value = v
    for mg in _WING_MERGES:
        ws.merge_cells(mg)
    return wb


def ensure_wing_template(filepath: str) -> bool:
    """기존 파일의 Wing 템플릿 구조가 올바른지 확인, 아니면 보정.

    Returns:
        True if 보정됨, False if 이미 정상.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    r4c1 = _safe_str(ws.cell(4, 1).value)
    col104h = _safe_str(ws.cell(2, 104).value)
    if "상품정보는 5행부터" in r4c1 and col104h == "대표이미지":
        wb.close()
        return False

    tpl = _load_wing_template()
    # 기존 병합 해제 (Row 1~4)
    for mg in list(ws.merged_cells.ranges):
        if mg.min_row <= 4:
            ws.unmerge_cells(str(mg))
    # 템플릿 적용
    for r in range(1, 5):
        if r in tpl:
            for c, v in tpl[r].items():
                ws.cell(r, c).value = v
    for mg in _WING_MERGES:
        ws.merge_cells(mg)
    wb.save(filepath)
    wb.close()
    return True


# ─── 검색어 태그 정리 ─────────────────────────────────────

def clean_search_tag(tag: str) -> str | None:
    """개별 검색어 태그 정리. 부적합하면 None 반환."""
    t = tag.strip()
    if not t or len(t) < 2:
        return None
    if t.isdigit():
        return None
    # 카테고리 코드 [] 포함 태그 제거
    if "[" in t or "]" in t:
        return None
    # 카테고리 경로 > 포함 태그 제거
    if ">" in t:
        return None
    # 길이 제한 (한글 20자)
    if len(t) > 20:
        t = t[:20]
    return t


def clean_search_tags_str(tags_str: str) -> str:
    """'/' 구분 검색어 문자열을 정리하여 반환."""
    if not tags_str:
        return ""
    tags = tags_str.split("/")
    cleaned = []
    seen = set()
    for tag in tags:
        t = clean_search_tag(tag)
        if t and t.lower() not in seen:
            seen.add(t.lower())
            cleaned.append(t)
    return "/".join(cleaned[:20])


# ─── 유틸 ────────────────────────────────────────────────

def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _safe_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _get(row_vals: tuple, idx: int):
    """튜플에서 안전하게 인덱스 접근."""
    if idx < len(row_vals):
        return row_vals[idx]
    return None


# ─── 파일 탐색 ───────────────────────────────────────────

def find_upload_files(base_dir: str, folder: str = "",
                      count: bool = True) -> list[dict]:
    """엑셀/ 하위 모든 .xlsx 파일 스캔 (temp ~$ 제외).

    Args:
        base_dir: 엑셀 디렉토리 경로
        folder: 하위 폴더 (빈 문자열이면 전체)
        count: True면 각 파일의 상품 수도 계산 (느릴 수 있음)

    Returns:
        [{path, folder, account, filename, count}, ...]
    """
    results = []
    search_dir = os.path.join(base_dir, folder) if folder else base_dir

    if not os.path.isdir(search_dir):
        return results

    for root, _dirs, files in os.walk(search_dir):
        for fname in files:
            if not fname.endswith(".xlsx") or fname.startswith("~$"):
                continue
            fpath = os.path.join(root, fname)
            rel = os.path.relpath(root, base_dir)
            account = fname.split("_")[0] if "_" in fname else ""
            cnt = -1
            if count:
                try:
                    cnt = count_products(fpath)
                except Exception:
                    cnt = -1
            results.append({
                "path": fpath,
                "folder": rel,
                "account": account,
                "filename": fname,
                "count": cnt,
            })

    results.sort(key=lambda x: (x["folder"], x["filename"]))
    return results


# ─── 데이터 읽기 ─────────────────────────────────────────

def read_products(filepath: str) -> list[dict]:
    """Row 5+ 데이터를 dict 리스트로 반환. iter_rows 사용으로 고속."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    products = []
    row_num = DATA_START_ROW
    for row_vals in ws.iter_rows(min_row=DATA_START_ROW, max_col=MAX_COL,
                                  values_only=True):
        name = _safe_str(_get(row_vals, _I_NAME))
        if not name:
            row_num += 1
            continue
        products.append({
            "row": row_num,
            "category": _safe_str(_get(row_vals, _I_CAT)),
            "name": name,
            "start_date": _safe_str(_get(row_vals, _I_START)),
            "end_date": _safe_str(_get(row_vals, _I_END)),
            "status": _safe_str(_get(row_vals, _I_STATUS)),
            "brand": _safe_str(_get(row_vals, _I_BRAND)),
            "maker": _safe_str(_get(row_vals, _I_MAKER)),
            "search": _safe_str(_get(row_vals, _I_SEARCH)),
            "price": _safe_int(_get(row_vals, _I_PRICE)),
            "discount_ref": _safe_int(_get(row_vals, _I_DISC_REF)),
            "stock": _safe_int(_get(row_vals, _I_STOCK)),
            "lead_time": _safe_int(_get(row_vals, _I_LEAD)),
            "seller_code": _safe_str(_get(row_vals, _I_SCODE)),
            "barcode": _safe_str(_get(row_vals, _I_BARCODE)),
        })
        row_num += 1
    wb.close()
    return products


def count_products(filepath: str) -> int:
    """데이터 행 수 (row5부터, 빈 행 제외). iter_rows 사용."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    cnt = 0
    for row_vals in ws.iter_rows(min_row=DATA_START_ROW, max_col=COL_NAME,
                                  values_only=True):
        if row_vals[_I_NAME]:
            cnt += 1
    wb.close()
    return cnt


# ─── 필터/삭제 ───────────────────────────────────────────

def filter_products(filepath: str, keywords: list[str],
                    column: int = COL_NAME, dry_run: bool = False) -> tuple[int, list[dict]]:
    """키워드 포함 행 삭제.

    Args:
        filepath: Excel 파일 경로
        keywords: 삭제 대상 키워드 목록 (OR 조건)
        column: 검색 대상 컬럼 (기본: 등록상품명)
        dry_run: True면 삭제하지 않고 매칭 목록만 반환

    Returns:
        (삭제 건수, 매칭된 상품 리스트)
    """
    # filter/delete는 non-read_only 모드 필요 → cell 접근 OK (행 수 적음)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    matched = []
    rows_to_delete = []

    for row in range(DATA_START_ROW, (ws.max_row or DATA_START_ROW) + 1):
        val = _safe_str(ws.cell(row, column).value)
        if not val:
            continue
        val_lower = val.lower()
        for kw in keywords:
            if kw.lower() in val_lower:
                matched.append({
                    "row": row,
                    "name": _safe_str(ws.cell(row, COL_NAME).value),
                    "keyword": kw,
                })
                rows_to_delete.append(row)
                break

    if not dry_run and rows_to_delete:
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        wb.save(filepath)
        wb.close()
        ensure_wing_template(filepath)
    else:
        wb.close()
    return len(matched), matched


def update_filename_count(filepath: str) -> str:
    """파일 내 실제 데이터 수로 파일명 업데이트.

    예: 007-ez_베스트셀러_375권.xlsx → 007-ez_베스트셀러_370권.xlsx
    """
    cnt = count_products(filepath)
    directory = os.path.dirname(filepath)
    fname = os.path.basename(filepath)

    m = re.match(r"^(.+_)(\d+)(권\.xlsx)$", fname)
    if not m:
        return filepath

    new_name = f"{m.group(1)}{cnt}{m.group(3)}"
    new_path = os.path.join(directory, new_name)

    if new_path != filepath:
        os.rename(filepath, new_path)

    return new_path


# ─── 검색 ────────────────────────────────────────────────

def search_products(base_dir: str, query: str, folder: str = "") -> list[dict]:
    """전체 파일에서 상품명에 query 포함된 행 검색."""
    results = []
    files = find_upload_files(base_dir, folder, count=False)
    query_lower = query.lower()

    for f in files:
        try:
            products = read_products(f["path"])
        except Exception:
            continue
        for p in products:
            if query_lower in p["name"].lower():
                results.append({
                    "file": f["filename"],
                    "folder": f["folder"],
                    "account": f["account"],
                    **p,
                })

    return results


# ─── 비교 ────────────────────────────────────────────────

def compare_accounts(base_dir: str, folder: str,
                     account_a: str, account_b: str) -> dict:
    """같은 폴더 내 두 계정의 상품 비교 (상품명 기준)."""
    files = find_upload_files(base_dir, folder, count=False)

    def _load_names(account: str) -> dict[str, dict]:
        names = {}
        for f in files:
            if f["account"] != account:
                continue
            try:
                for p in read_products(f["path"]):
                    names[p["name"]] = p
            except Exception:
                continue
        return names

    names_a = _load_names(account_a)
    names_b = _load_names(account_b)

    set_a = set(names_a.keys())
    set_b = set(names_b.keys())

    return {
        "only_a": sorted(set_a - set_b),
        "only_b": sorted(set_b - set_a),
        "common": sorted(set_a & set_b),
        "count_a": len(set_a),
        "count_b": len(set_b),
    }


# ─── 통계 ────────────────────────────────────────────────

def get_stats(base_dir: str, folder: str = "") -> dict:
    """폴더별/계정별 상품 수, 카테고리 분포, 가격대 분포."""
    files = find_upload_files(base_dir, folder, count=False)

    by_folder: dict[str, int] = Counter()
    by_account: dict[str, int] = Counter()
    categories: Counter = Counter()
    prices: list[int] = []
    total = 0
    file_count = 0

    for f in files:
        try:
            products = read_products(f["path"])
        except Exception:
            continue
        cnt = len(products)
        if cnt == 0:
            continue
        file_count += 1
        by_folder[f["folder"]] += cnt
        by_account[f["account"]] += cnt
        total += cnt
        for p in products:
            if p["category"]:
                top_cat = p["category"].split(">")[1].strip() if ">" in p["category"] else p["category"]
                categories[top_cat] += 1
            if p["price"] and p["price"] > 0:
                prices.append(p["price"])

    price_stats = {}
    if prices:
        prices.sort()
        price_stats = {
            "min": min(prices),
            "max": max(prices),
            "avg": sum(prices) // len(prices),
            "median": prices[len(prices) // 2],
        }

    return {
        "total": total,
        "files": file_count,
        "by_folder": dict(by_folder.most_common()),
        "by_account": dict(by_account.most_common()),
        "categories": dict(categories.most_common(20)),
        "price_stats": price_stats,
    }


# ─── 중복 체크 ───────────────────────────────────────────

def find_duplicates(filepath: str = "", base_dir: str = "",
                    folder: str = "") -> list[dict]:
    """파일 내 또는 파일 간 중복 상품 (상품명 기준)."""
    seen: dict[str, list[dict]] = {}

    if filepath:
        try:
            products = read_products(filepath)
        except Exception:
            return []
        fname = os.path.basename(filepath)
        for p in products:
            key = p["name"]
            entry = {"file": fname, "row": p["row"], "name": p["name"]}
            seen.setdefault(key, []).append(entry)
    elif base_dir:
        files = find_upload_files(base_dir, folder, count=False)
        for f in files:
            try:
                products = read_products(f["path"])
            except Exception:
                continue
            for p in products:
                key = p["name"]
                entry = {"file": f["filename"], "folder": f["folder"],
                         "row": p["row"], "name": p["name"]}
                seen.setdefault(key, []).append(entry)

    duplicates = []
    for name, entries in seen.items():
        if len(entries) > 1:
            duplicates.append({
                "name": name,
                "count": len(entries),
                "locations": entries,
            })

    duplicates.sort(key=lambda x: -x["count"])
    return duplicates


# ─── 검증 ────────────────────────────────────────────────

def validate_upload(filepath: str) -> list[dict]:
    """필수 필드 누락, 날짜 형식 오류, 가격 이상 등 체크."""
    issues = []
    date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    row_num = DATA_START_ROW
    for row_vals in ws.iter_rows(min_row=DATA_START_ROW, max_col=MAX_COL,
                                  values_only=True):
        name = _safe_str(_get(row_vals, _I_NAME))
        if not name:
            row_num += 1
            continue

        category = _safe_str(_get(row_vals, _I_CAT))
        brand = _safe_str(_get(row_vals, _I_BRAND))
        maker = _safe_str(_get(row_vals, _I_MAKER))

        if not category:
            issues.append({"row": row_num, "field": "카테고리", "issue": "필수 필드 누락"})
        if not brand:
            issues.append({"row": row_num, "field": "브랜드", "issue": "필수 필드 누락"})
        if not maker:
            issues.append({"row": row_num, "field": "제조사", "issue": "필수 필드 누락"})

        start_date = _safe_str(_get(row_vals, _I_START))
        if start_date and not date_pattern.match(start_date):
            issues.append({"row": row_num, "field": "판매시작일",
                           "issue": f"날짜 형식 오류: '{start_date}' (YYYY-MM-DD 필요)"})

        end_date = _safe_str(_get(row_vals, _I_END))
        if end_date and not date_pattern.match(end_date):
            issues.append({"row": row_num, "field": "판매종료일",
                           "issue": f"날짜 형식 오류: '{end_date}' (YYYY-MM-DD 필요)"})

        price = _safe_int(_get(row_vals, _I_PRICE))
        if price is None or price <= 0:
            issues.append({"row": row_num, "field": "판매가격",
                           "issue": f"가격 이상: {_get(row_vals, _I_PRICE)}"})

        discount_ref = _safe_int(_get(row_vals, _I_DISC_REF))
        if discount_ref is None or discount_ref <= 0:
            issues.append({"row": row_num, "field": "할인율기준가",
                           "issue": f"할인율기준가 누락: {_get(row_vals, _I_DISC_REF)}"})

        stock = _safe_int(_get(row_vals, _I_STOCK))
        if stock is None or stock < 0:
            issues.append({"row": row_num, "field": "재고수량",
                           "issue": f"재고 이상: {_get(row_vals, _I_STOCK)}"})

        lead_time = _safe_int(_get(row_vals, _I_LEAD))
        if lead_time is None or lead_time <= 0:
            issues.append({"row": row_num, "field": "출고리드타임",
                           "issue": f"출고리드타임 누락: {_get(row_vals, _I_LEAD)}"})

        row_num += 1

    wb.close()
    return issues


# ─── 필수 필드 자동 채우기 ────────────────────────────────

def fill_required_fields(filepath: str, *,
                         discount_ref_ratio: float = ORIGINAL_PRICE_RATIO,
                         default_stock: int = 1000,
                         default_lead_time: int = 1,
                         dry_run: bool = False) -> dict:
    """비어있는 필수 필드(할인율기준가, 재고수량, 출고리드타임)를 자동 채움.

    Args:
        filepath: Excel 파일 경로
        discount_ref_ratio: 할인율기준가 = 판매가격 x ratio (ORIGINAL_PRICE_RATIO 사용)
        default_stock: 기본 재고수량
        default_lead_time: 기본 출고리드타임 (일)
        dry_run: True면 수정하지 않고 채울 건수만 반환

    Returns:
        {"discount_ref": N, "stock": N, "lead_time": N, "total": N}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    filled = {"discount_ref": 0, "stock": 0, "lead_time": 0}

    for row in range(DATA_START_ROW, (ws.max_row or DATA_START_ROW) + 1):
        name = ws.cell(row, COL_NAME).value
        if not name:
            continue

        # 할인율기준가: 비어있으면 판매가격 x ratio
        disc = ws.cell(row, COL_DISCOUNT_REF).value
        if disc is None or _safe_str(disc) == "":
            price = ws.cell(row, COL_PRICE).value
            if price is not None:
                try:
                    calc = int(float(price) * discount_ref_ratio)
                    if not dry_run:
                        ws.cell(row, COL_DISCOUNT_REF).value = calc
                    filled["discount_ref"] += 1
                except (ValueError, TypeError):
                    pass

        # 재고수량: 비어있으면 기본값
        stock = ws.cell(row, COL_STOCK).value
        if stock is None or _safe_str(stock) == "":
            if not dry_run:
                ws.cell(row, COL_STOCK).value = default_stock
            filled["stock"] += 1

        # 출고리드타임: 비어있으면 기본값
        lead = ws.cell(row, COL_LEAD_TIME).value
        if lead is None or _safe_str(lead) == "":
            if not dry_run:
                ws.cell(row, COL_LEAD_TIME).value = default_lead_time
            filled["lead_time"] += 1

    filled["total"] = sum(filled.values())
    if not dry_run and filled["total"] > 0:
        wb.save(filepath)
        wb.close()
        ensure_wing_template(filepath)
    else:
        wb.close()
    return filled


# ─── 날짜 형식 수정 ──────────────────────────────────────

def fix_dates(filepath: str) -> int:
    """날짜 형식을 YYYY-MM-DD로 자동 수정. 수정 건수 반환."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    fixed = 0
    date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")

    for row in range(DATA_START_ROW, (ws.max_row or DATA_START_ROW) + 1):
        if not ws.cell(row, COL_NAME).value:
            continue
        for col in [COL_START_DATE, COL_END_DATE]:
            val = ws.cell(row, col).value
            if val is None:
                continue
            s = str(val).strip()
            if date_pattern.match(s):
                continue

            if hasattr(val, "strftime"):
                ws.cell(row, col).value = val.strftime("%Y-%m-%d")
                fixed += 1
                continue

            converted = re.sub(r"[/.]", "-", s)
            m = re.match(r"^(\d{4})(\d{2})(\d{2})$", s)
            if m:
                converted = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

            if date_pattern.match(converted) and converted != s:
                ws.cell(row, col).value = converted
                fixed += 1

    if fixed > 0:
        wb.save(filepath)
        wb.close()
        ensure_wing_template(filepath)
    else:
        wb.close()
    return fixed


# ─── 상품 추가 (TSV) ─────────────────────────────────────

# ─── SEO 태그 고도화 상수 ─────────────────────────────────

# 단편 → 완전체 교정 (None은 삭제)
FRAGMENT_FIX: dict[str, str | None] = {
    "확률과": "확률과통계", "생명과": "생명과학", "지구과": "지구과학",
    "생활과": "생활과윤리", "윤리와": "윤리와사상", "법과": "법과정치",
    "유형의": None, "상위권을": None, "위한": None, "명품": None,
    "기말고": "기말고사", "중간고": "중간고사",
}

# 불용어 (태그에서 제거)
STOPWORDS = {
    "교육과정", "개정", "전", "권", "년", "용", "최신", "반영",
    "포함", "제공", "무료", "선물", "사은품", "상철", "좌무선",
    "ver", "KIE", "세트", "구성", "에서", "으로",
    "부터", "까지", "통해", "위해", "위한", "대한", "관한",
}

# 조사/접속사 패턴 (토큰 끝에서 제거)
_JOSA_PATTERN = re.compile(
    r"(의|과|와|이|가|을|를|은|는|에|도|로|으로|에서|부터|까지|에게|한테|"
    r"보다|처럼|만큼|같이|대로|밖에|뿐|마다|조차|이나|나|든지|이든지)$"
)


def generate_search_tags(name: str, brand: str = "", category: str = "") -> list[str]:
    """상품명/브랜드/카테고리에서 검색어 태그 자동 생성 (최대 20개).

    추출 순서:
    1. 브랜드명
    2. SUBJECTS 매칭 (ad_report.py 상수)
    3. LEVELS 매칭
    4. 상품명 핵심 토큰 (2자 이상, 조사/접속사 제거)
    5. 카테고리 키워드
    6. 복합 키워드 조합 (브랜드+과목, 과목+학년)
    7. 연도 태그
    """
    from ad_report import SUBJECTS, LEVELS

    tags: list[str] = []
    seen: set[str] = set()

    def _add(tag: str):
        t = clean_search_tag(tag)
        if not t or t in seen:
            return
        if t in STOPWORDS:
            return
        seen.add(t)
        tags.append(t)

    # 1. 브랜드명
    if brand:
        _add(brand)

    # 2. SUBJECTS 매칭
    found_subjects = []
    for subj in SUBJECTS:
        if subj in name:
            _add(subj)
            found_subjects.append(subj)

    # 3. LEVELS 매칭
    found_levels = []
    for level in LEVELS:
        if level in name:
            _add(level)
            found_levels.append(level)

    # 4. 상품명 핵심 토큰
    tokens = re.findall(r"[가-힣a-zA-Z0-9]+", name)
    for tok in tokens:
        if len(tok) < 2:
            continue
        if tok.isdigit():
            continue
        # 조사 제거
        cleaned = _JOSA_PATTERN.sub("", tok)
        if len(cleaned) < 2:
            continue
        if cleaned in STOPWORDS:
            continue
        _add(cleaned)

    # 5. 카테고리 키워드 (> 분리, [코드] 제거)
    if category:
        cat_parts = [p.strip() for p in category.split(">")]
        for part in cat_parts:
            # [35065] 도서 → 도서 (카테고리 코드 제거)
            part = re.sub(r"\[.*?\]\s*", "", part).strip()
            if part and len(part) >= 2:
                _add(part)

    # 6. 복합 키워드 조합
    b = brand or ""
    for subj in found_subjects[:3]:
        if b:
            _add(f"{b} {subj}")
        for level in found_levels[:2]:
            _add(f"{subj} {level}")

    # 7. 연도 태그
    year_match = re.search(r"(202[4-9]|203[0-9])", name)
    if year_match:
        _add(year_match.group(1))
    else:
        _add("2026")

    return tags[:20]


# ─── 태그 최적화 ─────────────────────────────────────────

def optimize_tags(filepath: str, dry_run: bool = True) -> dict:
    """엑셀 파일의 검색어 태그를 고도화.

    기존 태그 파싱 → 단편 교정 → 새 태그 생성/병합 → Col 9에 쓰기.

    Returns:
        {"total": 상품수, "updated": 수정수, "avg_before": float, "avg_after": float}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    total = 0
    updated = 0
    sum_before = 0
    sum_after = 0

    for row in range(DATA_START_ROW, (ws.max_row or DATA_START_ROW) + 1):
        name = _safe_str(ws.cell(row, COL_NAME).value)
        if not name:
            continue
        total += 1

        brand = _safe_str(ws.cell(row, COL_BRAND).value)
        category = _safe_str(ws.cell(row, COL_CATEGORY).value)
        old_search = _safe_str(ws.cell(row, COL_SEARCH).value)

        # 기존 태그 파싱
        old_tags = [t.strip() for t in old_search.split("/") if t.strip()] if old_search else []
        sum_before += len(old_tags)

        # 기존 태그 교정 (단편 → 완전체, 불용어 제거)
        cleaned_old: list[str] = []
        seen_old: set[str] = set()
        for tag in old_tags:
            if tag in FRAGMENT_FIX:
                fixed = FRAGMENT_FIX[tag]
                if fixed is None:
                    continue  # 삭제 대상
                tag = fixed
            if tag in STOPWORDS or len(tag) < 2 or tag.isdigit():
                continue
            if tag not in seen_old:
                seen_old.add(tag)
                cleaned_old.append(tag)

        # 새 태그 생성
        new_tags = generate_search_tags(name, brand, category)

        # 병합 (기존 유효 태그 우선 + 새 태그 추가, 중복 제거, 형식 검증)
        merged: list[str] = []
        merged_set: set[str] = set()
        for t in cleaned_old:
            ct = clean_search_tag(t)
            if ct:
                low = ct.lower()
                if low not in merged_set:
                    merged_set.add(low)
                    merged.append(ct)
        for t in new_tags:
            ct = clean_search_tag(t)
            if ct:
                low = ct.lower()
                if low not in merged_set:
                    merged_set.add(low)
                    merged.append(ct)

        merged = merged[:20]
        sum_after += len(merged)

        new_search = "/".join(merged)
        if new_search != old_search:
            updated += 1
            if not dry_run:
                ws.cell(row, COL_SEARCH).value = new_search

    if not dry_run and updated > 0:
        wb.save(filepath)
        wb.close()
        ensure_wing_template(filepath)
    else:
        wb.close()

    return {
        "total": total,
        "updated": updated,
        "avg_before": round(sum_before / total, 1) if total else 0,
        "avg_after": round(sum_after / total, 1) if total else 0,
    }


# ─── SEO 감사 ────────────────────────────────────────────

def audit_seo(filepath: str) -> list[dict]:
    """SEO 감사 리포트 생성.

    체크 항목:
    - 상품명 길이 (37-80자 최적)
    - 검색어 태그 수 (10개 미만 경고)
    - 단편 태그 감지
    - 핵심 키워드 누락 (브랜드, 과목)
    - 중복 태그

    Returns:
        [{"row": int, "field": str, "issue": str, "suggestion": str}, ...]
    """
    from ad_report import SUBJECTS

    issues = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    row_num = DATA_START_ROW
    for row_vals in ws.iter_rows(min_row=DATA_START_ROW, max_col=MAX_COL,
                                  values_only=True):
        name = _safe_str(_get(row_vals, _I_NAME))
        if not name:
            row_num += 1
            continue

        brand = _safe_str(_get(row_vals, _I_BRAND))
        search = _safe_str(_get(row_vals, _I_SEARCH))

        # 1. 상품명 길이 체크
        name_len = len(name)
        if name_len < 30:
            issues.append({
                "row": row_num, "field": "상품명",
                "issue": f"너무 짧음 ({name_len}자)",
                "suggestion": "37-80자 권장. 핵심 키워드를 상품명에 포함하세요",
            })
        elif name_len > 80:
            issues.append({
                "row": row_num, "field": "상품명",
                "issue": f"너무 김 ({name_len}자)",
                "suggestion": "80자 이하 권장. 불필요한 수식어를 줄이세요",
            })

        # 2. 검색어 태그 수
        tags = [t.strip() for t in search.split("/") if t.strip()] if search else []
        tag_count = len(tags)
        if tag_count == 0:
            issues.append({
                "row": row_num, "field": "검색어",
                "issue": "검색어 태그 없음",
                "suggestion": "최소 10개 이상 태그 추가 필요 (최대 20개)",
            })
        elif tag_count < 10:
            issues.append({
                "row": row_num, "field": "검색어",
                "issue": f"태그 부족 ({tag_count}개)",
                "suggestion": f"현재 {tag_count}개 → 12개 이상 권장 (optimize로 자동 채우기 가능)",
            })

        # 3. 단편 태그 감지
        for tag in tags:
            if tag in FRAGMENT_FIX:
                fix = FRAGMENT_FIX[tag]
                if fix:
                    issues.append({
                        "row": row_num, "field": "검색어",
                        "issue": f"단편 태그: '{tag}'",
                        "suggestion": f"'{fix}'로 교정 필요",
                    })
                else:
                    issues.append({
                        "row": row_num, "field": "검색어",
                        "issue": f"무의미한 태그: '{tag}'",
                        "suggestion": "삭제 필요",
                    })

        # 4. 핵심 키워드 누락
        tag_joined = "/".join(tags).lower() if tags else ""
        if brand and brand.lower() not in tag_joined:
            issues.append({
                "row": row_num, "field": "검색어",
                "issue": f"브랜드 누락: '{brand}'",
                "suggestion": f"검색어에 브랜드명 '{brand}' 추가 필요",
            })

        found_subj_in_name = [s for s in SUBJECTS if s in name]
        if found_subj_in_name:
            missing = [s for s in found_subj_in_name if s.lower() not in tag_joined]
            for s in missing[:2]:
                issues.append({
                    "row": row_num, "field": "검색어",
                    "issue": f"과목 누락: '{s}'",
                    "suggestion": f"상품명에 '{s}'가 있으나 검색어에 없음",
                })

        # 5. 중복 태그
        tag_lower = [t.lower() for t in tags]
        dup_tags = [t for t in set(tag_lower) if tag_lower.count(t) > 1]
        for dt in dup_tags:
            issues.append({
                "row": row_num, "field": "검색어",
                "issue": f"중복 태그: '{dt}'",
                "suggestion": "중복 태그 제거 필요",
            })

        row_num += 1

    wb.close()
    return issues


# ─── 자동완성 (통합) ──────────────────────────────────

def auto_fill(filepath: str, *, dry_run: bool = False) -> dict:
    """모든 필수 필드를 한번에 자동 채우기.

    순서:
    1. fill_required_fields() -- 할인율기준가/재고/출고리드타임
    2. fix_dates() -- 날짜 형식 YYYY-MM-DD
    3. optimize_tags() -- 검색어 태그 자동 생성/최적화
    4. 상품상태 "판매중" 기본값
    5. 판매시작일 없으면 오늘 날짜
    6. ensure_wing_template() -- 템플릿 구조 보정

    Returns:
        {"fill": dict, "dates": int, "tags": dict,
         "status": int, "start_date": int, "template": bool}
    """
    from datetime import date

    result = {"fill": {}, "dates": 0, "tags": {}, "status": 0, "start_date": 0, "template": False}

    # 1. 필수 필드 (할인율기준가/재고/리드타임)
    result["fill"] = fill_required_fields(filepath, dry_run=dry_run)

    # 2. 날짜 형식 수정
    if not dry_run:
        result["dates"] = fix_dates(filepath)
    else:
        # dry_run에서도 날짜 문제 개수 파악
        date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        cnt = 0
        for row_vals in ws.iter_rows(min_row=DATA_START_ROW, max_col=MAX_COL, values_only=True):
            if not _safe_str(_get(row_vals, _I_NAME)):
                continue
            for idx in [_I_START, _I_END]:
                val = _get(row_vals, idx)
                if val is not None:
                    s = str(val).strip()
                    if s and not date_pattern.match(s):
                        cnt += 1
        wb.close()
        result["dates"] = cnt

    # 3. 태그 최적화
    result["tags"] = optimize_tags(filepath, dry_run=dry_run)

    # 4-5. 상품상태 + 판매시작일
    today = date.today().strftime("%Y-%m-%d")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    status_filled = 0
    date_filled = 0

    for row in range(DATA_START_ROW, (ws.max_row or DATA_START_ROW) + 1):
        name = ws.cell(row, COL_NAME).value
        if not name:
            continue

        # 상품상태 비어있으면 "판매중"
        status = _safe_str(ws.cell(row, COL_STATUS).value)
        if not status:
            if not dry_run:
                ws.cell(row, COL_STATUS).value = "판매중"
            status_filled += 1

        # 판매시작일 비어있으면 오늘
        start = _safe_str(ws.cell(row, COL_START_DATE).value)
        if not start:
            if not dry_run:
                ws.cell(row, COL_START_DATE).value = today
            date_filled += 1

    result["status"] = status_filled
    result["start_date"] = date_filled

    if not dry_run and (status_filled > 0 or date_filled > 0):
        wb.save(filepath)
    wb.close()

    # 6. 템플릿 보정
    if not dry_run:
        result["template"] = ensure_wing_template(filepath)

    return result


def add_products_from_tsv(filepath: str, tsv_path: str) -> int:
    """TSV 파일에서 상품 데이터 읽어 Excel에 추가. 추가 건수 반환."""
    import csv

    with open(tsv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        rows_data = list(reader)

    if not rows_data:
        return 0

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    next_row = (ws.max_row or DATA_START_ROW) + 1
    for r in range(DATA_START_ROW, next_row):
        if not ws.cell(r, COL_NAME).value:
            next_row = r
            break

    field_map = {
        "카테고리": COL_CATEGORY,
        "등록상품명": COL_NAME,
        "판매시작일": COL_START_DATE,
        "브랜드": COL_BRAND,
        "제조사": COL_MAKER,
        "검색어": COL_SEARCH,
        "판매가격": COL_PRICE,
        "재고수량": COL_STOCK,
        "업체상품코드": COL_SELLER_CODE,
        "바코드": COL_BARCODE,
    }

    added = 0
    for data in rows_data:
        for tsv_col, xl_col in field_map.items():
            if tsv_col in data and data[tsv_col]:
                val = data[tsv_col]
                if xl_col in (COL_PRICE, COL_STOCK):
                    try:
                        val = int(float(val))
                    except (ValueError, TypeError):
                        pass
                ws.cell(next_row, xl_col).value = val
        next_row += 1
        added += 1

    wb.save(filepath)
    wb.close()
    ensure_wing_template(filepath)
    return added

"""Excel 리포트 생성 (openpyxl)"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.config import AnalysisConfig


# 스타일 상수
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
AD_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
TOP10_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="1A237E")
SUBTITLE_FONT = Font(bold=True, size=12, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_table(ws, headers, rows, start_row=1):
    """헤더 + 데이터 행 기록"""
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    _style_header_row(ws, start_row, len(headers))

    for r_idx, row_data in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    return start_row + len(rows)


# ──────────────────────────────────────────────
# 시트 생성 함수들
# ──────────────────────────────────────────────

def _create_overview_sheet(wb, analysis):
    ws = wb.active
    ws.title = "개요"

    ws.cell(row=1, column=1, value="쿠팡 검색 알고리즘 분석 리포트").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"키워드: {analysis['keyword']}")
    ws.cell(row=3, column=1, value=f"분석 일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=4, column=1, value=f"총 상품 수: {analysis['total_products']}")

    # 요약 테이블
    row = 6
    ws.cell(row=row, column=1, value="분석 요약").font = SUBTITLE_FONT
    row += 1

    summary_items = []

    # 경쟁 강도
    ci = analysis["competition"].get("competition_index", 0)
    summary_items.append(("경쟁 강도", f"{ci}/100"))

    # 광고 비율
    ar = analysis["ad_patterns"].get("ratio", {})
    if ar:
        summary_items.append(("광고 비율", f"{ar.get('광고비율', 0)}%"))

    # 최적 가격대
    opt = analysis["price_distribution"].get("optimal_range", {})
    if opt:
        summary_items.append(("최적 가격대", f"{int(opt['하한']):,}원 ~ {int(opt['상한']):,}원"))

    # 키워드 포함율
    km = analysis["keyword_matching"]
    if km.get("inclusion_rate"):
        summary_items.append(("키워드 포함율", f"{km['inclusion_rate']}%"))

    # 핵심 순위 팩터
    fi = analysis["ranking_factors"].get("feature_importance", {})
    if fi:
        top_factor = max(fi, key=fi.get)
        summary_items.append(("핵심 순위 팩터", f"{top_factor} ({fi[top_factor]:.1%})"))

    headers = ["항목", "값"]
    rows = summary_items
    _write_table(ws, headers, rows, start_row=row)

    # 전략 요약
    strat = analysis["strategy"]
    row += len(rows) + 3
    ws.cell(row=row, column=1, value="전략 제안 요약").font = SUBTITLE_FONT
    row += 1

    if strat.get("actions"):
        headers = ["우선순위", "항목", "설명"]
        action_rows = [(a["우선순위"], a["항목"], a["설명"]) for a in strat["actions"]]
        _write_table(ws, headers, action_rows, start_row=row)

    _auto_width(ws)


def _create_raw_data_sheet(wb, analysis):
    ws = wb.create_sheet("원본데이터")
    df = analysis["dataframe"]

    col_map = {
        "exposure_order": "노출순서", "ad_type": "광고여부", "organic_rank": "자연검색순위",
        "product_name": "상품명", "original_price": "정가", "discount_rate": "할인율",
        "sale_price": "판매가", "rating": "평점", "review_count": "리뷰수",
        "delivery_type": "배송유형", "free_shipping": "무료배송", "product_id": "상품ID",
        "keyword_in_name": "키워드포함", "keyword_position": "키워드위치",
    }

    headers = list(col_map.values())
    cols = list(col_map.keys())

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, col in enumerate(cols, 1):
            val = row.get(col)
            if col == "keyword_in_name":
                val = "O" if val else "X"
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER

            # 조건부 서식
            if row.get("ad_type") == "AD":
                cell.fill = AD_FILL
            elif row.get("organic_rank") and row.get("organic_rank") <= 10:
                cell.fill = TOP10_FILL

    _auto_width(ws)


def _create_ranking_factors_sheet(wb, analysis):
    ws = wb.create_sheet("순위팩터")
    rf = analysis["ranking_factors"]

    ws.cell(row=1, column=1, value="순위 팩터 상관분석").font = TITLE_FONT

    # Spearman
    row = 3
    ws.cell(row=row, column=1, value="Spearman 순위상관").font = SUBTITLE_FONT
    row += 1
    if rf.get("spearman"):
        headers = ["팩터", "상관계수", "p-value", "유의성"]
        rows = []
        for factor, vals in rf["spearman"].items():
            corr = vals.get("correlation")
            pval = vals.get("p_value")
            sig = "***" if pval and pval < 0.01 else "**" if pval and pval < 0.05 else "*" if pval and pval < 0.1 else ""
            rows.append((factor, corr, pval, sig))
        last_row = _write_table(ws, headers, rows, start_row=row)
        row = last_row + 2

    # Feature Importance
    ws.cell(row=row, column=1, value="RandomForest Feature Importance").font = SUBTITLE_FONT
    row += 1
    fi = rf.get("feature_importance", {})
    if fi:
        headers = ["팩터", "중요도"]
        fi_rows = sorted(fi.items(), key=lambda x: x[1], reverse=True)
        last_row = _write_table(ws, headers, fi_rows, start_row=row)

        # 막대 차트
        chart = BarChart()
        chart.type = "col"
        chart.title = "순위 팩터 중요도"
        chart.y_axis.title = "중요도"
        chart.style = 10
        chart.width = 18
        chart.height = 12

        data_ref = Reference(ws, min_col=2, min_row=row, max_row=last_row)
        cats_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4

        ws.add_chart(chart, f"D{row}")
        row = last_row + 18

    # 다중회귀
    reg = rf.get("regression", {})
    if reg:
        ws.cell(row=row, column=1, value="다중회귀 분석").font = SUBTITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value="R²")
        ws.cell(row=row, column=2, value=reg.get("r_squared"))
        row += 1
        if reg.get("coefficients"):
            headers = ["팩터", "표준화 계수"]
            coeff_rows = list(reg["coefficients"].items())
            _write_table(ws, headers, coeff_rows, start_row=row)

    # VIF
    vif = rf.get("vif", {})
    if vif:
        row += len(reg.get("coefficients", {})) + 3
        ws.cell(row=row, column=1, value="VIF (다중공선성)").font = SUBTITLE_FONT
        row += 1
        headers = ["팩터", "VIF", "판정"]
        vif_rows = [(f, v, "주의" if v > 5 else "양호") for f, v in vif.items()]
        _write_table(ws, headers, vif_rows, start_row=row)

    _auto_width(ws)


def _create_keyword_sheet(wb, analysis):
    ws = wb.create_sheet("키워드분석")
    km = analysis["keyword_matching"]

    ws.cell(row=1, column=1, value="키워드 매칭 분석").font = TITLE_FONT

    row = 3
    items = [
        ("키워드", analysis["keyword"]),
        ("포함율", f"{km.get('inclusion_rate', 0)}%"),
    ]
    rd = km.get("rank_diff", {})
    if rd:
        items.append(("포함 평균순위", rd.get("포함_평균순위")))
        items.append(("미포함 평균순위", rd.get("미포함_평균순위")))

    mw = km.get("mann_whitney", {})
    if mw:
        items.append(("Mann-Whitney p-value", mw.get("p_value")))
        items.append(("통계적 유의미성", "유의미" if mw.get("significant") else "유의미하지 않음"))

    _write_table(ws, ["항목", "값"], items, start_row=row)
    row += len(items) + 2

    # 위치 분포
    pos = km.get("position_dist", {})
    if pos:
        ws.cell(row=row, column=1, value="키워드 위치 분포").font = SUBTITLE_FONT
        row += 1
        headers = ["위치", "상품수"]
        pos_rows = list(pos.items())
        last_row = _write_table(ws, headers, pos_rows, start_row=row)
        row = last_row + 2

    # 연관 키워드
    rw = km.get("related_words", [])
    if rw:
        ws.cell(row=row, column=1, value="연관 키워드 TOP 20").font = SUBTITLE_FONT
        row += 1
        headers = ["키워드", "출현 횟수"]
        rw_rows = [(w["word"], w["count"]) for w in rw]
        _write_table(ws, headers, rw_rows, start_row=row)

    _auto_width(ws)


def _create_ad_sheet(wb, analysis):
    ws = wb.create_sheet("광고분석")
    ap = analysis["ad_patterns"]

    ws.cell(row=1, column=1, value="광고 패턴 분석").font = TITLE_FONT

    row = 3
    ratio = ap.get("ratio", {})
    if ratio:
        items = [(k, v) for k, v in ratio.items()]
        last_row = _write_table(ws, ["항목", "값"], items, start_row=row)

        # 원형 차트
        chart = PieChart()
        chart.title = "광고 vs 자연검색"
        chart.style = 10
        chart.width = 14
        chart.height = 10

        # 파이차트 데이터 삽입
        pie_row = last_row + 2
        ws.cell(row=pie_row, column=1, value="광고")
        ws.cell(row=pie_row, column=2, value=ratio.get("광고수", 0))
        ws.cell(row=pie_row + 1, column=1, value="자연검색")
        ws.cell(row=pie_row + 1, column=2, value=ratio.get("자연검색수", 0))

        data_ref = Reference(ws, min_col=2, min_row=pie_row, max_row=pie_row + 1)
        cats_ref = Reference(ws, min_col=1, min_row=pie_row, max_row=pie_row + 1)
        chart.add_data(data_ref)
        chart.set_categories(cats_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True

        ws.add_chart(chart, "D3")
        row = pie_row + 2

    # 특성 비교
    comp = ap.get("comparison", {})
    if comp:
        row += 2
        ws.cell(row=row, column=1, value="광고 vs 자연검색 비교").font = SUBTITLE_FONT
        row += 1
        headers = ["유형", "평균가격", "평균리뷰수", "평균평점", "로켓배송비율"]
        comp_rows = []
        for label, data in comp.items():
            comp_rows.append((
                label,
                f"{int(data.get('평균가격', 0)):,}원" if data.get('평균가격') else "-",
                f"{int(data.get('평균리뷰수', 0)):,}개" if data.get('평균리뷰수') else "-",
                data.get("평균평점"),
                f"{data.get('로켓배송비율', 0)}%",
            ))
        _write_table(ws, headers, comp_rows, start_row=row)

    _auto_width(ws)


def _create_price_sheet(wb, analysis):
    ws = wb.create_sheet("가격분석")
    pd_result = analysis["price_distribution"]

    ws.cell(row=1, column=1, value="가격대 분석").font = TITLE_FONT

    row = 3
    st = pd_result.get("stats", {})
    if st:
        items = [(k, f"{int(v):,}원") for k, v in st.items()]
        last_row = _write_table(ws, ["통계", "값"], items, start_row=row)
        row = last_row + 2

    # 가격대 분포 + 히스토그램
    bins = pd_result.get("bins", [])
    if bins:
        ws.cell(row=row, column=1, value="가격대별 분포").font = SUBTITLE_FONT
        row += 1
        headers = ["가격대", "상품수"]
        bin_rows = [(b["range"], b["count"]) for b in bins]
        last_row = _write_table(ws, headers, bin_rows, start_row=row)

        chart = BarChart()
        chart.type = "col"
        chart.title = "가격대별 상품 분포"
        chart.y_axis.title = "상품수"
        chart.style = 10
        chart.width = 22
        chart.height = 12

        data_ref = Reference(ws, min_col=2, min_row=row, max_row=last_row)
        cats_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, "D3")
        row = last_row + 2

    # 상위10위 가격
    tp = pd_result.get("top10_price", {})
    if tp:
        row += 14
        ws.cell(row=row, column=1, value="상위 10위 가격대").font = SUBTITLE_FONT
        row += 1
        items = [(k, f"{int(v):,}원") for k, v in tp.items()]
        _write_table(ws, ["항목", "값"], items, start_row=row)
        row += len(items) + 2

    opt = pd_result.get("optimal_range", {})
    if opt:
        ws.cell(row=row, column=1, value="★ 최적 가격대").font = Font(bold=True, size=12, color="FF0000")
        row += 1
        ws.cell(row=row, column=1, value=f"{int(opt['하한']):,}원 ~ {int(opt['상한']):,}원")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)

    # 산점도 (가격 vs 순위)
    df = analysis["dataframe"]
    organic = df[df["ad_type"] == "자연검색"].dropna(subset=["sale_price", "organic_rank"])
    if len(organic) >= 3:
        scatter_start = last_row + 2 if bins else row + 2
        ws.cell(row=scatter_start, column=4, value="판매가")
        ws.cell(row=scatter_start, column=5, value="자연검색순위")
        for i, (_, r) in enumerate(organic.iterrows()):
            ws.cell(row=scatter_start + 1 + i, column=4, value=r["sale_price"])
            ws.cell(row=scatter_start + 1 + i, column=5, value=r["organic_rank"])

        chart = ScatterChart()
        chart.title = "가격 vs 순위"
        chart.x_axis.title = "판매가"
        chart.y_axis.title = "순위 (낮을수록 상위)"
        chart.style = 13
        chart.width = 18
        chart.height = 12

        x_ref = Reference(ws, min_col=4, min_row=scatter_start + 1, max_row=scatter_start + len(organic))
        y_ref = Reference(ws, min_col=5, min_row=scatter_start + 1, max_row=scatter_start + len(organic))
        from openpyxl.chart import Series
        series = Series(y_ref, x_ref, title="상품")
        chart.series.append(series)
        ws.add_chart(chart, f"D{scatter_start + len(organic) + 2}")

    _auto_width(ws)


def _create_competition_sheet(wb, analysis):
    ws = wb.create_sheet("경쟁분석")
    comp = analysis["competition"]

    ws.cell(row=1, column=1, value="경쟁 분석").font = TITLE_FONT

    row = 3
    ci = comp.get("competition_index", 0)
    ws.cell(row=row, column=1, value="경쟁 강도 지수")
    ws.cell(row=row, column=2, value=f"{ci}/100")
    level = "매우 높음" if ci >= 80 else "높음" if ci >= 60 else "보통" if ci >= 40 else "낮음"
    ws.cell(row=row, column=3, value=level)
    row += 2

    # 상위10 vs 나머지
    tvr = comp.get("top_vs_rest", {})
    if tvr:
        ws.cell(row=row, column=1, value="상위 10위 vs 나머지").font = SUBTITLE_FONT
        row += 1
        headers = ["구분", "상품수", "평균리뷰", "평균평점", "평균가격", "로켓배송"]
        rows = []
        for label, data in tvr.items():
            rows.append((
                label, data.get("상품수"),
                f"{int(data.get('평균리뷰수', 0)):,}",
                data.get("평균평점"),
                f"{int(data.get('평균가격', 0)):,}원",
                f"{data.get('로켓배송비율', 0)}%",
            ))
        last_row = _write_table(ws, headers, rows, start_row=row)
        row = last_row + 2

    # 진입 기준
    thr = comp.get("thresholds", {})
    if thr:
        ws.cell(row=row, column=1, value="상위 10위 진입 기준").font = SUBTITLE_FONT
        row += 1
        items = [(k, v) for k, v in thr.items()]
        _write_table(ws, ["항목", "값"], items, start_row=row)

    _auto_width(ws)


def _create_strategy_sheet(wb, analysis):
    ws = wb.create_sheet("전략제안")
    strat = analysis["strategy"]

    ws.cell(row=1, column=1, value="전략 제안").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"키워드: {analysis['keyword']}")

    row = 4

    # 가격 전략
    if strat.get("pricing"):
        ws.cell(row=row, column=1, value="가격 전략").font = SUBTITLE_FONT
        row += 1
        for k, v in strat["pricing"].items():
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            row += 1
        row += 1

    # 키워드 전략
    kw = strat.get("keyword", {})
    if kw:
        ws.cell(row=row, column=1, value="키워드 전략").font = SUBTITLE_FONT
        row += 1
        for k, v in kw.items():
            if isinstance(v, list):
                v = ", ".join(v)
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            row += 1
        row += 1

    # 리뷰 전략
    if strat.get("review"):
        ws.cell(row=row, column=1, value="리뷰 전략").font = SUBTITLE_FONT
        row += 1
        for k, v in strat["review"].items():
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            row += 1
        row += 1

    # 배송 전략
    if strat.get("delivery"):
        ws.cell(row=row, column=1, value="배송 전략").font = SUBTITLE_FONT
        row += 1
        for k, v in strat["delivery"].items():
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            row += 1
        row += 1

    # 액션 체크리스트
    if strat.get("actions"):
        row += 1
        ws.cell(row=row, column=1, value="액션 체크리스트").font = SUBTITLE_FONT
        row += 1
        headers = ["우선순위", "항목", "설명"]
        action_rows = [(a["우선순위"], a["항목"], a["설명"]) for a in strat["actions"]]
        _write_table(ws, headers, action_rows, start_row=row)

    _auto_width(ws)


def _create_sales_sheet(wb, analysis):
    ws = wb.create_sheet("판매량추정")
    se = analysis.get("sales_estimation", {})

    ws.cell(row=1, column=1, value="판매량 추정 (리뷰 기반)").font = TITLE_FONT

    row = 3
    ms = se.get("market_size", {})
    if ms:
        ws.cell(row=row, column=1, value="시장 규모").font = SUBTITLE_FONT
        row += 1
        items = [
            ("총 추정 판매량", f"{ms.get('총_추정_판매량', 0):,}개"),
            ("총 추정 매출", f"{ms.get('총_추정_매출', 0):,}원"),
            ("평균 추정 판매량", f"{int(ms.get('평균_추정_판매량', 0)):,}개"),
            ("리뷰-구매 전환율", f"{ms.get('전환율_기준', 0)}"),
        ]
        last_row = _write_table(ws, ["항목", "값"], items, start_row=row)
        row = last_row + 2

    # 집중도
    conc = se.get("concentration", {})
    if conc:
        ws.cell(row=row, column=1, value="판매 집중도").font = SUBTITLE_FONT
        row += 1
        items = [
            ("상위10 판매량 점유율", f"{conc.get('상위10_판매량_점유율', 0)}%"),
            ("상위10 매출 점유율", f"{conc.get('상위10_매출_점유율', 0)}%"),
        ]
        last_row = _write_table(ws, ["항목", "값"], items, start_row=row)
        row = last_row + 2

    # TOP 10 매출 테이블 + 차트
    top_products = se.get("top_products", [])
    if top_products:
        ws.cell(row=row, column=1, value="TOP 10 매출 상품").font = SUBTITLE_FONT
        row += 1
        headers = ["순위", "상품명", "판매가", "리뷰수", "추정 판매량", "추정 매출"]
        tp_rows = [
            (p["순위"], p["상품명"], f"{p['판매가']:,}원", f"{p['리뷰수']:,}",
             f"{p['추정_판매량']:,}", f"{p['추정_매출']:,}원")
            for p in top_products
        ]
        last_row = _write_table(ws, headers, tp_rows, start_row=row)

        # 매출 막대차트 (수치 데이터 별도 삽입)
        chart_data_row = last_row + 2
        ws.cell(row=chart_data_row, column=1, value="상품")
        ws.cell(row=chart_data_row, column=2, value="추정매출")
        for i, p in enumerate(top_products):
            ws.cell(row=chart_data_row + 1 + i, column=1, value=f"{p['순위']}위")
            ws.cell(row=chart_data_row + 1 + i, column=2, value=p["추정_매출"])

        chart = BarChart()
        chart.type = "col"
        chart.title = "TOP 10 추정 매출"
        chart.y_axis.title = "추정 매출 (원)"
        chart.style = 10
        chart.width = 20
        chart.height = 12

        data_ref = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + len(top_products))
        cats_ref = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(top_products))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"D{row}")

    _auto_width(ws)


def _create_rank_tracking_sheet(wb, analysis):
    ws = wb.create_sheet("순위변동")
    rt = analysis.get("rank_tracking", {})

    ws.cell(row=1, column=1, value="시계열 순위 변동").font = TITLE_FONT

    row = 3
    if not rt.get("tracking_available"):
        ws.cell(row=row, column=1, value=rt.get("note", "스냅샷 부족"))
        _auto_width(ws)
        return

    # 개요
    items = [
        ("이전 스냅샷", rt.get("prev_snapshot", "")[:19]),
        ("최신 스냅샷", rt.get("latest_snapshot", "")[:19]),
        ("공통 상품수", rt.get("common_products", 0)),
        ("변동성 지수", f"{rt.get('volatility_index', 0)}/100"),
    ]
    last_row = _write_table(ws, ["항목", "값"], items, start_row=row)
    row = last_row + 2

    GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    # 상승 상품
    risers = rt.get("risers", [])
    if risers:
        ws.cell(row=row, column=1, value="상승 상품").font = SUBTITLE_FONT
        row += 1
        headers = ["상품ID", "이전순위", "현재순위", "변동"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, len(headers))
        for r_idx, r_data in enumerate(risers, row + 1):
            vals = [r_data["product_id"], r_data["이전순위"], r_data["현재순위"], f"+{r_data['변동']}"]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = THIN_BORDER
                cell.fill = GREEN_FILL
        row = row + len(risers) + 2

    # 하락 상품
    fallers = rt.get("fallers", [])
    if fallers:
        ws.cell(row=row, column=1, value="하락 상품").font = SUBTITLE_FONT
        row += 1
        headers = ["상품ID", "이전순위", "현재순위", "변동"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, len(headers))
        for r_idx, r_data in enumerate(fallers, row + 1):
            vals = [r_data["product_id"], r_data["이전순위"], r_data["현재순위"], str(r_data["변동"])]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = THIN_BORDER
                cell.fill = RED_FILL
        row = row + len(fallers) + 2

    ws.cell(row=row, column=1, value=f"안정 상품: {rt.get('stable_count', 0)}개")

    _auto_width(ws)


def _create_seller_sheet(wb, analysis):
    ws = wb.create_sheet("셀러집중도")
    sc = analysis.get("seller_concentration", {})

    ws.cell(row=1, column=1, value="셀러 집중도 분석").font = TITLE_FONT

    row = 3
    if sc.get("seller_count", 0) == 0:
        ws.cell(row=row, column=1, value="셀러 데이터 없음")
        _auto_width(ws)
        return

    # HHI 요약
    hhi = sc.get("hhi", 0)
    structure = sc.get("market_structure", "")
    items = [
        ("셀러 수", sc["seller_count"]),
        ("HHI", hhi),
        ("시장 구조", structure),
        ("판정 기준", "HHI < 1500: 경쟁적 / 1500~2500: 중간 / 2500+: 고도집중"),
    ]
    last_row = _write_table(ws, ["항목", "값"], items, start_row=row)
    row = last_row + 2

    # TOP 5 셀러
    top_sellers = sc.get("top_sellers", [])
    if top_sellers:
        ws.cell(row=row, column=1, value="TOP 5 셀러").font = SUBTITLE_FONT
        row += 1
        headers = ["셀러ID", "상품수", "점유율"]
        ts_rows = [(s["seller_id"], s["상품수"], f"{s['점유율']}%") for s in top_sellers]
        last_row = _write_table(ws, headers, ts_rows, start_row=row)

        # 원형 차트
        pie_row = last_row + 2
        ws.cell(row=pie_row, column=1, value="기타")
        other_share = 100 - sum(s["점유율"] for s in top_sellers)
        ws.cell(row=pie_row, column=2, value=round(max(other_share, 0), 1))
        for i, s in enumerate(top_sellers):
            ws.cell(row=pie_row + 1 + i, column=1, value=s["seller_id"])
            ws.cell(row=pie_row + 1 + i, column=2, value=s["점유율"])

        chart = PieChart()
        chart.title = "셀러 점유율"
        chart.style = 10
        chart.width = 14
        chart.height = 10

        data_ref = Reference(ws, min_col=2, min_row=pie_row, max_row=pie_row + len(top_sellers))
        cats_ref = Reference(ws, min_col=1, min_row=pie_row, max_row=pie_row + len(top_sellers))
        chart.add_data(data_ref)
        chart.set_categories(cats_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        ws.add_chart(chart, f"D{row}")

    _auto_width(ws)


# ──────────────────────────────────────────────
# 메인: 리포트 생성
# ──────────────────────────────────────────────

def generate_report(analysis: dict, config: AnalysisConfig = None) -> str:
    """분석 결과를 Excel 리포트로 생성. 파일 경로 반환."""
    config = config or AnalysisConfig()
    config.ensure_dirs()

    if not analysis:
        print("  분석 데이터가 없습니다.")
        return ""

    wb = Workbook()

    print("  시트 생성 중...")
    _create_overview_sheet(wb, analysis)
    _create_raw_data_sheet(wb, analysis)
    _create_ranking_factors_sheet(wb, analysis)
    _create_keyword_sheet(wb, analysis)
    _create_ad_sheet(wb, analysis)
    _create_price_sheet(wb, analysis)
    _create_competition_sheet(wb, analysis)
    _create_sales_sheet(wb, analysis)
    _create_rank_tracking_sheet(wb, analysis)
    _create_seller_sheet(wb, analysis)
    _create_strategy_sheet(wb, analysis)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_{analysis['keyword']}_{timestamp}.xlsx"
    filepath = os.path.join(config.reports_dir, filename)

    wb.save(filepath)
    print(f"  리포트 저장: {filepath}")
    return filepath

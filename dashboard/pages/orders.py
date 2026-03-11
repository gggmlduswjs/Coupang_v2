"""
주문/배송 통합 페이지
====================
탭1: 결제완료 (ACCEPT) → 발주확인
탭2: 상품준비중 (INSTRUCT) → Stepper (다운로드→한진→송장등록)
탭3: 배송현황 (DEPARTURE+) → 조회 전용
탭4: 📊 운영현황 → 검색/KPI/대기주문/배치요약/다운로드/이력
"""
import io
import logging
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from st_aggrid import AgGrid, GridOptionsBuilder

from core.api.wing_client import CoupangWingError
from core.constants import (
    is_gift_item,
    match_publisher_from_text,
    resolve_distributor,
)
from dashboard.utils import (
    create_wing_client,
    query_df,
    query_df_cached,
)
from dashboard.services.order_data import (
    load_all_orders_live,
    get_instruct_orders,
    get_instruct_by_box,
    clear_order_caches,
    fmt_krw_short,
    STATUS_MAP,
    build_delivery_rows,
    build_delivery_excel_bytes,
)
from dashboard.services.order_service import (
    load_hanjin_creds as _load_hanjin_creds,
    save_hanjin_creds as _save_hanjin_creds,
    update_orders_status_after_invoice as _update_orders_status,
)
from dashboard.services.invoice_matcher import (
    list_batches,
    load_latest_batch,
    match_invoices,
    check_registerable,
)
from core.database import SessionLocal
from core.models.delivery_log import DeliveryListLog
from core.models.purchase_order_log import PurchaseOrderLog

logger = logging.getLogger(__name__)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# KPI 색상 뱃지
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_KPI_COLORS = {
    "결제완료": "🔴",
    "상품준비중": "🟡",
    "배송지시": "🔵",
    "배송중": "🟢",
    "배송완료(30일)": "⚪",
}


def render(selected_account, accounts_df, account_names):
    st.title("주문/배송")

    # ── 상단 컨트롤 + 글로벌 계정 필터 ──
    _top_c1, _top_c2, _top_c3 = st.columns([2, 3, 2])
    with _top_c1:
        if st.button("🔄 새로고침", key="btn_live_refresh", use_container_width=True,
                     help="WING API에서 실시간 주문 조회", type="primary"):
            clear_order_caches()
            for _sk in ("_step_delivery_list", "_step_hanjin", "_step_invoice",
                        "_t2_current_step"):
                st.session_state.pop(_sk, None)
            st.rerun()
    with _top_c2:
        # 글로벌 계정 필터 — 모든 탭에 적용
        _global_accts = st.multiselect(
            "계정", account_names, default=account_names,
            key="global_acct_filter", label_visibility="collapsed",
        )
    with _top_c3:
        _last_synced = st.session_state.get("order_last_synced")
        if _last_synced:
            st.caption(f"마지막: {_last_synced}")

    # ── 데이터 로드 ──
    _all_orders = load_all_orders_live(accounts_df)

    # 글로벌 계정 필터 적용
    if not _all_orders.empty and _global_accts:
        _all_orders = _all_orders[_all_orders["계정"].isin(_global_accts)].copy()

    def _filter_status(df, status):
        if df.empty:
            return pd.DataFrame()
        return df[df["상태"] == status].copy()

    def _kpi_count(df, status):
        sub = _filter_status(df, status)
        if sub.empty:
            return {}
        return sub.groupby("계정")["묶음배송번호"].nunique().to_dict()

    _accept_all = _filter_status(_all_orders, "ACCEPT")
    _instruct_live = _filter_status(_all_orders, "INSTRUCT")
    _instruct_all = _instruct_live[~_instruct_live["취소"]].copy() if not _instruct_live.empty else pd.DataFrame()

    _kpi_accept = _accept_all.groupby("계정")["묶음배송번호"].nunique().to_dict() if not _accept_all.empty else {}
    _kpi_instruct = _instruct_all.groupby("계정")["묶음배송번호"].nunique().to_dict() if not _instruct_all.empty else {}
    _kpi_departure = _kpi_count(_all_orders, "DEPARTURE")
    _kpi_delivering = _kpi_count(_all_orders, "DELIVERING")
    _kpi_final = _kpi_count(_all_orders, "FINAL_DELIVERY")

    # ── 상단 KPI (컬러 뱃지) ──
    _kc1, _kc2, _kc3, _kc4, _kc5 = st.columns(5)

    def _render_kpi(col, label, counts):
        total = sum(counts.values())
        badge = _KPI_COLORS.get(label, "")
        col.metric(f"{badge} {label}", f"{total:,}건")
        if counts:
            parts = [f"{k}: {v}" for k, v in sorted(counts.items())]
            col.caption(" | ".join(parts))

    _render_kpi(_kc1, "결제완료", _kpi_accept)
    _render_kpi(_kc2, "상품준비중", _kpi_instruct)
    _render_kpi(_kc3, "배송지시", _kpi_departure)
    _render_kpi(_kc4, "배송중", _kpi_delivering)
    _render_kpi(_kc5, "배송완료(30일)", _kpi_final)

    st.divider()

    # ── 4탭 ──
    _tab1, _tab2, _tab3, _tab4 = st.tabs(["결제완료", "상품준비중", "배송현황", "📊 운영현황"])

    # ══════════════════════════════════════
    # 탭1: 결제완료 (ACCEPT) → 발주확인
    # ══════════════════════════════════════
    with _tab1:
        # ── 발주확인 완료 토스트 (rerun 후 표시) ──
        if "_ack_success_count" in st.session_state:
            _ack_cnt = st.session_state.pop("_ack_success_count")
            st.toast(f"발주확인 완료 ({_ack_cnt}건) — 상품준비중 탭에서 배송리스트를 다운로드하세요.", icon="✅")

        if _accept_all.empty:
            st.info("결제완료 주문이 없습니다. 새로고침으로 확인하세요.")
        else:
            _accept_display = _accept_all.copy()
            _accept_display["상품/옵션/수량"] = _accept_display.apply(
                lambda r: f"{r['상품명']} / {r['옵션명']} / {int(r['수량'])}권", axis=1
            )
            _accept_display["수취인/연락처"] = _accept_display.apply(
                lambda r: f"{r['수취인']}" + (f" ({r['수취인전화번호']})" if r.get('수취인전화번호') else ""), axis=1
            )
            _accept_display["배송상태"] = _accept_display["상태"].map(STATUS_MAP)
            _accept_display["결제금액"] = _accept_display["결제금액"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "0")

            _accept_total = _accept_all["묶음배송번호"].nunique()
            _accept_amount = int(_accept_all["결제금액"].apply(lambda x: int(str(x).replace(",", "")) if pd.notna(x) else 0).sum()) if not _accept_all.empty else 0

            _display_cols = [
                "주문번호", "상품/옵션/수량", "수취인/연락처", "수취인주소", "배송상태",
                "주문일시", "묶음배송번호", "계정", "결제금액",
            ]
            _grid_df = _accept_display[_display_cols].rename(columns={"수취인주소": "배송지"})

            gb = GridOptionsBuilder.from_dataframe(_grid_df)
            gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=50)
            gb.configure_default_column(resizable=True, sorteable=True, filterable=True)
            gb.configure_column("상품/옵션/수량", width=350)
            gb.configure_column("배송지", width=250)
            gb.configure_column("수취인/연락처", width=130)
            gb.configure_column("주문번호", width=120)
            gb.configure_column("묶음배송번호", width=120)
            gb.configure_column("주문일시", width=140)
            gb.configure_selection(
                "multiple", use_checkbox=True,
                header_checkbox=True,
                pre_selected_rows=list(range(len(_grid_df))),
            )
            grid_opts = gb.build()
            _grid_result = AgGrid(
                _grid_df, gridOptions=grid_opts, height=500, theme="streamlit",
                key="t1_accept_grid", update_mode="SELECTION_CHANGED",
            )

            _selected_rows = _grid_result.get("selected_rows", None)
            _has_selection = False
            if _selected_rows is not None:
                if isinstance(_selected_rows, pd.DataFrame) and not _selected_rows.empty:
                    _selected_df = _selected_rows
                    _has_selection = True
                elif isinstance(_selected_rows, list) and len(_selected_rows) > 0:
                    _selected_df = pd.DataFrame(_selected_rows)
                    _has_selection = True

            if _has_selection:
                _sel_box_ids = _selected_df["묶음배송번호"].unique().tolist()
                _sel_data = _accept_all[_accept_all["묶음배송번호"].isin(_sel_box_ids)].copy()
                _sel_boxes = len(_sel_box_ids)
            else:
                _sel_data = _accept_all.copy()
                _sel_boxes = _accept_total

            # ── 발주확인 액션바 ──
            _ab1, _ab2 = st.columns([3, 1])
            with _ab1:
                if st.button(f"발주확인 ({_sel_boxes}건 → 상품준비중)", type="primary", key="t1_btn_ack",
                             use_container_width=True):
                    _total_success = 0
                    _total_fail = 0

                    for _aid, _grp in _sel_data.groupby("_account_id"):
                        _acct_name = _grp.iloc[0]["계정"]
                        _acct_row = accounts_df[accounts_df["id"] == _aid]
                        if _acct_row.empty:
                            st.error(f"[{_acct_name}] 계정 정보를 찾을 수 없습니다.")
                            continue
                        _client = create_wing_client(_acct_row.iloc[0])
                        if not _client:
                            st.error(f"[{_acct_name}] WING API 클라이언트 생성 실패")
                            continue

                        _ack_ids = _grp["묶음배송번호"].unique().tolist()
                        try:
                            _ack_result = _client.acknowledge_ordersheets([int(x) for x in _ack_ids])
                            _success_ids = []
                            _fail_items = []
                            if isinstance(_ack_result, dict) and "data" in _ack_result:
                                _resp_data = _ack_result["data"]
                                _resp_code = _resp_data.get("responseCode")
                                _resp_list = _resp_data.get("responseList", [])
                                for _item in _resp_list:
                                    if _item.get("succeed"):
                                        _success_ids.append(_item["shipmentBoxId"])
                                    else:
                                        _fail_items.append(_item)
                                if _resp_code == 0:
                                    st.toast(f"[{_acct_name}] 완료: {len(_success_ids)}건", icon="✅")
                                elif _resp_code == 1:
                                    st.warning(f"[{_acct_name}] 부분 성공: {len(_success_ids)}건 성공, {len(_fail_items)}건 실패")
                                    for _fi in _fail_items:
                                        st.error(f"  {_fi.get('shipmentBoxId')}: {_fi.get('resultMessage', '')}")
                                elif _resp_code == 99:
                                    st.error(f"[{_acct_name}] 전체 실패: {_resp_data.get('responseMessage', '')}")
                                else:
                                    _success_ids = [int(x) for x in _ack_ids]
                            else:
                                _success_ids = [int(x) for x in _ack_ids]

                            _total_success += len(_success_ids)
                            _total_fail += len(_fail_items)

                        except CoupangWingError as e:
                            st.error(f"[{_acct_name}] API 오류: {e}")
                            _total_fail += len(_ack_ids)

                    if _total_success > 0:
                        clear_order_caches()
                        st.session_state["_ack_success_count"] = _total_success
                        st.rerun()
            with _ab2:
                pass  # 우측 여백

            # ── 주문 취소 (접힌 상태) ──
            with st.expander("주문 취소", expanded=False):
                _render_cancel_section(accounts_df, account_names, _accept_all, _instruct_live)

    # ══════════════════════════════════════
    # 탭2: 상품준비중 (INSTRUCT) — Stepper
    # ══════════════════════════════════════
    with _tab2:
        # ── Flash 토스트 (rerun 후 표시) ──
        if "_flash_messages" in st.session_state:
            for _ftype, _fmsg in st.session_state.pop("_flash_messages"):
                _icon = {"success": "✅", "warning": "⚠️", "error": "❌"}.get(_ftype, "ℹ️")
                st.toast(_fmsg, icon=_icon)

        if _instruct_all.empty:
            st.info("상품준비중 주문이 없습니다. 결제완료 탭에서 발주확인 후 여기에 표시됩니다.")
        else:
            _inst_by_box = get_instruct_by_box(_instruct_all)
            _inst_total = len(_inst_by_box)
            _inst_amount = int(_inst_by_box["결제금액"].sum()) if not _inst_by_box.empty else 0

            # ── Stepper 헤더 ──
            _step_done = {
                "dl": st.session_state.get("_step_delivery_list", False),
                "hj": st.session_state.get("_step_hanjin", False),
                "inv": st.session_state.get("_step_invoice", False),
            }

            # 현재 활성 단계 결정
            if not _step_done["dl"]:
                _auto_step = 0
            elif not _step_done["hj"]:
                _auto_step = 1
            elif not _step_done["inv"]:
                _auto_step = 2
            else:
                _auto_step = 2  # 모두 완료 시 마지막 단계

            _current_step = st.session_state.get("_t2_current_step", _auto_step)

            _step_labels = [
                ("dl", "① 다운로드"),
                ("hj", "② 한진"),
                ("inv", "③ 송장등록"),
            ]

            # Stepper 바
            _sc = st.columns(len(_step_labels))
            for _i, (_col, (_skey, _slabel)) in enumerate(zip(_sc, _step_labels)):
                _done = _step_done[_skey]
                _active = (_i == _current_step)
                if _done:
                    _prefix = "✅"
                elif _active:
                    _prefix = "▶"
                else:
                    _prefix = "⬜"
                if _col.button(
                    f"{_prefix} {_slabel}",
                    key=f"t2_step_btn_{_i}",
                    use_container_width=True,
                    type="primary" if _active else "secondary",
                ):
                    st.session_state["_t2_current_step"] = _i
                    st.rerun()

            st.caption(f"상품준비중 {_inst_total:,}건 · ₩{fmt_krw_short(_inst_amount)}")

            # 주문 그리드 — 접힌 상태로 제공 (항상 전체 선택이 기본)
            with st.expander(f"주문 목록 ({_inst_total}건)", expanded=False):
                _inst_display = _inst_by_box[["계정", "묶음배송번호", "주문번호", "상품명", "수량", "결제금액", "주문일", "수취인"]].copy()
                _inst_display["결제금액_표시"] = _inst_display["결제금액"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "0")
                _grid_cols = ["계정", "묶음배송번호", "주문번호", "상품명", "수량", "결제금액_표시", "주문일", "수취인"]
                _inst_grid_df = _inst_display[_grid_cols].rename(columns={"결제금액_표시": "결제금액"})

                gb_inst = GridOptionsBuilder.from_dataframe(_inst_grid_df)
                gb_inst.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
                gb_inst.configure_default_column(resizable=True, sorteable=True, filterable=True)
                gb_inst.configure_column("상품명", width=350)
                gb_inst.configure_selection(
                    "multiple", use_checkbox=True,
                    header_checkbox=True,
                    pre_selected_rows=list(range(len(_inst_grid_df))),
                )
                _inst_grid_result = AgGrid(
                    _inst_grid_df, gridOptions=gb_inst.build(), height=400, theme="streamlit",
                    key="t2_instruct_grid", update_mode="SELECTION_CHANGED",
                )

                _t2_selected_rows = _inst_grid_result.get("selected_rows", None)
                _t2_has_sel = False
                if _t2_selected_rows is not None:
                    if isinstance(_t2_selected_rows, pd.DataFrame) and not _t2_selected_rows.empty:
                        _t2_sel_df = _t2_selected_rows
                        _t2_has_sel = True
                    elif isinstance(_t2_selected_rows, list) and len(_t2_selected_rows) > 0:
                        _t2_sel_df = pd.DataFrame(_t2_selected_rows)
                        _t2_has_sel = True

                if _t2_has_sel:
                    _sel_box_ids = _t2_sel_df["묶음배송번호"].unique().tolist()
                    st.session_state["_t2_sel_box_ids"] = _sel_box_ids
                    st.caption(f"선택: {len(_t2_sel_df)}건 ({len(_sel_box_ids)}묶음) — 체크 해제한 주문은 엑셀/송장에서 제외")
                else:
                    st.session_state.pop("_t2_sel_box_ids", None)

            # 선택 필터 적용
            _t2_sel_ids = st.session_state.get("_t2_sel_box_ids")
            if _t2_sel_ids:
                _t2_filtered = _instruct_all[_instruct_all["묶음배송번호"].isin(_t2_sel_ids)].copy()
            else:
                _t2_filtered = _instruct_all.copy()

            st.divider()

            # ── Stepper 본문: 활성 단계만 렌더 ──
            if _current_step == 0:
                # ── ① 다운로드 단계 ──
                _render_delivery_list(_t2_filtered, accounts_df)
                # 발주서 (접힌 상태)
                _render_purchase_order(_t2_filtered, accounts_df, key_prefix="t2")

            elif _current_step == 1:
                # ── ② 한진 단계 ──
                _render_hanjin_nfocus()

                # 이전 단계 요약
                if _step_done["dl"]:
                    _last_bid = st.session_state.get("_last_batch_id")
                    if _last_bid:
                        _dl_cnt = query_df(
                            "SELECT COUNT(*) AS cnt FROM delivery_list_logs WHERE batch_id = :bid",
                            {"bid": _last_bid},
                        )
                        _cnt = int(_dl_cnt.iloc[0]["cnt"]) if not _dl_cnt.empty else 0
                        st.caption(f"이전 단계: {_cnt}건 다운로드 완료")

            elif _current_step == 2:
                # ── ③ 송장등록 단계 ──
                # 한진 미완료 시 안내
                if not _step_done["hj"]:
                    st.warning("② 한진 단계를 먼저 완료하세요. 한진에서 출력자료등록 엑셀을 다운로드한 후 여기서 업로드합니다.")

                _render_invoice_upload(_t2_filtered, accounts_df)

    # ══════════════════════════════════════
    # 탭3: 배송현황 (DEPARTURE+) — 조회 전용
    # ══════════════════════════════════════
    with _tab3:
        _t3_data = _filter_status(_all_orders, "DEPARTURE")

        _t3_total = _t3_data["묶음배송번호"].nunique() if not _t3_data.empty else 0
        _t3_amount = int(_t3_data["결제금액"].sum()) if not _t3_data.empty else 0
        _t3k1, _t3k2 = st.columns(2)
        _t3k1.metric("배송지시 주문", f"{_t3_total:,}건")
        _t3k2.metric("총 금액", f"₩{fmt_krw_short(_t3_amount)}")

        if _t3_data.empty:
            st.info("배송지시 주문이 없습니다. 송장 등록이 완료되면 여기에 표시됩니다.")
        else:
            _t3_display = _t3_data.copy()
            _t3_display["상품/옵션/수량"] = _t3_display.apply(
                lambda r: f"{r['상품명']} / {r['옵션명']} / {int(r['수량'])}권", axis=1
            )
            _HANJIN_TRACK = "https://www.hanjin.com/kor/CMS/DeliveryMgr/WaybillResult.do?wblnumText2={}&schLang=KR"
            _t3_display["배송추적"] = _t3_display["운송장번호"].apply(
                lambda x: _HANJIN_TRACK.format(str(x).strip()) if x and str(x).strip() else None
            )
            _t3_cols = ["주문번호", "상품/옵션/수량", "수취인", "운송장번호", "배송추적",
                        "수취인주소", "주문일시", "묶음배송번호", "계정"]
            _t3_show = _t3_display[_t3_cols].rename(columns={"수취인주소": "배송지"})

            st.dataframe(
                _t3_show,
                hide_index=True, use_container_width=True, height=500,
                column_config={
                    "상품/옵션/수량": st.column_config.TextColumn(width="large"),
                    "배송지": st.column_config.TextColumn(width="medium"),
                    "배송추적": st.column_config.LinkColumn("배송추적", display_text="추적"),
                    "묶음배송번호": st.column_config.NumberColumn(format="%d"),
                },
            )

        # ── 반품/교환 요청 확인 ──
        with st.expander("반품/교환 요청 확인", expanded=False):
            if st.button("반품/교환 요청 조회 (오늘)", key="t3_btn_returns"):
                _return_results = []
                _t3_from = date.today().isoformat()
                _t3_to = date.today().isoformat()
                _checked_aids = set()
                _aid_list = accounts_df["id"].tolist()
                _clients_cache = {}
                for _aid in _aid_list:
                    if int(_aid) in _checked_aids:
                        continue
                    _checked_aids.add(int(_aid))
                    _acct_row = accounts_df[accounts_df["id"] == int(_aid)]
                    if _acct_row.empty:
                        continue
                    _acct_row = _acct_row.iloc[0]
                    _client = create_wing_client(_acct_row)
                    if not _client:
                        continue
                    _clients_cache[int(_aid)] = _client
                    try:
                        for _rs_code, _rs_label in [("RU", "출고중지"), ("UC", "반품접수"), ("CC", "수거완료")]:
                            _reqs = _client.get_all_return_requests(_t3_from, _t3_to, status=_rs_code)
                            for _req in _reqs:
                                _return_results.append({
                                    "_account_id": int(_aid),
                                    "_receipt_id": _req.get("receiptId"),
                                    "_cancel_count": _req.get("cancelCountSum", 1),
                                    "계정": _acct_row["account_name"],
                                    "유형": _rs_label,
                                    "주문번호": _req.get("orderId"),
                                    "상품명": ((_req.get("returnItems") or [{}])[0].get("vendorItemName", ""))[:40] if _req.get("returnItems") else "",
                                    "사유": _req.get("cancelReason", ""),
                                    "접수일시": str(_req.get("createdAt", ""))[:16].replace("T", " "),
                                })
                    except Exception as e:
                        st.warning(f"[{_acct_row['account_name']}] 조회 실패: {e}")

                st.session_state["_t3_return_results"] = _return_results
                st.session_state["_t3_return_clients"] = _clients_cache

            # 결과 표시 (세션에서 읽어서 rerun 후에도 유지)
            _return_results = st.session_state.get("_t3_return_results")
            _clients_cache = st.session_state.get("_t3_return_clients", {})
            if _return_results is not None:
                if not _return_results:
                    st.success("오늘 반품/교환/출고중지 요청이 없습니다.")
                else:
                    _ru_items = [r for r in _return_results if r["유형"] == "출고중지"]
                    _other_items = [r for r in _return_results if r["유형"] != "출고중지"]

                    # 출고중지(RU) — 액션 가능
                    if _ru_items:
                        st.warning(f"출고중지 요청 {len(_ru_items)}건 — 미출고 시 '출고중지완료' 처리 필요")
                        _ru_df = pd.DataFrame(_ru_items)
                        st.dataframe(
                            _ru_df[["계정", "주문번호", "상품명", "사유", "접수일시"]],
                            hide_index=True, use_container_width=True,
                        )
                        # 출고중지완료 처리 버튼
                        _ru_confirm = st.checkbox(
                            f"출고중지 {len(_ru_items)}건을 '출고중지완료' 처리 (해당 주문 미출고 확인)",
                            key="t3_ru_confirm",
                        )
                        if _ru_confirm:
                            if st.button(f"출고중지완료 처리 ({len(_ru_items)}건)", key="t3_btn_ru_complete", type="primary"):
                                _ru_ok = 0
                                _ru_fail = 0
                                for _ri in _ru_items:
                                    _aid = _ri["_account_id"]
                                    _rid = _ri["_receipt_id"]
                                    _cnt = _ri["_cancel_count"]
                                    _cl = _clients_cache.get(_aid)
                                    if not _cl or not _rid:
                                        _ru_fail += 1
                                        continue
                                    try:
                                        _cl.stop_shipment(receipt_id=int(_rid), cancel_count=int(_cnt))
                                        _ru_ok += 1
                                    except Exception as e:
                                        st.error(f"[{_ri['계정']}] 주문 {_ri['주문번호']}: {e}")
                                        _ru_fail += 1
                                if _ru_ok > 0:
                                    st.toast(f"출고중지완료 {_ru_ok}건 처리", icon="✅")
                                    st.session_state.pop("_t3_return_results", None)
                                    st.session_state.pop("_t3_return_clients", None)
                                    clear_order_caches()
                                    st.rerun()
                                if _ru_fail > 0:
                                    st.error(f"실패 {_ru_fail}건")

                    # 반품접수/수거완료 — 조회만
                    if _other_items:
                        st.info(f"반품접수/수거완료 {len(_other_items)}건 (조회 전용)")
                        _ot_df = pd.DataFrame(_other_items)
                        st.dataframe(
                            _ot_df[["계정", "유형", "주문번호", "상품명", "사유", "접수일시"]],
                            hide_index=True, use_container_width=True,
                        )

    # ══════════════════════════════════════
    # 탭4: 📊 운영현황 (오늘 현황 / 이력 검색)
    # ══════════════════════════════════════
    with _tab4:
        _render_order_stats(_all_orders, accounts_df)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 하위 렌더 함수들
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def _render_order_stats(all_orders, accounts_df):
    """탭4: 운영 현황 — 단일 스크롤 페이지 (검색 → KPI → 대기주문 → 배치요약 → 다운로드 → 이력)"""

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Section 0: 빠른 검색 (최상단)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    _sc1, _sc2 = st.columns([4, 1])
    with _sc1:
        _search_q = st.text_input(
            "🔍", placeholder="수취인 또는 상품명 검색",
            key="t4_quick_search", label_visibility="collapsed",
        )
    with _sc2:
        _search_days = st.selectbox(
            "기간", [7, 14, 30], index=2,
            format_func=lambda d: f"{d}일",
            key="t4_search_days", label_visibility="collapsed",
        )

    if _search_q and _search_q.strip():
        _q = _search_q.strip()
        _results_frames = []

        # 1) API 실시간 데이터 (all_orders) 인메모리 필터
        if not all_orders.empty:
            _mask = (
                all_orders["수취인"].str.contains(_q, case=False, na=False)
                | all_orders["상품명"].str.contains(_q, case=False, na=False)
                | all_orders["옵션명"].str.contains(_q, case=False, na=False)
            )
            _api_hits = all_orders[_mask].copy()
            if not _api_hits.empty:
                _api_hits["_source"] = "API"
                _results_frames.append(_api_hits)

        # 2) DB 과거 데이터 (FINAL_DELIVERY 등)
        _db_hits = query_df("""
            SELECT
                a.account_name AS "계정",
                o.receiver_name AS "수취인",
                o.vendor_item_name AS "상품명",
                o.shipping_count AS "수량",
                o.status AS "상태",
                o.invoice_number AS "운송장번호",
                CAST(o.ordered_at + INTERVAL '9 hours' AS timestamp) AS "주문일시",
                o.shipment_box_id AS "묶음배송번호"
            FROM orders o
            JOIN accounts a ON o.account_id = a.id
            WHERE o.ordered_at >= CURRENT_DATE - :days * INTERVAL '1 day'
              AND o.canceled = false
              AND (o.receiver_name LIKE :kw OR o.vendor_item_name LIKE :kw)
            ORDER BY o.ordered_at DESC
            LIMIT 200
        """, {"days": _search_days, "kw": f"%{_q}%"})

        if not _db_hits.empty:
            _db_hits["_source"] = "DB"
            _results_frames.append(_db_hits)

        if _results_frames:
            _combined = pd.concat(_results_frames, ignore_index=True)
            # 중복 제거 (묶음배송번호 기준, API 우선)
            _combined["_priority"] = _combined["_source"].map({"API": 0, "DB": 1})
            _combined = _combined.sort_values("_priority").drop_duplicates(
                subset=["묶음배송번호"], keep="first"
            )
            _combined["상태(한글)"] = _combined["상태"].map(STATUS_MAP).fillna(_combined["상태"])
            _combined["주문일시"] = pd.to_datetime(_combined["주문일시"], errors="coerce").dt.strftime("%m/%d %H:%M")

            _show_cols = ["주문일시", "계정", "수취인", "상품명", "수량", "상태(한글)", "운송장번호"]
            _show_cols = [c for c in _show_cols if c in _combined.columns]
            st.caption(f"검색 결과: {len(_combined)}건")
            st.dataframe(
                _combined[_show_cols], hide_index=True, use_container_width=True,
                column_config={"상품명": st.column_config.TextColumn(width="large")},
            )
        else:
            st.info(f"'{_q}' 검색 결과 없음 (최근 {_search_days}일)")

        st.divider()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Section 1: KPI
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    _kpi = query_df("""
        SELECT
            (SELECT COUNT(DISTINCT shipment_box_id) FROM orders
             WHERE (ordered_at + INTERVAL '9 hours')::date = CURRENT_DATE
               AND canceled = false) AS today_orders,
            (SELECT COUNT(DISTINCT batch_id) FROM delivery_list_logs
             WHERE (downloaded_at + INTERVAL '9 hours')::date = CURRENT_DATE) AS today_dl_batches,
            (SELECT COUNT(*) FROM delivery_list_logs
             WHERE registered = true
               AND (downloaded_at + INTERVAL '9 hours')::date = CURRENT_DATE) AS today_registered,
            (SELECT COUNT(*) FROM delivery_list_logs
             WHERE registered = false) AS total_pending,
            (SELECT COUNT(DISTINCT batch_id) FROM purchase_order_logs
             WHERE (ordered_at + INTERVAL '9 hours')::date = CURRENT_DATE) AS today_po_batches
    """)
    _k = _kpi.iloc[0] if not _kpi.empty else {}

    _c1, _c2, _c3, _c4, _c5 = st.columns(5)
    _c1.metric("오늘 주문", f"{int(_k.get('today_orders', 0)):,}건")
    _c2.metric("배송리스트", f"{int(_k.get('today_dl_batches', 0))}회")
    _c3.metric("송장 등록", f"{int(_k.get('today_registered', 0)):,}건")
    _c4.metric("미등록 (전체)", f"{int(_k.get('total_pending', 0)):,}건",
               delta=f"{int(_k.get('total_pending', 0))}건" if int(_k.get("total_pending", 0)) > 0 else None,
               delta_color="inverse")
    _c5.metric("발주서", f"{int(_k.get('today_po_batches', 0))}회")

    st.divider()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Section 2: 처리 대기 주문 (ACCEPT / INSTRUCT)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    if not all_orders.empty:
        _pending = all_orders[
            all_orders["상태"].isin(["ACCEPT", "INSTRUCT"]) & ~all_orders["취소"]
        ].copy()
    else:
        _pending = pd.DataFrame()

    if _pending.empty:
        st.info("처리 대기 주문 없음")
    else:
        # 합포장 감지: 묶음배송번호에 2건 이상 → 📦
        _box_counts = _pending.groupby("묶음배송번호").size()
        _multi_boxes = set(_box_counts[_box_counts > 1].index)
        _pending["합포장"] = _pending["묶음배송번호"].apply(lambda x: "📦" if x in _multi_boxes else "")

        _accept_pending = _pending[_pending["상태"] == "ACCEPT"]
        _instruct_pending = _pending[_pending["상태"] == "INSTRUCT"]

        st.subheader(f"처리 대기 주문 ({len(_pending)}건)")

        _pc1, _pc2 = st.columns(2)
        _pending_cols = ["합포장", "계정", "수취인", "옵션명", "수량", "주문일시"]

        with _pc1:
            _n_accept = len(_accept_pending)
            st.markdown(f"**결제완료 ({_n_accept}건)**")
            if _accept_pending.empty:
                st.caption("없음")
            else:
                _ad = _accept_pending[_pending_cols].copy()
                _ad["주문일시"] = pd.to_datetime(_ad["주문일시"], errors="coerce").dt.strftime("%m/%d %H:%M")
                st.dataframe(_ad, hide_index=True, use_container_width=True, height=300)

        with _pc2:
            _n_instruct = len(_instruct_pending)
            st.markdown(f"**상품준비중 ({_n_instruct}건)**")
            if _instruct_pending.empty:
                st.caption("없음")
            else:
                _id = _instruct_pending[_pending_cols].copy()
                _id["주문일시"] = pd.to_datetime(_id["주문일시"], errors="coerce").dt.strftime("%m/%d %H:%M")
                st.dataframe(_id, hide_index=True, use_container_width=True, height=300)

        if _multi_boxes:
            st.caption(f"📦 합포장 {len(_multi_boxes)}건 — 같은 묶음배송번호에 2건 이상")

    st.divider()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Section 3: 배치 요약 (2컬럼)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    _bc1, _bc2 = st.columns(2)

    with _bc1:
        st.markdown("**최근 배송리스트**")
        _dl_batches = query_df("""
            SELECT
                d.batch_id,
                (MIN(d.downloaded_at) + INTERVAL '9 hours')::timestamp AS 다운로드일시,
                COUNT(*) AS 총건수,
                SUM(CASE WHEN d.registered THEN 1 ELSE 0 END) AS 등록완료,
                SUM(CASE WHEN NOT d.registered THEN 1 ELSE 0 END) AS 미등록
            FROM delivery_list_logs d
            WHERE d.batch_id IS NOT NULL
            GROUP BY d.batch_id
            ORDER BY MIN(d.downloaded_at) DESC
            LIMIT 5
        """)
        if _dl_batches.empty:
            st.caption("이력 없음")
        else:
            _dl_batches["다운로드일시"] = pd.to_datetime(_dl_batches["다운로드일시"]).dt.strftime("%m/%d %H:%M")
            _dl_batches["상태"] = _dl_batches.apply(
                lambda r: "✅ 전체 등록" if r["미등록"] == 0 else f"⏳ 미등록 {int(r['미등록'])}건", axis=1
            )
            st.dataframe(
                _dl_batches[["다운로드일시", "총건수", "등록완료", "미등록", "상태"]],
                hide_index=True, use_container_width=True,
                column_config={
                    "총건수": st.column_config.NumberColumn(format="%d건"),
                    "등록완료": st.column_config.NumberColumn(format="%d건"),
                    "미등록": st.column_config.NumberColumn(format="%d건"),
                },
            )

    with _bc2:
        st.markdown("**최근 발주서**")
        _po_batches = query_df("""
            SELECT
                p.batch_id,
                (MIN(p.ordered_at) + INTERVAL '9 hours')::timestamp AS 발주일시,
                COUNT(*) AS 총건수,
                SUM(p.quantity) AS 총수량,
                STRING_AGG(DISTINCT COALESCE(p.distributor, '미지정'), ', ') AS 거래처
            FROM purchase_order_logs p
            WHERE p.batch_id IS NOT NULL
            GROUP BY p.batch_id
            ORDER BY MIN(p.ordered_at) DESC
            LIMIT 5
        """)
        if _po_batches.empty:
            st.caption("이력 없음")
        else:
            _po_batches["발주일시"] = pd.to_datetime(_po_batches["발주일시"]).dt.strftime("%m/%d %H:%M")
            st.dataframe(
                _po_batches[["발주일시", "총건수", "총수량", "거래처"]],
                hide_index=True, use_container_width=True,
                column_config={
                    "총건수": st.column_config.NumberColumn(format="%d건"),
                    "총수량": st.column_config.NumberColumn(format="%d권"),
                },
            )

    st.divider()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Section 4: 다운로드 (expander)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    with st.expander("📋 당일 합산 발주서"):
        _po_orders_raw = query_df("""
            SELECT DISTINCT ON (o.shipment_box_id)
                   o.shipment_box_id AS "묶음배송번호",
                   o.order_id AS "주문번호",
                   o.seller_product_name AS "상품명",
                   o.vendor_item_name AS "옵션명",
                   o.shipping_count AS "수량",
                   o.order_price AS "결제금액",
                   o.account_id AS "_account_id",
                   o.vendor_item_id AS "_vendor_item_id",
                   o.seller_product_id AS "_seller_product_id"
            FROM orders o
            WHERE (o.ordered_at + INTERVAL '9 hours')::date = CURRENT_DATE
              AND o.canceled = false
            ORDER BY o.shipment_box_id, o.updated_at DESC
        """)

        if _po_orders_raw.empty:
            st.info("오늘 주문이 없습니다.")
        else:
            _po_enriched = _enrich_purchase_order_data(_po_orders_raw)
            _po_enriched["거래처"] = _po_enriched["거래처"].fillna("미지정")
            _po_enriched["출판사"] = _po_enriched["출판사"].fillna("")

            _po_enriched["_group_key"] = _po_enriched.apply(
                lambda r: r["ISBN"] if r.get("ISBN") else r["도서명"], axis=1
            )
            _po_agg = _po_enriched.groupby(["거래처", "출판사", "_group_key"]).agg(
                도서명=("도서명", "first"), ISBN=("ISBN", "first"), 주문수량=("수량", "sum"),
            ).reset_index().drop(columns=["_group_key"])
            _po_agg = _po_agg.sort_values(["거래처", "출판사", "도서명"])

            _po_summary = _po_agg.groupby("거래처").agg(
                종수=("도서명", "count"), 총수량=("주문수량", "sum"),
            ).reset_index().sort_values("총수량", ascending=False)
            st.dataframe(
                _po_summary, hide_index=True, use_container_width=True,
                column_config={
                    "종수": st.column_config.NumberColumn(format="%d종"),
                    "총수량": st.column_config.NumberColumn(format="%d권"),
                },
            )

            with st.expander(f"발주 상세 ({len(_po_agg)}종 / {int(_po_agg['주문수량'].sum())}권)"):
                st.dataframe(
                    _po_agg, hide_index=True, use_container_width=True,
                    column_config={
                        "도서명": st.column_config.TextColumn(width="large"),
                        "주문수량": st.column_config.NumberColumn(format="%d권"),
                    },
                )

            _store_name = st.session_state.get("order_store_name", "잉글리쉬존")
            _xl_buf = _build_purchase_order_excel(_po_agg, _store_name)
            _today_str = date.today().strftime("%Y%m%d")
            st.download_button(
                "📥 당일 합산 발주서 다운로드",
                data=_xl_buf,
                file_name=f"발주서_합산_{_today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t4_po_merged_dl",
            )
            st.caption(f"오늘 주문 {len(_po_orders_raw)}건 기준 (orders 테이블 직접 조회)")

    with st.expander("📦 당일 극동 출고 엑셀"):
        # orders 테이블에서 직접 조회 → enrich → books 정보 JOIN
        _gk_orders_raw = query_df("""
            SELECT DISTINCT ON (o.shipment_box_id)
                   o.shipment_box_id AS "묶음배송번호",
                   o.order_id AS "주문번호",
                   o.seller_product_name AS "상품명",
                   o.vendor_item_name AS "옵션명",
                   o.shipping_count AS "수량",
                   o.order_price AS "결제금액",
                   o.account_id AS "_account_id",
                   o.vendor_item_id AS "_vendor_item_id",
                   o.seller_product_id AS "_seller_product_id"
            FROM orders o
            WHERE (o.ordered_at + INTERVAL '9 hours')::date = CURRENT_DATE
              AND o.canceled = false
            ORDER BY o.shipment_box_id, o.updated_at DESC
        """)

        if _gk_orders_raw.empty:
            st.info("오늘 주문이 없습니다.")
        else:
            _gk_enriched = _enrich_purchase_order_data(_gk_orders_raw)
            # books 테이블에서 정가/저자/출판년도/공급률 추가
            _gk_book_info = query_df_cached("""
                SELECT b.isbn, b.list_price, COALESCE(b.author, '') AS author,
                       b.year AS pub_year, pub.supply_rate
                FROM books b
                LEFT JOIN publishers pub ON b.publisher_id = pub.id
                WHERE b.isbn IS NOT NULL AND b.isbn != ''
            """)
            _gk_book_map = {}
            if not _gk_book_info.empty:
                for _, _br in _gk_book_info.iterrows():
                    _gk_book_map[str(_br["isbn"])] = {
                        "정가": _br["list_price"],
                        "저자": str(_br["author"]) if pd.notna(_br["author"]) else "",
                        "출판년도": _br["pub_year"],
                        "공급률": _br["supply_rate"],
                    }
            _gk_po = _gk_enriched.copy()
            _gk_po["상품바코드"] = _gk_po["ISBN"].fillna("").astype(str)
            _gk_po["도서명"] = _gk_po["도서명"].fillna("").astype(str)
            _gk_po["정가"] = _gk_po["상품바코드"].map(lambda x: _gk_book_map.get(x, {}).get("정가"))
            _gk_po["저자"] = _gk_po["상품바코드"].map(lambda x: _gk_book_map.get(x, {}).get("저자", ""))
            _gk_po["출판년도"] = _gk_po["상품바코드"].map(lambda x: _gk_book_map.get(x, {}).get("출판년도"))
            _gk_po["공급률"] = _gk_po["상품바코드"].map(lambda x: _gk_book_map.get(x, {}).get("공급률"))

            _gk_key = _gk_po.apply(lambda r: r["상품바코드"] if r["상품바코드"] else r["도서명"], axis=1)
            _gk_po["_key"] = _gk_key
            _gk_agg = _gk_po.groupby("_key").agg(
                상품바코드=("상품바코드", "first"),
                상품명=("도서명", "first"),
                정가=("정가", "first"),
                수량=("수량", "sum"),
                공급률=("공급률", "first"),
                출판사=("출판사", "first"),
                저자=("저자", "first"),
                출판년도=("출판년도", "first"),
            ).reset_index(drop=True)

            _gk_show = _gk_agg[["상품바코드", "상품명", "수량", "출판사"]].copy()
            st.dataframe(_gk_show, hide_index=True, use_container_width=True)

            _gk_result = pd.DataFrame()
            _gk_result["NO."] = range(1, len(_gk_agg) + 1)
            _gk_result["상품바코드"] = _gk_agg["상품바코드"].values
            _gk_result["상품명"] = _gk_agg["상품명"].values
            _gk_result["#"] = ""
            _gk_result["정 가"] = _gk_agg["정가"].apply(lambda x: int(x) if pd.notna(x) else 0).values
            _gk_result["수 량"] = _gk_agg["수량"].values
            _gk_result["%"] = _gk_agg["공급률"].apply(lambda x: f"{x*100:.0f}" if pd.notna(x) and x else "").values
            _gk_result["단 가"] = _gk_agg.apply(
                lambda r: int(r["정가"] * r["공급률"]) if pd.notna(r["공급률"]) and r["공급률"] and pd.notna(r["정가"]) else (int(r["정가"]) if pd.notna(r["정가"]) else 0),
                axis=1
            ).values
            _gk_result["금 액"] = (_gk_result["단 가"] * _gk_result["수 량"]).values
            _gk_result[""] = ""
            _gk_result["출판사"] = _gk_agg["출판사"].apply(lambda x: str(x) if pd.notna(x) else "").values
            _gk_result["저자"] = _gk_agg["저자"].apply(lambda x: str(x) if pd.notna(x) else "").values
            _gk_result["출판년도"] = _gk_agg["출판년도"].apply(lambda x: str(int(x)) if pd.notna(x) else "").values

            _gk_buf = io.BytesIO()
            with pd.ExcelWriter(_gk_buf, engine="openpyxl") as writer:
                _gk_result.to_excel(writer, sheet_name="극동", index=False)
            _gk_buf.seek(0)

            st.download_button(
                f"📥 극동 다운로드 ({len(_gk_agg)}종 / {int(_gk_agg['수량'].sum())}권)",
                _gk_buf.getvalue(),
                file_name=f"극동_{date.today().strftime('%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t4_gk_xlsx_dl",
            )

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Section 5: 이력 검색 (expander, collapsed)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    with st.expander("📂 이력 검색", expanded=False):
        _render_history_search()


def _render_history_search():
    """이력 검색 — 배치 상세 + 주문 DB 조회"""

    _search_mode = st.radio(
        "검색 대상", ["배송리스트", "발주서", "주문 DB"],
        horizontal=True, key="t4_search_mode",
    )

    if _search_mode == "배송리스트":
        _dl_batches = query_df("""
            SELECT
                d.batch_id,
                (MIN(d.downloaded_at) + INTERVAL '9 hours')::timestamp AS 다운로드일시,
                COUNT(*) AS 총건수,
                SUM(CASE WHEN d.registered THEN 1 ELSE 0 END) AS 등록완료,
                SUM(CASE WHEN NOT d.registered THEN 1 ELSE 0 END) AS 미등록
            FROM delivery_list_logs d
            WHERE d.batch_id IS NOT NULL
            GROUP BY d.batch_id
            ORDER BY MIN(d.downloaded_at) DESC
            LIMIT 20
        """)
        if _dl_batches.empty:
            st.info("배송리스트 이력이 없습니다.")
            return

        _dl_batches["다운로드일시"] = pd.to_datetime(_dl_batches["다운로드일시"]).dt.strftime("%m/%d %H:%M")
        _dl_batches["상태"] = _dl_batches.apply(
            lambda r: "✅ 전체 등록" if r["미등록"] == 0 else f"⏳ 미등록 {int(r['미등록'])}건", axis=1
        )
        st.dataframe(
            _dl_batches[["다운로드일시", "총건수", "등록완료", "미등록", "상태"]],
            hide_index=True, use_container_width=True,
            column_config={
                "총건수": st.column_config.NumberColumn(format="%d건"),
                "등록완료": st.column_config.NumberColumn(format="%d건"),
                "미등록": st.column_config.NumberColumn(format="%d건"),
            },
        )

        _dl_options = []
        for _, _b in _dl_batches.iterrows():
            _dl_options.append(f"{_b['다운로드일시']} ({int(_b['총건수'])}건, {_b['상태']})")
        _dl_sel = st.selectbox("배치 상세 조회", range(len(_dl_options)),
                               format_func=lambda i: _dl_options[i], key="t4_dl_batch_sel")
        _dl_sel_bid = _dl_batches.iloc[_dl_sel]["batch_id"]

        _dl_detail = query_df("""
            SELECT
                d.seq_no AS 순번,
                d.shipment_box_id AS 묶음배송번호,
                d.order_id AS 주문번호,
                d.receiver_name AS 수취인,
                d.buyer_name AS 구매자,
                a.account_name AS 계정,
                d.registered AS 송장등록
            FROM delivery_list_logs d
            JOIN accounts a ON d.account_id = a.id
            WHERE d.batch_id = :bid
            ORDER BY d.seq_no
        """, {"bid": _dl_sel_bid})

        if not _dl_detail.empty:
            _reg_cnt = int(_dl_detail["송장등록"].sum())
            _unreg_cnt = len(_dl_detail) - _reg_cnt
            st.caption(f"총 {len(_dl_detail)}건 — 등록 {_reg_cnt} / 미등록 {_unreg_cnt}")
            st.dataframe(
                _dl_detail,
                hide_index=True, use_container_width=True,
                column_config={
                    "묶음배송번호": st.column_config.NumberColumn(format="%d"),
                    "주문번호": st.column_config.NumberColumn(format="%d"),
                    "송장등록": st.column_config.CheckboxColumn("송장등록", disabled=True),
                },
            )

    elif _search_mode == "발주서":
        _po_batches = query_df("""
            SELECT
                p.batch_id,
                (MIN(p.ordered_at) + INTERVAL '9 hours')::timestamp AS 발주일시,
                COUNT(*) AS 총건수,
                SUM(p.quantity) AS 총수량,
                STRING_AGG(DISTINCT COALESCE(p.distributor, '미지정'), ', ') AS 거래처
            FROM purchase_order_logs p
            WHERE p.batch_id IS NOT NULL
            GROUP BY p.batch_id
            ORDER BY MIN(p.ordered_at) DESC
            LIMIT 20
        """)
        if _po_batches.empty:
            st.info("발주서 이력이 없습니다.")
            return

        _po_batches["발주일시"] = pd.to_datetime(_po_batches["발주일시"]).dt.strftime("%m/%d %H:%M")
        st.dataframe(
            _po_batches[["발주일시", "총건수", "총수량", "거래처"]],
            hide_index=True, use_container_width=True,
            column_config={
                "총건수": st.column_config.NumberColumn(format="%d건"),
                "총수량": st.column_config.NumberColumn(format="%d권"),
            },
        )

        _po_options = []
        for _, _b in _po_batches.iterrows():
            _po_options.append(f"{_b['발주일시']} ({int(_b['총건수'])}건, {int(_b['총수량'])}권)")
        _po_sel = st.selectbox("배치 상세 조회", range(len(_po_options)),
                               format_func=lambda i: _po_options[i], key="t4_po_batch_sel")
        _po_sel_bid = _po_batches.iloc[_po_sel]["batch_id"]

        _po_detail = query_df("""
            SELECT
                p.book_title AS 도서명,
                p.isbn,
                p.publisher AS 출판사,
                COALESCE(p.distributor, '') AS 거래처,
                p.quantity AS 수량,
                p.shipment_box_id AS 묶음배송번호,
                a.account_name AS 계정
            FROM purchase_order_logs p
            JOIN accounts a ON p.account_id = a.id
            WHERE p.batch_id = :bid
            ORDER BY p.distributor NULLS LAST, p.book_title
        """, {"bid": _po_sel_bid})

        if not _po_detail.empty:
            _dist_summary = _po_detail.groupby("거래처").agg(
                건수=("도서명", "count"), 수량=("수량", "sum")
            ).reset_index()
            _dist_summary = _dist_summary[_dist_summary["거래처"] != ""]
            if not _dist_summary.empty:
                st.caption("거래처별: " + " / ".join(
                    f"{r['거래처']} {int(r['수량'])}권" for _, r in _dist_summary.iterrows()
                ))
            st.dataframe(
                _po_detail,
                hide_index=True, use_container_width=True,
                column_config={
                    "묶음배송번호": st.column_config.NumberColumn(format="%d"),
                    "도서명": st.column_config.TextColumn(width="large"),
                },
            )

    else:  # 주문 DB
        _search_kw = st.text_input(
            "🔍 검색 (수취인/상품명)",
            placeholder="김선희, 수능특강 등",
            key="t4_search_keyword",
        )

        _fc1, _fc2, _fc3 = st.columns([2, 2, 1])
        _today = date.today()
        with _fc1:
            _date_range = st.date_input("기간", value=[_today - timedelta(days=30), _today], key="t4_date_range")
        with _fc2:
            _status_options = list(STATUS_MAP.keys())
            _status_sel = st.multiselect("상태", _status_options,
                                         default=_status_options,
                                         format_func=lambda x: STATUS_MAP.get(x, x),
                                         key="t4_status_filter")
        with _fc3:
            _inc_canceled = st.checkbox("취소 포함", value=False, key="t4_inc_canceled")

        if len(_date_range) == 2:
            _start, _end = _date_range
        else:
            _start = _end = _date_range[0] if _date_range else _today

        _params = {"start_date": str(_start), "end_date": str(_end)}
        _where_extra = ""
        if _search_kw and _search_kw.strip():
            _where_extra = " AND (o.receiver_name LIKE :kw OR o.vendor_item_name LIKE :kw)"
            _params["kw"] = f"%{_search_kw.strip()}%"

        _orders_df = query_df(f"""
            SELECT
                a.account_name AS 계정,
                o.shipment_box_id AS 묶음배송번호,
                o.order_id AS 주문번호,
                o.receiver_name AS 수취인,
                o.vendor_item_name AS 상품명,
                o.shipping_count AS 수량,
                o.order_price AS 결제금액,
                o.status AS 상태코드,
                o.invoice_number AS 운송장번호,
                o.delivery_company_name AS 택배사,
                o.canceled AS 취소,
                CAST(o.ordered_at + INTERVAL '9 hours' AS timestamp) AS 주문일시
            FROM orders o
            JOIN accounts a ON o.account_id = a.id
            WHERE o.ordered_at >= :start_date
              AND o.ordered_at < CAST(:end_date AS date) + INTERVAL '1 day'
              {_where_extra}
            ORDER BY o.ordered_at DESC
            LIMIT 500
        """, _params)

        if _orders_df.empty:
            st.info("해당 기간에 주문 데이터가 없습니다.")
        else:
            if _status_sel:
                _orders_df = _orders_df[_orders_df["상태코드"].isin(_status_sel)]
            if not _inc_canceled:
                _orders_df = _orders_df[~_orders_df["취소"]]

            if _orders_df.empty:
                st.info("필터 조건에 맞는 데이터가 없습니다.")
            else:
                _orders_df["상태"] = _orders_df["상태코드"].map(STATUS_MAP).fillna(_orders_df["상태코드"])
                _status_pivot = _orders_df.groupby(["계정", "상태"]).size().unstack(fill_value=0).reset_index()
                st.dataframe(_status_pivot, hide_index=True, use_container_width=True)

                _orders_df["주문일시"] = pd.to_datetime(_orders_df["주문일시"]).dt.strftime("%m/%d %H:%M")
                _display_cols = ["주문일시", "계정", "묶음배송번호", "주문번호", "수취인", "상품명", "수량", "결제금액", "상태", "운송장번호"]
                if _inc_canceled:
                    _display_cols.append("취소")
                st.caption(f"조회 결과: {len(_orders_df)}건" + (" (최대 500건)" if len(_orders_df) == 500 else ""))
                st.dataframe(
                    _orders_df[_display_cols],
                    hide_index=True, use_container_width=True,
                    column_config={
                        "묶음배송번호": st.column_config.NumberColumn(format="%d"),
                        "주문번호": st.column_config.NumberColumn(format="%d"),
                        "결제금액": st.column_config.NumberColumn(format="%d원"),
                        "상품명": st.column_config.TextColumn(width="large"),
                    },
                )


def _render_cancel_section(accounts_df, account_names, _accept_all, _instruct_live):
    """주문 취소 UI"""
    st.caption("ACCEPT/INSTRUCT 상태의 주문을 취소합니다.")

    _cancel_acct = st.selectbox("취소할 계정", account_names, key="t1_cancel_acct")
    _cancel_acct_row = None
    if _cancel_acct and not accounts_df.empty:
        _mask = accounts_df["account_name"] == _cancel_acct
        if _mask.any():
            _cancel_acct_row = accounts_df[_mask].iloc[0]

    if _cancel_acct_row is not None:
        _cancel_account_id = int(_cancel_acct_row["id"])
        _cancel_client = create_wing_client(_cancel_acct_row)

        _cancel_frames = []
        if not _accept_all.empty:
            _cancel_frames.append(_accept_all)
        if not _instruct_live.empty:
            _cancel_frames.append(_instruct_live)
        _cancel_all = pd.concat(_cancel_frames, ignore_index=True) if _cancel_frames else pd.DataFrame()
        _cancelable = pd.DataFrame()
        if not _cancel_all.empty:
            _cancel_acct_df = _cancel_all[
                (_cancel_all["_account_id"] == _cancel_account_id) & (~_cancel_all["취소"])
            ].copy()
            if not _cancel_acct_df.empty:
                _cancelable = _cancel_acct_df.rename(columns={"_vendor_item_id": "옵션ID"})[
                    ["주문번호", "옵션ID", "상품명", "수량", "결제금액", "상태", "주문일"]
                ].copy()

        if _cancelable.empty:
            st.info(f"[{_cancel_acct}] 취소 가능한 주문이 없습니다.")
        else:
            _cancelable_display = _cancelable.copy()
            _cancelable_display["상태"] = _cancelable_display["상태"].map(lambda x: STATUS_MAP.get(x, x))
            st.dataframe(_cancelable_display, use_container_width=True, hide_index=True)

            _cancel_reasons = {
                "SOLD_OUT": "재고 소진",
                "PRICE_ERROR": "가격 오류",
                "PRODUCT_ERROR": "상품 정보 오류",
                "OTHER": "기타 사유",
            }
            _sel_reason = st.selectbox("취소 사유", list(_cancel_reasons.keys()),
                                        format_func=lambda x: _cancel_reasons[x],
                                        key="t1_cancel_reason")
            _cancel_detail = st.text_input("상세 사유", value=_cancel_reasons[_sel_reason], key="t1_cancel_detail")

            _confirm_cancel = st.checkbox(
                f"{len(_cancelable)}건을 정말 취소하시겠습니까? (되돌릴 수 없음)",
                key="t1_cancel_confirm",
            )
            if _confirm_cancel:
                if st.button(f"주문 취소 ({len(_cancelable)}건)", type="secondary", key="t1_btn_cancel"):
                    if _cancel_client:
                        try:
                            _cancel_groups = _cancelable.groupby("주문번호")
                            _cancel_count = 0
                            for _oid, _group in _cancel_groups:
                                _vids = [int(x) for x in _group["옵션ID"].tolist() if pd.notna(x)]
                                _cnts = [int(x) for x in _group["수량"].tolist()]
                                if _vids:
                                    _cancel_client.cancel_order(
                                        order_id=int(_oid),
                                        vendor_item_ids=_vids,
                                        receipt_counts=_cnts,
                                        cancel_reason_category=_sel_reason,
                                        cancel_reason=_cancel_detail,
                                    )
                                    _cancel_count += len(_vids)
                            st.toast(f"취소 요청 완료: {_cancel_count}건", icon="✅")
                            clear_order_caches()
                            st.rerun()
                        except CoupangWingError as e:
                            st.error(f"API 오류: {e}")
                    else:
                        st.error("WING API 클라이언트를 생성할 수 없습니다.")


def _enrich_purchase_order_data(orders_df):
    """주문 DataFrame에 도서명/ISBN/출판사/거래처 enrichment (발주서 + 배송리스트 공용)"""
    import re as _re

    _pub_list = query_df_cached("SELECT name FROM publishers WHERE is_active = true ORDER BY LENGTH(name) DESC")
    _pub_names = _pub_list["name"].tolist() if not _pub_list.empty else []

    def _match_pub(row):
        result = match_publisher_from_text(str(row.get("옵션명") or ""), _pub_names)
        if not result:
            result = match_publisher_from_text(str(row.get("상품명") or ""), _pub_names)
        return result

    _isbn_lookup = query_df_cached("""
        SELECT l.coupang_product_id, l.isbn as isbn,
               b.title as db_title, l.product_name as listing_name
        FROM listings l
        LEFT JOIN books b ON l.isbn = b.isbn AND l.isbn IS NOT NULL AND l.isbn != ''
        WHERE l.coupang_product_id IS NOT NULL
    """)
    _isbn_map = {}
    if not _isbn_lookup.empty:
        for _, _r in _isbn_lookup.iterrows():
            _isbn_map[str(_r["coupang_product_id"])] = {
                "isbn": str(_r["isbn"]) if pd.notna(_r["isbn"]) else "",
                "title": str(_r["db_title"]) if pd.notna(_r["db_title"]) else "",
                "listing_name": str(_r["listing_name"]) if pd.notna(_r["listing_name"]) else "",
            }

    _TITLE_RE_PATTERNS = [
        r'\s*[-–]\s*2\d{3}\s*개정\s*교육과정.*$',
        r'\s+2\d{3}\s*개정\s*교육과정.*$',
        r'\s*[-–]\s*202\d학년도\s*수능\s*연계.*$',
        r'\s*\(202\d년?\s*수능대비\).*$',
        r'\s*\(202\d학년도\s*수능대비\).*$',
        r'\s*:\s*202\d학년도\s*수능.*$',
        r'\s*:\s*슝슝.*$',
        r'\s*:\s*동영상\s*강의.*$',
        r'\s*:\s*유형의\s*완성.*$',
        r'\s*#.*$',
        r'\s+사은품증정\s+\S+.*$',
        r'\s+/\s*본교재.*$',
        r'\s+\d+rd\s+edition.*$',
        r'\(2\d{3}년용\)',
        r'\s+고등\s+한국교육방송공사.*$',
        r'\s+한국교육방송공사.*$',
        r'\s+고등학교\s*[123]학년.*$',
        r'\s+고등\s*[123]학년.*$',
    ]

    def _clean_title(title: str) -> str:
        if title.startswith("사은품+"):
            title = title[4:]
        if "," in title:
            title = title[:title.index(",")].strip()
        for pat in _TITLE_RE_PATTERNS:
            title = _re.sub(pat, "", title, flags=_re.IGNORECASE).strip()
        if " : " in title:
            parts = title.split(" : ")
            if len(parts[0]) >= 10:
                title = parts[0].strip()
        return title.strip()

    def _resolve_book_info(row):
        spid = str(row.get("_seller_product_id", ""))
        info = _isbn_map.get(spid, {})
        isbn = info.get("isbn", "")
        title = info.get("title", "")
        if not title:
            title = info.get("listing_name", "")
        if not title:
            title = str(row.get("상품명", "")).strip()
        if not title:
            title = str(row.get("옵션명", "")).strip()
        return pd.Series({"도서명": _clean_title(title), "ISBN": isbn})

    df = orders_df.copy()
    df[["도서명", "ISBN"]] = df.apply(_resolve_book_info, axis=1)
    df["출판사"] = df.apply(_match_pub, axis=1)
    df["거래처"] = df["출판사"].apply(resolve_distributor)
    return df


def _save_purchase_order_to_db(enriched_df):
    """enriched 주문 DataFrame을 purchase_order_logs에 저장. batch_id 반환."""
    from uuid import uuid4
    from dashboard.utils import engine as _eng

    _batch_id = str(uuid4())
    with _eng.connect() as _conn:
        for _, _r in enriched_df.iterrows():
            _conn.execute(
                PurchaseOrderLog.__table__.insert().values(
                    batch_id=_batch_id,
                    shipment_box_id=int(_r["묶음배송번호"]),
                    order_id=int(_r.get("주문번호", 0) or 0),
                    account_id=int(_r.get("_account_id", 0)),
                    vendor_item_id=int(_r.get("_vendor_item_id", 0) or 0),
                    book_title=str(_r.get("도서명", "")).strip()[:500],
                    isbn=str(_r.get("ISBN", "")).strip()[:50],
                    publisher=str(_r.get("출판사", "")).strip()[:100],
                    distributor=str(_r.get("거래처", "")).strip()[:50],
                    quantity=int(_r.get("수량", 1)),
                )
            )
        _conn.commit()
    return _batch_id


def _build_purchase_order_excel(agg_df, store_name="잉글리쉬존"):
    """거래처별 발주서 엑셀 생성. BytesIO 반환."""
    from openpyxl.styles import Font as _OXFont, Alignment as _OXAlign

    _xl_buf = io.BytesIO()
    with pd.ExcelWriter(_xl_buf, engine="openpyxl") as writer:
        _dist_order = ["백석", "강우사", "하람", "서부", "동아", "제일", "일신", "대성", "북전", "대원", "일반"]
        _all_dists = sorted(
            agg_df["거래처"].unique(),
            key=lambda d: _dist_order.index(d) if d in _dist_order else 99,
        )
        for _dname in _all_dists:
            _sdf = agg_df[agg_df["거래처"] == _dname][["도서명", "출판사", "주문수량"]].copy()
            if _sdf.empty:
                continue
            _sdf = _sdf.sort_values(["출판사", "도서명"]).reset_index(drop=True)
            _sdf["주문수량"] = _sdf["주문수량"].astype(int)
            _safe = _dname[:31].replace("/", "_").replace("\\", "_")
            _sdf.to_excel(writer, sheet_name=_safe, index=False, header=False, startrow=1)
            ws = writer.sheets[_safe]
            ws.merge_cells("A1:C1")
            _t = ws.cell(row=1, column=1)
            _t.value = f"{store_name} 주문"
            _t.font = _OXFont(name="맑은 고딕", size=11)
            _t.alignment = _OXAlign(horizontal="center", vertical="center")
            for _r in range(2, ws.max_row + 1):
                ws.cell(_r, 1).font = _OXFont(name="맑은 고딕", size=10)
                ws.cell(_r, 1).alignment = _OXAlign(horizontal="left", vertical="center")
                ws.cell(_r, 2).font = _OXFont(name="맑은 고딕", size=10)
                ws.cell(_r, 2).alignment = _OXAlign(horizontal="left", vertical="center")
                ws.cell(_r, 3).font = _OXFont(name="맑은 고딕", size=10)
                ws.cell(_r, 3).alignment = _OXAlign(horizontal="center", vertical="center")
            ws.column_dimensions["A"].width = 57.5
            ws.column_dimensions["B"].width = 9.0
            ws.column_dimensions["C"].width = 13.0
    _xl_buf.seek(0)
    return _xl_buf


def _render_purchase_order(instruct_all, accounts_df, key_prefix="t2"):
    """발주서 생성 — compact 레이아웃"""

    with st.expander(f"📋 발주서 ({len(instruct_all)}건)", expanded=False):
        if instruct_all.empty:
            st.info("발주서 대상 주문이 없습니다.")
        else:
            _dist_orders = _enrich_purchase_order_data(instruct_all)

            _isbn_found = _dist_orders["ISBN"].apply(lambda x: bool(x and str(x).strip())).sum()
            _isbn_missing = len(_dist_orders) - _isbn_found
            if _isbn_missing > 0:
                st.caption("ISBN 없음: {}/{}건 (삭제된 상품 또는 세트물은 정상)".format(_isbn_missing, len(_dist_orders)))

            _store_name = st.text_input(
                "가게명 (발주서 첫 줄에 표시)",
                value=st.session_state.get("order_store_name", "잉글리쉬존"),
                key=f"{key_prefix}_store_name",
                help="예: 잉글리쉬존, 북마트"
            )
            st.session_state["order_store_name"] = _store_name

            _dist_summary = _dist_orders.groupby("거래처").agg(
                건수=("도서명", "count"), 수량합계=("수량", "sum"), 금액합계=("결제금액", "sum"),
            ).reset_index().sort_values("건수", ascending=False)
            _dist_summary["금액합계"] = _dist_summary["금액합계"].apply(lambda x: f"{int(x):,}")
            st.dataframe(_dist_summary, hide_index=True, use_container_width=True)

            _dist_orders["_group_key"] = _dist_orders.apply(
                lambda r: r["ISBN"] if r.get("ISBN") else r["도서명"], axis=1
            )
            _agg = _dist_orders.groupby(["거래처", "출판사", "_group_key"]).agg(
                도서명=("도서명", "first"), ISBN=("ISBN", "first"), 주문수량=("수량", "sum"),
            ).reset_index().drop(columns=["_group_key"])
            _agg = _agg.sort_values(["거래처", "출판사", "도서명"])

            _xl_buf = _build_purchase_order_excel(_agg, _store_name)

            _ord_date_to_str = date.today().isoformat()
            if "주문일" in _dist_orders.columns:
                _dist_dates = _dist_orders["주문일"].dropna()
                if not _dist_dates.empty:
                    _ord_date_to_str = str(_dist_dates.max())
            _file_date = _ord_date_to_str.replace("-", "")[2:]

            st.download_button(
                "📥 발주서 다운로드",
                _xl_buf.getvalue(),
                file_name=f"쿠팡{_file_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{key_prefix}_dist_xlsx_dl",
                type="primary",
                use_container_width=True,
            )

            _dist_names_sorted = _dist_summary["거래처"].tolist()
            _dist_filter = st.multiselect(
                "거래처 필터", _dist_names_sorted,
                default=_dist_names_sorted, key=f"{key_prefix}_dist_filter",
            )
            _filtered_agg = _agg[_agg["거래처"].isin(_dist_filter)] if _dist_filter else _agg
            _show_agg = _filtered_agg[["거래처", "ISBN", "출판사", "도서명", "주문수량"]].copy()

            gb2 = GridOptionsBuilder.from_dataframe(_show_agg)
            gb2.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
            gb2.configure_default_column(resizable=True, sorteable=True, filterable=True)
            gb2.configure_column("도서명", width=350)
            gb2.configure_column("주문수량", width=80)
            AgGrid(_show_agg, gridOptions=gb2.build(), height=500, theme="streamlit", key=f"{key_prefix}_dist_grid")



def _render_geukdong_excel(instruct_all, accounts_df, key_prefix="t2"):
    """극동 엑셀 — compact 레이아웃"""
    with st.expander("📦 극동 엑셀", expanded=False):
        if instruct_all.empty:
            st.info("대상 주문이 없습니다.")
            return

        _gk_orders = instruct_all.copy()

        _gk_isbn_lookup = query_df_cached("""
            SELECT l.coupang_product_id,
                   l.isbn as "ISBN",
                   b.title as "DB도서명",
                   l.product_name as 리스팅도서명,
                   b.list_price as 정가,
                   COALESCE(b.author, '') as 저자,
                   b.year as 출판년도,
                   pub.name as 출판사,
                   pub.supply_rate as 공급률
            FROM listings l
            LEFT JOIN books b ON l.isbn = b.isbn AND l.isbn IS NOT NULL AND l.isbn != ''
            LEFT JOIN publishers pub ON b.publisher_id = pub.id
            WHERE l.coupang_product_id IS NOT NULL
        """)
        _gk_map = {}
        if not _gk_isbn_lookup.empty:
            for _, _r in _gk_isbn_lookup.iterrows():
                _gk_map[str(_r["coupang_product_id"])] = {
                    "ISBN": str(_r["ISBN"]) if pd.notna(_r["ISBN"]) else "",
                    "DB도서명": str(_r["DB도서명"]) if pd.notna(_r["DB도서명"]) else "",
                    "리스팅도서명": str(_r["리스팅도서명"]) if pd.notna(_r["리스팅도서명"]) else "",
                    "정가": _r["정가"] if pd.notna(_r["정가"]) else 0,
                    "저자": str(_r["저자"]) if pd.notna(_r["저자"]) else "",
                    "출판년도": _r["출판년도"] if pd.notna(_r["출판년도"]) else None,
                    "출판사": str(_r["출판사"]) if pd.notna(_r["출판사"]) else "",
                    "공급률": _r["공급률"] if pd.notna(_r["공급률"]) else None,
                }

        def _gk_enrich(row):
            info = _gk_map.get(str(row.get("_seller_product_id", "")), {})
            return pd.Series({
                "ISBN": info.get("ISBN", ""),
                "DB도서명": info.get("DB도서명", ""),
                "리스팅도서명": info.get("리스팅도서명", ""),
                "정가": info.get("정가", 0),
                "저자": info.get("저자", ""),
                "출판년도": info.get("출판년도", None),
                "출판사": info.get("출판사", ""),
                "공급률": info.get("공급률", None),
            })

        _gk_extra = _gk_orders.apply(_gk_enrich, axis=1)
        _gk_orders = pd.concat([_gk_orders, _gk_extra], axis=1)

        if _gk_orders.empty:
            st.info("극동 대상 주문이 없습니다.")
            return

        def _resolve_gk_title(r):
            if pd.notna(r.get("DB도서명")) and r["DB도서명"]:
                return str(r["DB도서명"]).strip()
            if pd.notna(r.get("리스팅도서명")) and r["리스팅도서명"]:
                return str(r["리스팅도서명"]).strip()
            return str(r["옵션명"]).strip()

        _gk_orders["도서명"] = _gk_orders.apply(_resolve_gk_title, axis=1)
        _gk_orders["ISBN_clean"] = _gk_orders["ISBN"].apply(lambda x: str(x).strip() if pd.notna(x) and x else "")

        _gk_orders["_key"] = _gk_orders.apply(lambda r: r["ISBN_clean"] if r["ISBN_clean"] else r["도서명"], axis=1)
        _gk_agg = _gk_orders.groupby("_key").agg(
            상품바코드=("ISBN_clean", "first"),
            상품명=("도서명", "first"),
            정가=("정가", "first"),
            수량=("수량", "sum"),
            공급률=("공급률", "first"),
            출판사=("출판사", "first"),
            저자=("저자", "first"),
            출판년도=("출판년도", "first"),
        ).reset_index(drop=True)

        _gk_total_amount = int(_gk_agg.apply(
            lambda r: (r["정가"] * r["공급률"] * r["수량"]) if pd.notna(r["공급률"]) and r["공급률"] and pd.notna(r["정가"]) else 0,
            axis=1
        ).sum()) if not _gk_agg.empty else 0
        _gk_k1, _gk_k2, _gk_k3 = st.columns(3)
        _gk_k1.metric("출고 품목", f"{len(_gk_agg)}종")
        _gk_k2.metric("출고 수량", f"{int(_gk_agg['수량'].sum())}권")
        _gk_k3.metric("총 금액", f"₩{fmt_krw_short(_gk_total_amount)}")

        _gk_show = _gk_agg[["상품바코드", "상품명", "수량", "출판사"]].copy()
        st.dataframe(_gk_show, hide_index=True, use_container_width=True)

        _gk_result = pd.DataFrame()
        _gk_result["NO."] = range(1, len(_gk_agg) + 1)
        _gk_result["상품바코드"] = _gk_agg["상품바코드"].values
        _gk_result["상품명"] = _gk_agg["상품명"].values
        _gk_result["#"] = ""
        _gk_result["정 가"] = _gk_agg["정가"].apply(lambda x: int(x) if pd.notna(x) else 0).values
        _gk_result["수 량"] = _gk_agg["수량"].values
        _gk_result["%"] = _gk_agg["공급률"].apply(lambda x: f"{x*100:.0f}" if pd.notna(x) and x else "").values
        _gk_result["단 가"] = _gk_agg.apply(
            lambda r: int(r["정가"] * r["공급률"]) if pd.notna(r["공급률"]) and r["공급률"] and pd.notna(r["정가"]) else (int(r["정가"]) if pd.notna(r["정가"]) else 0),
            axis=1
        ).values
        _gk_result["금 액"] = (_gk_result["단 가"] * _gk_result["수 량"]).values
        _gk_result[""] = ""
        _gk_result["출판사"] = _gk_agg["출판사"].apply(lambda x: str(x) if pd.notna(x) else "").values
        _gk_result["저자"] = _gk_agg["저자"].apply(lambda x: str(x) if pd.notna(x) else "").values
        _gk_result["출판년도"] = _gk_agg["출판년도"].apply(lambda x: str(int(x)) if pd.notna(x) else "").values

        _gk_buf = io.BytesIO()
        with pd.ExcelWriter(_gk_buf, engine="openpyxl") as writer:
            _gk_result.to_excel(writer, sheet_name="극동", index=False)
        _gk_buf.seek(0)

        st.download_button(
            f"극동 다운로드 ({len(_gk_agg)}종 / {int(_gk_agg['수량'].sum())}권)",
            _gk_buf.getvalue(),
            file_name=f"극동_{date.today().strftime('%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_gk_xlsx_dl",
            type="primary",
            use_container_width=True,
        )


def _render_delivery_list(instruct_all, accounts_df=None):
    """배송리스트 다운로드 — 메인 CTA"""
    if instruct_all.empty:
        st.info("상품준비중 주문이 없습니다.")
        return

    try:
        _dl_orders = instruct_all.copy()
        _acct_counts = _dl_orders.groupby("계정").size().reset_index(name="건수")
        st.dataframe(_acct_counts, hide_index=True)

        # ── 출고중지 사전 확인 ──
        if accounts_df is not None:
            if st.button("출고중지 사전 확인", key="t2_pre_ru_check",
                         help="배송리스트 다운로드 전 출고중지(RU) 요청이 있는지 미리 확인"):
                _ru_hits = []
                _dl_oids = set(int(x) for x in _dl_orders["주문번호"].unique() if pd.notna(x) and int(x) > 0)
                _t2_from = (date.today() - timedelta(days=14)).isoformat()
                _t2_to = date.today().isoformat()
                for _aid in _dl_orders["_account_id"].unique():
                    _acct_row = accounts_df[accounts_df["id"] == int(_aid)]
                    if _acct_row.empty:
                        continue
                    _acct_row = _acct_row.iloc[0]
                    _client = create_wing_client(_acct_row)
                    if not _client:
                        continue
                    try:
                        _reqs = _client.get_all_return_requests(_t2_from, _t2_to, status="RU")
                        for _req in _reqs:
                            _oid = _req.get("orderId")
                            if _oid and int(_oid) in _dl_oids:
                                _ru_hits.append({
                                    "계정": _acct_row["account_name"],
                                    "주문번호": _oid,
                                    "사유": _req.get("cancelReason", ""),
                                })
                    except Exception as e:
                        st.warning(f"[{_acct_row['account_name']}] RU 조회 실패: {e}")
                if _ru_hits:
                    st.warning(f"출고중지 요청 {len(_ru_hits)}건 — 해당 주문은 송장 등록 시 자동 제외됩니다.")
                    st.dataframe(pd.DataFrame(_ru_hits), hide_index=True, use_container_width=True)
                else:
                    st.success("출고중지 요청 없음 — 안전하게 배송리스트를 다운로드하세요.")

        # 공유 함수로 엑셀 생성
        _xl_bytes, _dl_df = build_delivery_excel_bytes(_dl_orders, sort_and_color=True)

        # 세션에 저장 (송장 매칭용)
        st.session_state["_delivery_list_df"] = _dl_df.copy()

        # ── 배송리스트 DB 기록 (엑셀 생성 시점에 바로 저장) ──
        _dl_batch_key = "_dl_batch_saved_boxes"
        _dl_box_set = set(int(b) for b in _dl_df["묶음배송번호"].unique())
        if st.session_state.get(_dl_batch_key) != _dl_box_set:
            try:
                from uuid import uuid4
                from sqlalchemy.dialects.postgresql import insert as pg_insert
                from dashboard.utils import engine as _eng

                _batch_id = str(uuid4())
                _seen_boxes = {}
                _box_seq = 0
                for _, _r in _dl_df.iterrows():
                    _box_id = int(_r["묶음배송번호"])
                    if _box_id not in _seen_boxes:
                        _box_seq += 1
                        _seen_boxes[_box_id] = {
                            "shipment_box_id": _box_id,
                            "account_id": int(_r["_account_id"]),
                            "order_id": int(_r.get("주문번호", 0) or 0),
                            "vendor_item_id": int(_r.get("_vendor_item_id", 0) or 0),
                            "receiver_name": str(_r.get("수취인이름", "")).strip(),
                            "buyer_name": str(_r.get("구매자", "")).strip(),
                            "seq_no": _box_seq,
                            "batch_id": _batch_id,
                        }

                with _eng.connect() as _conn:
                    for _vals in _seen_boxes.values():
                        _stmt = pg_insert(DeliveryListLog.__table__).values(**_vals)
                        _stmt = _stmt.on_conflict_do_update(
                            index_elements=["shipment_box_id"],
                            set_={
                                "account_id": _vals["account_id"],
                                "order_id": _vals["order_id"],
                                "vendor_item_id": _vals["vendor_item_id"],
                                "receiver_name": _vals["receiver_name"],
                                "buyer_name": _vals["buyer_name"],
                                "seq_no": _vals["seq_no"],
                                "batch_id": _batch_id,
                                "downloaded_at": datetime.utcnow(),
                                "registered": False,
                            },
                        )
                        _conn.execute(_stmt)
                    _conn.commit()
                st.session_state["_last_batch_id"] = _batch_id
                st.session_state[_dl_batch_key] = _dl_box_set
                logger.info(f"배송리스트 {len(_seen_boxes)}건 DB 기록 완료 (batch={_batch_id[:8]})")
            except Exception as e:
                logger.error(f"배송리스트 DB 기록 실패: {e}")
                st.error(f"배송리스트 DB 기록 실패: {e}")

        # ── 동일 수취인 합배송 방지 안내 ──
        _recv_groups = _dl_df.groupby("수취인이름")["묶음배송번호"].apply(
            lambda x: list(x.unique())
        )
        _multi_recv = _recv_groups[_recv_groups.apply(len) > 1]
        if not _multi_recv.empty:
            _suffix_names = [n for n in _multi_recv.index if n.endswith(")")]
            if _suffix_names:
                st.info(
                    f"동일 수취인 {len(_multi_recv)}명 — "
                    "한진 합배송 방지를 위해 수취인 이름에 구분자를 자동 추가했습니다."
                )

        # ── 이전 다운로드 이력 확인 (정보 표시, 차단 안 함) ──
        _current_boxes = list(int(b) for b in _dl_df["묶음배송번호"])
        try:
            _db = SessionLocal()
            _prev_logs = _db.query(
                DeliveryListLog.shipment_box_id,
                DeliveryListLog.downloaded_at,
            ).filter(
                DeliveryListLog.shipment_box_id.in_(_current_boxes),
                DeliveryListLog.registered == False,
            ).all()
            _db.close()
            _prev_map = {r[0]: r[1] for r in _prev_logs}
        except Exception as e:
            logger.warning(f"배송리스트 이력 체크 실패: {e}")
            _prev_map = {}
        if _prev_map:
            _overlap_detail = _dl_df[_dl_df["묶음배송번호"].isin(_prev_map.keys())][
                ["번호", "묶음배송번호", "주문번호", "수취인이름", "등록상품명"]
            ].copy()
            _overlap_detail["이전 다운로드"] = _overlap_detail["묶음배송번호"].map(
                lambda x: _prev_map.get(int(x))
            ).apply(lambda dt: (dt + timedelta(hours=9)).strftime("%m/%d %H:%M") if dt else "")
            st.warning(
                f"다운로드 후 송장 미등록 주문 {len(_prev_map)}건 — "
                "한진에 이미 올렸다면 **중복 송장**에 주의하세요."
            )
            with st.expander(f"미등록 상세 ({len(_prev_map)}건)"):
                st.dataframe(_overlap_detail, hide_index=True, use_container_width=True)
    except Exception as e:
        st.error(f"배송리스트 생성 오류: {e}")
        logger.exception("배송리스트 생성 오류")
        return

    # 책별 픽킹 요약
    _pick_summary = (
        _dl_df.groupby("등록상품명")
        .agg(건수=("묶음배송번호", "count"), 총수량=("구매수(수량)", "sum"))
        .sort_index()
        .reset_index()
    )
    _pick_summary.columns = ["도서명", "주문건수", "총수량"]
    with st.expander(f"책별 픽킹 요약 ({len(_pick_summary)}종)", expanded=True):
        st.dataframe(_pick_summary, hide_index=True, use_container_width=True,
                     column_config={"도서명": st.column_config.TextColumn(width="large")})

    if st.download_button(
        f"배송리스트 다운로드 ({len(_dl_orders)}건)",
        _xl_bytes,
        file_name=f"DeliveryList({date.today().isoformat()})_통합.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="t2_dl_delivery_list",
        type="primary",
        use_container_width=True,
    ):
        st.session_state["_step_delivery_list"] = True

        # 발주 스냅샷 자동 저장
        try:
            _enriched = _enrich_purchase_order_data(_dl_orders)
            _po_batch = _save_purchase_order_to_db(_enriched)
        except Exception as e:
            logger.warning(f"발주 스냅샷 저장 실패: {e}")

    st.caption("Sheet1: 한진택배 업로드용 (책 순 정렬) | Sheet2: 픽킹리스트")


def _render_hanjin_nfocus():
    """한진 N-Focus — 복귀 컨텍스트 포함"""
    st.markdown("### 한진 N-Focus")
    st.caption("배송리스트 엑셀 → N-Focus 업로드 → 오류체크 → 출력 → 출력자료등록 엑셀 다운로드")

    # 작업 순서 안내
    st.markdown("""
    1. **업로드** — 배송리스트 엑셀을 N-Focus에 업로드
    2. **오류체크** — 오류 없으면 다음 단계
    3. **출력** — 운송장 출력
    4. **출력자료등록** — 엑셀 다운로드 (③ 송장등록에서 사용)
    """)

    _hc1, _hc2 = st.columns([3, 1])
    with _hc1:
        st.link_button(
            "한진 N-Focus 열기",
            "https://focus.hanjin.com/release/listup",
            type="primary", use_container_width=True,
        )
    with _hc2:
        if st.button("완료 →", key="t2_hanjin_done_btn", use_container_width=True, type="secondary"):
            st.session_state["_step_hanjin"] = True
            st.session_state["_t2_current_step"] = 2  # 자동으로 송장등록 단계로
            st.rerun()

    # 완료 상태 표시
    if st.session_state.get("_step_hanjin", False):
        st.success("한진 작업 완료 — 출력자료등록 엑셀을 다운로드했다면 ③ 송장등록 단계로 이동하세요.")


def _render_invoice_upload(instruct_all, accounts_df):
    """송장 등록 — 한진 출력자료 업로드 → 자동 매칭 → 쿠팡 등록"""
    st.caption("한진 출력자료등록 엑셀을 업로드하면 자동으로 배송리스트와 매칭하여 쿠팡에 송장을 등록합니다.")

    _inv_file = st.file_uploader("한진 출력자료등록 엑셀 (운송장번호 포함)", type=["xlsx", "xls"], key="t2_inv_file")
    if _inv_file is None:
        return

    try:
        _inv_df = pd.read_excel(_inv_file)
    except Exception as e:
        st.error(f"엑셀 파일 읽기 오류: {e}")
        return

    # ── 1. 배치 자동 감지 ──
    _batches = list_batches()
    _batch_df = None
    if _batches is not None and not _batches.empty:
        _auto_idx = 0
        if len(_batches) > 1:
            _batch_options = []
            for _, _b in _batches.iterrows():
                _dl_dt = _b["downloaded_at"]
                _dt_str = _dl_dt.strftime("%m/%d %H:%M") if hasattr(_dl_dt, "strftime") else str(_dl_dt)[:16]
                _batch_options.append(f"{_dt_str} ({_b['count']}건)")
            _auto_idx = st.selectbox(
                "배치", range(len(_batch_options)),
                format_func=lambda i: _batch_options[i],
                key="t2_batch_select",
            )
        _sel_batch_id = _batches.iloc[_auto_idx]["batch_id"]
        _batch_df = load_latest_batch(batch_id=_sel_batch_id)
        if _batch_df is not None:
            _dl_at = _batch_df["_downloaded_at"].iloc[0]
            _dl_date = _dl_at.strftime("%m/%d %H:%M") if hasattr(_dl_at, "strftime") else str(_dl_at)[:16]
            st.caption(f"배치: {_dl_date} ({len(_batch_df)}건)")
    else:
        st.caption("DB 배치 없음 — 이름 매칭(fallback) 사용")

    # ── 2. 매칭 ──
    _matched_df, _method = match_invoices(_inv_df, _batch_df)

    if _matched_df is None or _matched_df.empty:
        st.warning("매칭 결과가 없습니다. 엑셀 형식을 확인하세요. ('순번/운송장번호' 또는 '묶음배송번호/주문번호/운송장번호' 필요)")
        return

    st.success(f"매칭 완료: {len(_matched_df)}건 ({_method})")

    # ── 3. 등록 가능 여부 분류 ──
    _result = check_registerable(_matched_df, instruct_all, _batch_df)
    _reg_df = _result["registerable"]
    _shipped_df = _result["already_shipped"]
    _summary = _result["summary"]

    st.info(f"등록 가능: {_summary['등록가능']}건 / 이미 출고: {_summary['이미출고']}건")

    if not _shipped_df.empty:
        with st.expander(f"이미 출고된 주문 ({_summary['이미출고']}건)"):
            st.dataframe(_shipped_df[["묶음배송번호", "주문번호", "운송장번호"]].drop_duplicates(), hide_index=True)

    if _reg_df.empty:
        st.info("등록할 송장이 없습니다.")
        return

    # 계정별 건수
    _acct_id_map = dict(zip(accounts_df["id"].astype(int), accounts_df["account_name"]))
    _reg_df["계정"] = _reg_df["_account_id"].astype(int).map(_acct_id_map)
    _acct_summary = _reg_df.groupby("계정").size().reset_index(name="송장건수")
    st.dataframe(_acct_summary, hide_index=True)

    # ── 4. 출고중지요청 체크 ──
    _stop_orders, _safe_df = _check_stop_shipment_requests(_reg_df, accounts_df)

    if not _stop_orders.empty:
        st.warning(f"출고중지요청 {len(_stop_orders)}건 감지 — 해당 주문은 제외됩니다.")
        with st.expander(f"출고중지요청 상세 ({len(_stop_orders)}건)", expanded=True):
            _stop_display = _stop_orders[["계정", "주문번호", "묶음배송번호", "_receipt_id", "_cancel_count", "_cancel_reason"]].copy()
            _stop_display.columns = ["계정", "주문번호", "묶음배송번호", "접수번호", "취소수량", "취소사유"]
            st.dataframe(_stop_display, hide_index=True)

            if st.button("출고중지완료 처리 (미출고 확인)", key="t2_btn_stop_shipment", type="secondary"):
                _stop_ok = 0
                _stop_fail = 0
                for _aid, _sg in _stop_orders.groupby("_account_id"):
                    _aid = int(_aid)
                    _acct_row = accounts_df[accounts_df["id"] == _aid]
                    if _acct_row.empty:
                        continue
                    _acct_row = _acct_row.iloc[0]
                    _client = create_wing_client(_acct_row)
                    if not _client:
                        continue
                    for _, _sr in _sg.iterrows():
                        try:
                            _client.stop_shipment(int(_sr["_receipt_id"]), int(_sr["_cancel_count"]))
                            _stop_ok += 1
                        except Exception as e:
                            _stop_fail += 1
                            st.error(f"[{_acct_row['account_name']}] 접수번호 {_sr['_receipt_id']}: {e}")
                if _stop_ok > 0:
                    st.toast(f"출고중지완료: {_stop_ok}건 성공" + (f", {_stop_fail}건 실패" if _stop_fail else ""), icon="✅")
                    clear_order_caches()

    if _safe_df.empty:
        if not _stop_orders.empty:
            st.info("출고중지 건을 제외하면 등록할 송장이 없습니다.")
        return

    # ── 5. 묶음배송 중복 제거 + 등록 ──
    _before_dedup = len(_safe_df)
    _safe_df = _safe_df.drop_duplicates(subset=["묶음배송번호"], keep="first").copy()
    if _before_dedup != len(_safe_df):
        st.caption(f"묶음배송 중복 제거: {_before_dedup}행 → {len(_safe_df)}건")

    if st.button(f"전체 송장 등록 ({len(_safe_df)}건)", key="t2_btn_bulk_invoice", type="primary"):
        _total_success = 0
        _total_fail = 0
        _success_items = []

        for _aid, _grp in _safe_df.groupby("_account_id"):
            _aid = int(_aid)
            _acct_row = accounts_df[accounts_df["id"] == _aid]
            if _acct_row.empty:
                continue
            _acct_row = _acct_row.iloc[0]
            _client = create_wing_client(_acct_row)
            if not _client:
                st.error(f"[{_acct_row['account_name']}] API 클라이언트 생성 실패")
                continue

            _inv_data = []
            for _, _r in _grp.iterrows():
                _vid = int(_r["_vendor_item_id"]) if pd.notna(_r.get("_vendor_item_id")) else 0
                _inv_data.append({
                    "shipmentBoxId": int(_r["묶음배송번호"]),
                    "orderId": int(_r["주문번호"]),
                    "vendorItemId": _vid,
                    "deliveryCompanyCode": "HANJIN",
                    "invoiceNumber": str(_r["운송장번호"]).strip(),
                    "splitShipping": False,
                    "preSplitShipped": False,
                    "estimatedShippingDate": "",
                })

            try:
                _result = _client.upload_invoice(_inv_data)
                _s_cnt = 0
                _f_cnt = 0
                if isinstance(_result, dict) and "data" in _result:
                    for _ri in _result["data"].get("responseList", []):
                        if _ri.get("succeed"):
                            _s_cnt += 1
                            _success_items.append({
                                "shipmentBoxId": _ri.get("shipmentBoxId"),
                                "invoiceNumber": str(_ri.get("invoiceNumber", "")),
                                "deliveryCompanyCode": "HANJIN",
                            })
                        else:
                            _f_cnt += 1
                            st.error(f"  [{_acct_row['account_name']}] {_ri.get('shipmentBoxId')}: {_ri.get('resultMessage', '')}")
                else:
                    _s_cnt = len(_inv_data)
                    _success_items.extend(_inv_data)
                _total_success += _s_cnt
                _total_fail += _f_cnt
                st.info(f"[{_acct_row['account_name']}] 성공 {_s_cnt}건" + (f", 실패 {_f_cnt}건" if _f_cnt else ""))
            except Exception as e:
                _total_fail += len(_inv_data)
                st.error(f"[{_acct_row['account_name']}] API 오류: {e}")

        if _success_items:
            _update_orders_status(_success_items)
            try:
                _success_box_ids = [int(s["shipmentBoxId"]) for s in _success_items]
                from dashboard.utils import engine as _eng
                from sqlalchemy import text as _sa_text
                with _eng.connect() as _conn:
                    _conn.execute(
                        _sa_text("UPDATE delivery_list_logs SET registered = TRUE WHERE shipment_box_id = ANY(:ids)"),
                        {"ids": _success_box_ids},
                    )
                    _conn.commit()
            except Exception as e:
                logger.warning(f"배송리스트 로그 등록마킹 실패: {e}")

        if _total_success > 0:
            st.session_state["_step_invoice"] = True
            _msg = f"송장 등록 완료: 총 {_total_success}건 성공"
            if _total_fail:
                _msg += f", {_total_fail}건 실패"
            st.session_state["_flash_messages"] = [("success", _msg)]
            clear_order_caches()
            st.session_state.pop("_delivery_list_df", None)
            st.rerun()
        elif _total_fail > 0:
            st.error(f"전체 실패: {_total_fail}건 — 원인 확인 후 재시도하세요.")


def _check_stop_shipment_requests(matched_df, accounts_df):
    """송장 등록 전 출고중지요청(RU) 체크 — 해당 주문 분리 반환.

    Returns:
        (stop_orders_df, safe_df): 출고중지 대상 / 안전한 송장 등록 대상
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    _today = date.today()
    _from = (_today - timedelta(days=14)).isoformat()
    _to = _today.isoformat()

    _stop_order_ids = {}

    def _fetch_stops(acct_row, client):
        try:
            reqs = client.get_all_return_requests(_from, _to, status="RU")
            return acct_row, reqs
        except Exception as e:
            logger.warning(f"[{acct_row['account_name']}] 출고중지요청 조회 실패: {e}")
            return acct_row, []

    _acct_ids_in_matched = matched_df["_account_id"].astype(int).unique()
    _tasks = []
    for _aid in _acct_ids_in_matched:
        _acct_row = accounts_df[accounts_df["id"] == int(_aid)]
        if _acct_row.empty:
            continue
        _acct_row = _acct_row.iloc[0]
        _client = create_wing_client(_acct_row)
        if _client:
            _tasks.append((_acct_row, _client))

    if _tasks:
        with ThreadPoolExecutor(max_workers=min(len(_tasks), 10)) as pool:
            futures = [pool.submit(_fetch_stops, a, c) for a, c in _tasks]
            for f in as_completed(futures):
                acct_row, reqs = f.result()
                for req in reqs:
                    _oid = req.get("orderId")
                    if _oid:
                        _stop_order_ids[int(_oid)] = {
                            "receipt_id": req.get("receiptId"),
                            "cancel_count": req.get("cancelCountSum", 1),
                            "cancel_reason": req.get("cancelReason", ""),
                            "account_id": int(acct_row["id"]),
                            "account_name": acct_row["account_name"],
                        }

    if not _stop_order_ids:
        return pd.DataFrame(), matched_df

    _matched = matched_df.copy()
    _matched["_oid_int"] = _matched["주문번호"].astype(int)
    _is_stopped = _matched["_oid_int"].isin(_stop_order_ids.keys())

    _stop_df = _matched[_is_stopped].copy()
    _safe_df = _matched[~_is_stopped].copy()

    if not _stop_df.empty:
        _stop_df["_receipt_id"] = _stop_df["_oid_int"].map(lambda x: _stop_order_ids.get(x, {}).get("receipt_id", ""))
        _stop_df["_cancel_count"] = _stop_df["_oid_int"].map(lambda x: _stop_order_ids.get(x, {}).get("cancel_count", 1))
        _stop_df["_cancel_reason"] = _stop_df["_oid_int"].map(lambda x: _stop_order_ids.get(x, {}).get("cancel_reason", ""))
        _acct_id_map = dict(zip(accounts_df["id"].astype(int), accounts_df["account_name"]))
        _stop_df["계정"] = _stop_df["_account_id"].astype(int).map(_acct_id_map)

    _stop_df = _stop_df.drop(columns=["_oid_int"], errors="ignore")
    _safe_df = _safe_df.drop(columns=["_oid_int"], errors="ignore")

    return _stop_df, _safe_df

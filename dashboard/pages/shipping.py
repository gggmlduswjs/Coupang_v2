"""
배송/송장 관리 페이지
====================
STEP 1: 주문 확인 → STEP 2: 배송리스트 → STEP 3: 한진 송장 → STEP 4: 쿠팡 등록
"""
import io
import logging
import sys
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from st_aggrid import AgGrid, GridOptionsBuilder

from dashboard.utils import create_wing_client
from dashboard.services.order_data import (
    load_all_orders_live,
    get_instruct_orders,
    get_instruct_by_box,
    clear_order_caches,
    fmt_krw_short,
)
from dashboard.services.order_service import (
    load_hanjin_creds as _load_hanjin_creds,
    save_hanjin_creds as _save_hanjin_creds,
)

logger = logging.getLogger(__name__)


def render(selected_account, accounts_df, account_names):
    st.title("배송/송장 관리")

    # ── 공유 데이터 로드 (WING API 실시간) ──
    _all_orders = load_all_orders_live(accounts_df)
    _instruct_all = get_instruct_orders(_all_orders)
    _inst_by_box = get_instruct_by_box(_instruct_all)

    # ── 상단 컨트롤 ──
    _top_c1, _top_c2 = st.columns([2, 5])
    with _top_c1:
        if st.button("🔄 주문 새로고침", key="ship_btn_refresh", use_container_width=True,
                     help="WING API에서 실시간 주문 조회", type="primary"):
            clear_order_caches()
            st.rerun()
    with _top_c2:
        _last_synced = st.session_state.get("order_last_synced")
        if _last_synced:
            st.caption(f"마지막 조회: {_last_synced} (WING API 실시간)")

    st.caption("상품준비중 주문 배송 처리: 주문 확인 → 배송리스트 → 한진 송장 발급 → 쿠팡 송장 등록")

    # ── Step 워크플로우 표시 ──
    _s1, _s2, _s3, _s4 = st.columns(4)
    with _s1:
        st.markdown("**STEP 1** 주문 확인")
    with _s2:
        st.markdown("**STEP 2** 배송리스트")
    with _s3:
        st.markdown("**STEP 3** 한진 송장")
    with _s4:
        st.markdown("**STEP 4** 쿠팡 등록")

    st.divider()

    # ── STEP 1: 상품준비중 주문 확인 ──
    st.subheader("STEP 1: 상품준비중 주문 확인")

    _inst_total = len(_inst_by_box)
    _inst_amount = int(_inst_by_box["결제금액"].sum()) if not _inst_by_box.empty else 0

    _ik1, _ik2 = st.columns(2)
    _ik1.metric("상품준비중 주문", f"{_inst_total:,}건")
    _ik2.metric("총 금액", f"₩{fmt_krw_short(_inst_amount)}")

    if _inst_by_box.empty:
        st.info("상품준비중(INSTRUCT) 상태의 주문이 없습니다.")
    else:
        _inst_display = _inst_by_box[["계정", "묶음배송번호", "주문번호", "상품명", "수량", "결제금액", "주문일", "수취인"]].copy()
        _inst_display["결제금액"] = _inst_display["결제금액"].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "0")

        gb = GridOptionsBuilder.from_dataframe(_inst_display)
        gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
        gb.configure_default_column(resizable=True, sorteable=True, filterable=True)
        gb.configure_column("상품명", width=350)
        grid_opts = gb.build()
        AgGrid(_inst_display, gridOptions=grid_opts, height=400, theme="streamlit", key="ship_instruct_grid")

    st.divider()

    # ── STEP 2: 배송리스트 다운로드 ──
    st.subheader("STEP 2: 배송리스트 다운로드")

    if _instruct_all.empty:
        st.info("상품준비중 주문이 없습니다.")
    else:
        _dl_orders = _instruct_all.copy()

        # 계정별 건수 표시
        _acct_counts = _dl_orders.groupby("계정").size().reset_index(name="건수")
        st.dataframe(_acct_counts, hide_index=True)

        # 쿠팡 DeliveryList 형식 (40컬럼) 생성
        _dl_rows = []
        for _idx, (_i, _row) in enumerate(_dl_orders.iterrows(), 1):
            _dl_rows.append({
                "번호": _idx,
                "묶음배송번호": int(_row["묶음배송번호"]),
                "주문번호": int(_row["주문번호"]),
                "택배사": "한진택배",
                "운송장번호": "",
                "분리배송 Y/N": "분리배송가능" if _row.get("분리배송가능") else "분리배송불가",
                "분리배송 출고예정일": "",
                "주문시 출고예정일": _row.get("주문시출고예정일", ""),
                "출고일(발송일)": "",
                "주문일": _row.get("주문일시", _row.get("주문일", "")),
                "등록상품명": str(_row.get("상품명") or ""),
                "등록옵션명": _row.get("옵션명", ""),
                "노출상품명(옵션명)": f"{_row.get('상품명', '')}, {_row.get('옵션명', '')}",
                "노출상품ID": str(_row.get("_seller_product_id", "")),
                "옵션ID": str(_row.get("_vendor_item_id", "")),
                "최초등록등록상품명/옵션명": _row.get("최초등록상품옵션명", ""),
                "업체상품코드": _row.get("업체상품코드", ""),
                "바코드": "",
                "결제액": int(_row.get("결제금액", 0)),
                "배송비구분": _row.get("배송비구분", ""),
                "배송비": _row.get("배송비", 0),
                "도서산간 추가배송비": int(_row.get("도서산간추가배송비", 0)),
                "구매수(수량)": int(_row.get("수량", 0)),
                "옵션판매가(판매단가)": int(_row.get("판매단가", 0) or _row.get("결제금액", 0)),
                "구매자": _row.get("구매자", ""),
                "구매자전화번호": _row.get("구매자전화번호", ""),
                "수취인이름": _row.get("수취인", ""),
                "수취인전화번호": _row.get("수취인전화번호", ""),
                "우편번호": _row.get("우편번호", ""),
                "수취인 주소": _row.get("수취인주소", ""),
                "배송메세지": _row.get("배송메세지", ""),
                "상품별 추가메시지": "",
                "주문자 추가메시지": "",
                "배송완료일": "",
                "구매확정일자": "",
                "개인통관번호(PCCC)": _row.get("개인통관번호", ""),
                "통관용수취인전화번호": _row.get("통관용전화번호", ""),
                "기타": "",
                "결제위치": _row.get("결제위치", ""),
                "배송유형": "판매자 배송",
            })

        _dl_df = pd.DataFrame(_dl_rows)

        # NaN 처리 — 정렬 오작동 방지
        _dl_df["등록상품명"] = _dl_df["등록상품명"].fillna("").astype(str)

        # 책 이름 순 정렬 (같은 책끼리 묶임) → 번호 재부여
        _dl_df = _dl_df.sort_values(["등록상품명", "묶음배송번호"]).reset_index(drop=True)
        _dl_df["번호"] = range(1, len(_dl_df) + 1)

        # ── 책별 픽킹 요약 표시 (가나다 순) ──
        _pick_summary = (
            _dl_df.groupby("등록상품명")
            .agg(건수=("묶음배송번호", "count"), 총수량=("구매수(수량)", "sum"))
            .sort_index()
            .reset_index()
        )
        _pick_summary.columns = ["도서명", "주문건수", "총수량"]
        with st.expander(f"📚 책별 픽킹 요약 ({len(_pick_summary)}종)", expanded=True):
            st.dataframe(_pick_summary, hide_index=True, use_container_width=True,
                         column_config={"도서명": st.column_config.TextColumn(width="large")})

        # 엑셀 생성 — Sheet1: 배송리스트(책 순 정렬), Sheet2: 픽킹리스트
        _dl_buf = io.BytesIO()
        with pd.ExcelWriter(_dl_buf, engine="openpyxl") as writer:
            # Sheet1: 한진 업로드용 배송리스트
            _dl_df.to_excel(writer, sheet_name="Delivery", index=False)
            ws = writer.sheets["Delivery"]
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill, Font
            # 텍스트 포맷 (지수 표기 방지)
            for col_name in ["묶음배송번호", "주문번호", "노출상품ID", "옵션ID"]:
                if col_name in _dl_df.columns:
                    col_idx = _dl_df.columns.get_loc(col_name)
                    col_letter = get_column_letter(col_idx + 1)
                    for row_idx in range(2, len(_dl_df) + 2):
                        cell = ws[f"{col_letter}{row_idx}"]
                        cell.value = str(int(cell.value)) if cell.value is not None else ""
                        cell.number_format = "@"
            # 같은 책 첫 행에 색상 표시 (눈에 띄게)
            _prev_book = None
            _fill = PatternFill(start_color="D9E8FB", end_color="D9E8FB", fill_type="solid")
            for row_idx, book in enumerate(_dl_df["등록상품명"], start=2):
                if book != _prev_book:
                    for c in range(1, len(_dl_df.columns) + 1):
                        ws.cell(row=row_idx, column=c).fill = _fill
                    _prev_book = book

            # Sheet2: 픽킹 리스트 (출력용)
            _pick_summary.to_excel(writer, sheet_name="픽킹리스트", index=False)
            ws2 = writer.sheets["픽킹리스트"]
            ws2.column_dimensions["A"].width = 60
            ws2.column_dimensions["B"].width = 12
            ws2.column_dimensions["C"].width = 12

        _dl_buf.seek(0)

        st.download_button(
            f"📦 배송리스트 다운로드 ({len(_dl_orders)}건, 책별 정렬)",
            _dl_buf.getvalue(),
            file_name=f"DeliveryList({date.today().isoformat()})_통합.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ship_dl_delivery_list",
            type="primary",
            use_container_width=True,
        )
        st.caption("Sheet1: 한진택배 업로드용 (책 순 정렬) | Sheet2: 픽킹리스트")

    st.divider()

    # ── STEP 3: 한진 N-Focus 송장 발급 ──
    st.subheader("STEP 3: 한진 N-Focus 송장 발급")
    st.caption("STEP 2에서 다운로드한 배송리스트 엑셀을 업로드하면 N-Focus에서 송장을 자동 발급합니다.")

    _hanjin_creds = _load_hanjin_creds()
    if not _hanjin_creds.get("user_id"):
        with st.expander("한진 N-Focus 로그인 설정", expanded=True):
            _hj_id = st.text_input("N-Focus 아이디", key="ship_hanjin_id")
            _hj_pw = st.text_input("N-Focus 비밀번호", type="password", key="ship_hanjin_pw")
            if st.button("저장", key="ship_hanjin_save"):
                if _hj_id and _hj_pw:
                    _save_hanjin_creds(_hj_id, _hj_pw)
                    st.success("한진 크레덴셜 저장 완료")
                    st.rerun()
                else:
                    st.warning("아이디와 비밀번호를 모두 입력하세요.")
    else:
        with st.expander(f"N-Focus 계정: {_hanjin_creds['user_id']}", expanded=False):
            _hj_id = st.text_input("N-Focus 아이디", value=_hanjin_creds.get("user_id", ""), key="ship_hanjin_id_edit")
            _hj_pw = st.text_input("N-Focus 비밀번호", type="password", key="ship_hanjin_pw_edit")
            if st.button("변경 저장", key="ship_hanjin_save_edit"):
                if _hj_id and _hj_pw:
                    _save_hanjin_creds(_hj_id, _hj_pw)
                    st.success("한진 크레덴셜 업데이트 완료")
                    st.rerun()

    _nfocus_file = st.file_uploader(
        "배송리스트 엑셀 (STEP 2에서 다운로드한 파일)",
        type=["xlsx", "xls"],
        key="ship_nfocus_upload",
    )

    _nfocus_disabled = (
        not _hanjin_creds.get("user_id")
        or _nfocus_file is None
        or st.session_state.get("nfocus_running", False)
    )

    if st.button(
        "한진 N-Focus 자동 처리",
        key="ship_btn_nfocus",
        type="primary",
        disabled=_nfocus_disabled,
        use_container_width=True,
    ):
        st.session_state["nfocus_running"] = True
        try:
            from operations.hanjin_nfocus import HanjinNFocusClient, HanjinNFocusError

            _hc = _load_hanjin_creds()
            with st.status("N-Focus 처리 중...", expanded=True) as status:
                with HanjinNFocusClient(
                    user_id=_hc["user_id"],
                    password=_hc["password"],
                    headless=False,
                ) as client:
                    _nf_result = client.process_full_workflow(
                        excel_bytes=_nfocus_file.getvalue(),
                        filename=_nfocus_file.name,
                        progress_callback=lambda msg: st.write(msg),
                    )

                if _nf_result["success"]:
                    status.update(label="N-Focus 처리 완료!", state="complete")
                    st.success(
                        f"정상 출력: {_nf_result['registered']}건"
                        + (f" / 오류: {_nf_result['error']}건" if _nf_result["error"] else "")
                    )
                    if _nf_result["error_details"]:
                        with st.expander("오류 상세"):
                            for _err in _nf_result["error_details"]:
                                st.warning(_err)
                    if _nf_result["invoice_excel"]:
                        st.download_button(
                            "운송장 엑셀 다운로드 → STEP 4에서 업로드",
                            _nf_result["invoice_excel"],
                            file_name=f"Invoice_{date.today().isoformat()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="ship_dl_nfocus_invoice",
                        )
                else:
                    status.update(label="N-Focus 처리 실패", state="error")
        except ImportError:
            st.error("playwright 미설치. `pip install playwright && playwright install chromium`")
        except Exception as e:
            st.error(f"N-Focus 오류: {e}")
        finally:
            st.session_state["nfocus_running"] = False

    st.divider()

    # ── STEP 4: 송장 엑셀 업로드 (쿠팡 등록) ──
    st.subheader("STEP 4: 쿠팡 송장 등록")
    st.caption("운송장번호가 포함된 엑셀을 업로드하면 각 계정별로 쿠팡에 자동 등록됩니다.")

    _inv_file = st.file_uploader("송장 엑셀 파일 (운송장번호 포함)", type=["xlsx", "xls"], key="ship_inv_file_upload")
    _inv_df = None

    if _inv_file is not None:
        try:
            _inv_df = pd.read_excel(_inv_file)
        except Exception as e:
            st.error(f"엑셀 파일 읽기 오류: {e}")

    if _inv_df is not None:
        try:
            # 컬럼명 확인
            _need_cols = ["묶음배송번호", "주문번호", "운송장번호"]
            _missing = [c for c in _need_cols if c not in _inv_df.columns]
            if _missing:
                st.error(f"필수 컬럼 누락: {_missing}")
            else:
                # 운송장번호가 있는 행만
                _inv_filled = _inv_df[_inv_df["운송장번호"].notna() & (_inv_df["운송장번호"] != "")].copy()

                if _inv_filled.empty:
                    st.warning("운송장번호가 입력된 행이 없습니다.")
                else:
                    st.success(f"송장번호 입력된 주문: {len(_inv_filled)}건")

                    # 옵션ID 컬럼 확인
                    _has_option_id = "옵션ID" in _inv_filled.columns

                    # INSTRUCT 주문과 매칭하여 계정 정보 연결
                    _inv_merged = _inv_filled.copy()
                    if not _instruct_all.empty:
                        _match_cols = _instruct_all[["묶음배송번호", "_account_id", "_vendor_item_id", "주문번호"]].copy()
                        _match_cols["묶음배송번호"] = _match_cols["묶음배송번호"].astype(str)
                        _inv_merged["묶음배송번호"] = _inv_merged["묶음배송번호"].astype(str)
                        _inv_merged["주문번호"] = _inv_merged["주문번호"].astype(str)
                        _match_cols["주문번호"] = _match_cols["주문번호"].astype(str)

                        _inv_merged = _inv_merged.merge(
                            _match_cols.drop_duplicates(subset=["묶음배송번호", "주문번호"]),
                            on=["묶음배송번호", "주문번호"], how="left",
                        )

                    # 계정별로 분리하여 표시
                    if "_account_id" not in _inv_merged.columns:
                        st.error("상품준비중 주문과 매칭할 수 없습니다. 먼저 배송리스트를 다운로드하세요.")
                    else:
                        _matched = _inv_merged[_inv_merged["_account_id"].notna()]
                        _unmatched = _inv_merged[_inv_merged["_account_id"].isna()]

                        if not _unmatched.empty:
                            st.warning(f"매칭 안 된 주문: {len(_unmatched)}건 (이미 발송됐거나 취소된 주문)")

                        if _matched.empty:
                            st.info("등록할 송장이 없습니다.")
                        else:
                            # 계정별 건수
                            _acct_id_map = dict(zip(accounts_df["id"].astype(int), accounts_df["account_name"]))
                            _matched["계정"] = _matched["_account_id"].astype(int).map(_acct_id_map)
                            _acct_summary = _matched.groupby("계정").size().reset_index(name="송장건수")
                            st.dataframe(_acct_summary, hide_index=True)

                            if st.button(f"전체 송장 등록 ({len(_matched)}건)", key="ship_btn_bulk_invoice", type="primary"):
                                _total_success = 0
                                _total_fail = 0

                                for _aid, _grp in _matched.groupby("_account_id"):
                                    _aid = int(_aid)
                                    _acct_row = accounts_df[accounts_df["id"] == _aid]
                                    if _acct_row.empty:
                                        continue
                                    _acct_row = _acct_row.iloc[0]
                                    _client = create_wing_client(_acct_row)
                                    if not _client:
                                        st.error(f"[{_acct_row['account_name']}] API 클라이언트 생성 실패")
                                        continue

                                    _inv_data = []
                                    for _, _r in _grp.iterrows():
                                        _vid = int(_r["_vendor_item_id"]) if pd.notna(_r.get("_vendor_item_id")) else 0
                                        if not _vid and _has_option_id:
                                            _vid = int(_r["옵션ID"]) if pd.notna(_r.get("옵션ID")) else 0
                                        _inv_data.append({
                                            "shipmentBoxId": int(_r["묶음배송번호"]),
                                            "orderId": int(_r["주문번호"]),
                                            "vendorItemId": _vid,
                                            "deliveryCompanyCode": "HANJIN",
                                            "invoiceNumber": str(_r["운송장번호"]).strip(),
                                            "splitShipping": False,
                                            "preSplitShipped": False,
                                            "estimatedShippingDate": "",
                                        })

                                    try:
                                        _result = _client.upload_invoice(_inv_data)
                                        _s_cnt = 0
                                        _f_cnt = 0
                                        if isinstance(_result, dict) and "data" in _result:
                                            for _ri in _result["data"].get("responseList", []):
                                                if _ri.get("succeed"):
                                                    _s_cnt += 1
                                                else:
                                                    _f_cnt += 1
                                                    st.error(f"  [{_acct_row['account_name']}] {_ri.get('shipmentBoxId')}: {_ri.get('resultMessage', '')}")
                                        else:
                                            _s_cnt = len(_inv_data)
                                        _total_success += _s_cnt
                                        _total_fail += _f_cnt
                                        st.info(f"[{_acct_row['account_name']}] 성공 {_s_cnt}건" + (f", 실패 {_f_cnt}건" if _f_cnt else ""))
                                    except Exception as e:
                                        _total_fail += len(_inv_data)
                                        st.error(f"[{_acct_row['account_name']}] API 오류: {e}")

                                if _total_success > 0:
                                    st.success(f"송장 등록 완료: 총 {_total_success}건 성공" + (f", {_total_fail}건 실패" if _total_fail else ""))
                                    clear_order_caches()
                                    st.rerun()
                                elif _total_fail > 0:
                                    st.error(f"전체 실패: {_total_fail}건")

        except Exception as e:
            st.error(f"엑셀 파일 읽기 오류: {e}")

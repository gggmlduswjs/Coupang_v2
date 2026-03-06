"""광고 데이터 처리 — UI 독립 함수들.

ads.py에서 추출된 Excel/데이터 처리 순수 함수.
"""

import pandas as pd


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """DataFrame -> Excel bytes (단일 시트)"""
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        ws = writer.sheets[sheet_name]
        style_excel_header(ws, len(df.columns), len(df), sheet_name)
    return buf.getvalue()


def style_excel_header(ws, num_cols: int, num_rows: int, title: str = ""):
    """Excel 시트 헤더 스타일링 (export_order_sheets.py 패턴)"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 타이틀 행
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols, 1))
        cell = ws.cell(row=1, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="center")

    # 헤더 행 (row 2)
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=2, column=ci)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # 데이터 행 테두리
    for ri in range(3, 3 + num_rows):
        for ci in range(1, num_cols + 1):
            ws.cell(row=ri, column=ci).border = thin_border

    # 열 너비 자동 조정 (최소 12, 최대 40)
    for ci in range(1, num_cols + 1):
        header_val = str(ws.cell(row=2, column=ci).value or "")
        max_len = max(len(header_val) * 2, 12)  # 한글 2배
        for ri in range(3, min(3 + num_rows, 53)):  # 최대 50행 샘플
            val = str(ws.cell(row=ri, column=ci).value or "")
            max_len = max(max_len, min(len(val) + 2, 40))
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = min(max_len, 40)


def create_efficiency_excel(total_spend, wasted_spend, waste_pct, efficient_kw_count,
                             zero_conv_kw, good_kw, zero_conv_prod, good_prod,
                             camp_df, recs_df, d_from, d_to):
    """효율 리포트 멀티시트 Excel 생성"""
    from io import BytesIO

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        period_label = f"{d_from} ~ {d_to}"

        # Sheet 1: 요약
        summary = pd.DataFrame([
            {"항목": "총 광고비", "값": f"{total_spend:,}원"},
            {"항목": "낭비 광고비 (전환 0건 키워드)", "값": f"{wasted_spend:,}원"},
            {"항목": "낭비 비율", "값": f"{waste_pct}%"},
            {"항목": "효율 키워드 수 (ROAS>=200%)", "값": f"{efficient_kw_count}개"},
            {"항목": "분석 기간", "값": period_label},
        ])
        summary.to_excel(writer, sheet_name="요약", index=False, startrow=1)
        ws = writer.sheets["요약"]
        style_excel_header(ws, 2, len(summary), f"광고 효율 요약 ({period_label})")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Sheet 2: 비효율 키워드
        if not zero_conv_kw.empty:
            zero_conv_kw.to_excel(writer, sheet_name="비효율 키워드", index=False, startrow=1)
            ws = writer.sheets["비효율 키워드"]
            style_excel_header(ws, len(zero_conv_kw.columns), len(zero_conv_kw),
                                f"전환 0건 키워드 ({period_label})")
        else:
            pd.DataFrame({"메시지": ["전환 0건 키워드 없음"]}).to_excel(
                writer, sheet_name="비효율 키워드", index=False)

        # Sheet 3: 효율 키워드
        if not good_kw.empty:
            good_kw.to_excel(writer, sheet_name="효율 키워드", index=False, startrow=1)
            ws = writer.sheets["효율 키워드"]
            style_excel_header(ws, len(good_kw.columns), len(good_kw),
                                f"효율 키워드 ROAS>=200% ({period_label})")
        else:
            pd.DataFrame({"메시지": ["ROAS>=200% 키워드 없음"]}).to_excel(
                writer, sheet_name="효율 키워드", index=False)

        # Sheet 4: 비효율 상품
        if not zero_conv_prod.empty:
            zero_conv_prod.to_excel(writer, sheet_name="비효율 상품", index=False, startrow=1)
            ws = writer.sheets["비효율 상품"]
            style_excel_header(ws, len(zero_conv_prod.columns), len(zero_conv_prod),
                                f"전환 0건 상품 ({period_label})")
        else:
            pd.DataFrame({"메시지": ["전환 0건 상품 없음"]}).to_excel(
                writer, sheet_name="비효율 상품", index=False)

        # Sheet 5: 효율 상품
        if not good_prod.empty:
            good_prod.to_excel(writer, sheet_name="효율 상품", index=False, startrow=1)
            ws = writer.sheets["효율 상품"]
            style_excel_header(ws, len(good_prod.columns), len(good_prod),
                                f"효율 상품 ROAS>=200% ({period_label})")
        else:
            pd.DataFrame({"메시지": ["ROAS>=200% 상품 없음"]}).to_excel(
                writer, sheet_name="효율 상품", index=False)

        # Sheet 6: 캠페인 비교
        if not camp_df.empty:
            camp_df.to_excel(writer, sheet_name="캠페인 비교", index=False, startrow=1)
            ws = writer.sheets["캠페인 비교"]
            style_excel_header(ws, len(camp_df.columns), len(camp_df),
                                f"캠페인별 성과 ({period_label})")
        else:
            pd.DataFrame({"메시지": ["캠페인 데이터 없음"]}).to_excel(
                writer, sheet_name="캠페인 비교", index=False)

        # Sheet 7: 개선 제안
        if recs_df is not None and not recs_df.empty:
            recs_df.to_excel(writer, sheet_name="개선 제안", index=False, startrow=1)
            ws = writer.sheets["개선 제안"]
            style_excel_header(ws, len(recs_df.columns), len(recs_df),
                                f"개선 제안 ({period_label})")
        else:
            pd.DataFrame({"메시지": ["조치가 필요한 항목 없음"]}).to_excel(
                writer, sheet_name="개선 제안", index=False)

    return buf.getvalue()


def generate_recommendations(kw_df, prod_df, camp_df):
    """키워드/상품/캠페인별 조치 권장 목록 생성"""
    recs = []

    # ── 키워드 조치 ──
    if not kw_df.empty:
        for _, r in kw_df.iterrows():
            conv = int(r.get("전환주문", 0) or 0)
            spend = int(r.get("광고비", 0) or 0)
            roas = float(r.get("ROAS(%)", 0) or 0)
            clicks = int(r.get("클릭수", 0) or 0)
            impressions = int(r.get("노출수", 0) or 0)
            ctr = (clicks / impressions * 100) if impressions > 0 else 0
            name = r.get("키워드", "")
            campaign = r.get("캠페인", "")

            if conv == 0 and ctr >= 5:
                recs.append({"대상유형": "키워드", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "상품페이지 점검", "우선순위": "높음",
                             "사유": f"CTR {ctr:.1f}%로 높으나 전환 0건"})
            elif conv == 0 and spend >= 5000:
                recs.append({"대상유형": "키워드", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "키워드 중지", "우선순위": "높음",
                             "사유": f"전환 0건, 광고비 {spend:,}원 소진"})
            elif conv == 0 and spend < 5000:
                recs.append({"대상유형": "키워드", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "모니터링", "우선순위": "낮음",
                             "사유": "전환 0건, 데이터 부족"})
            elif roas < 100:
                recs.append({"대상유형": "키워드", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "입찰가 하향 또는 중지", "우선순위": "높음",
                             "사유": f"ROAS {roas:.0f}% 적자"})
            elif roas < 200:
                recs.append({"대상유형": "키워드", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "입찰가 하향 검토", "우선순위": "중간",
                             "사유": f"ROAS {roas:.0f}% 저효율"})
            elif roas >= 500 and clicks >= 10:
                recs.append({"대상유형": "키워드", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "입찰가 상향 검토", "우선순위": "중간",
                             "사유": f"ROAS {roas:.0f}%, 클릭 {clicks}회 -- 확대 여지"})

    # ── 상품 조치 ──
    if not prod_df.empty:
        for _, r in prod_df.iterrows():
            conv = int(r.get("전환주문", 0) or 0)
            spend = int(r.get("광고비", 0) or 0)
            roas = float(r.get("ROAS(%)", 0) or 0)
            name = r.get("상품명", "")
            campaign = r.get("캠페인", "")

            if conv == 0 and spend >= 10000:
                recs.append({"대상유형": "상품", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "광고 중지", "우선순위": "높음",
                             "사유": f"전환 0건, 광고비 {spend:,}원 소진"})
            elif conv > 0 and roas < 100:
                recs.append({"대상유형": "상품", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "예산 축소 또는 중지", "우선순위": "높음",
                             "사유": f"ROAS {roas:.0f}% 적자"})
            elif conv > 0 and roas >= 500:
                recs.append({"대상유형": "상품", "캠페인": campaign, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "예산 확대 검토", "우선순위": "중간",
                             "사유": f"ROAS {roas:.0f}% 고효율"})

    # ── 캠페인 조치 ──
    if not camp_df.empty:
        for _, r in camp_df.iterrows():
            roas = float(r.get("ROAS(%)", 0) or 0)
            spend = int(r.get("광고비", 0) or 0)
            name = r.get("캠페인", "")

            if roas < 100:
                recs.append({"대상유형": "캠페인", "캠페인": name, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "캠페인 예산 축소", "우선순위": "높음",
                             "사유": f"ROAS {roas:.0f}% 적자"})
            elif roas < 200:
                recs.append({"대상유형": "캠페인", "캠페인": name, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "키워드 정리 필요", "우선순위": "중간",
                             "사유": f"ROAS {roas:.0f}% 저효율"})
            elif roas >= 300:
                recs.append({"대상유형": "캠페인", "캠페인": name, "이름": name,
                             "ROAS(%)": roas, "광고비": spend,
                             "조치": "예산 확대 검토", "우선순위": "중간",
                             "사유": f"ROAS {roas:.0f}% 고효율"})

    if not recs:
        return pd.DataFrame()

    recs_df = pd.DataFrame(recs)
    # 우선순위 정렬: 높음 -> 중간 -> 낮음, 같은 순위 내 광고비 내림차순
    priority_order = {"높음": 0, "중간": 1, "낮음": 2}
    recs_df["_sort"] = recs_df["우선순위"].map(priority_order)
    recs_df = (recs_df.sort_values(["_sort", "광고비"], ascending=[True, False])
               .drop(columns="_sort").reset_index(drop=True))
    return recs_df

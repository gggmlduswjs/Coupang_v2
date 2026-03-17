"""주문 공유 데이터 로직 — orders.py / shipping.py 양쪽에서 사용.

ACCEPT/INSTRUCT: WING API 실시간 (정확한 책 목록)
그 외 (DEPARTURE/DELIVERING/FINAL_DELIVERY): DB (1분 동기화)
"""

import logging
import time
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st

from dashboard.utils import create_wing_client, query_df
from dashboard.services.order_service import (
    extract_price as _extract_price,
    parse_dt as _parse_dt,
    save_ordersheets_to_db as _save_ordersheets_to_db,
)

logger = logging.getLogger(__name__)

LIVE_STATUSES = ["ACCEPT", "INSTRUCT", "DEPARTURE", "DELIVERING", "FINAL_DELIVERY", "NONE_TRACKING"]

STATUS_MAP = {
    "ACCEPT": "결제완료", "INSTRUCT": "상품준비중", "DEPARTURE": "출고완료",
    "DELIVERING": "배송중", "FINAL_DELIVERY": "배송완료", "NONE_TRACKING": "추적불가",
}

# API 실시간 조회 대상 (정확도 필수)
_API_STATUSES = ["ACCEPT", "INSTRUCT"]
# DB 조회 대상 (백그라운드 동기화, 속도 우선)
_DB_STATUSES = ["DEPARTURE", "DELIVERING", "FINAL_DELIVERY", "NONE_TRACKING"]


def _api_row(acct_name, acct_id, status, os_data, item):
    """API 응답 → DataFrame 행 변환 (쿠팡 OPEN API v5 스펙 기준)"""
    vid = item.get("vendorItemId") or os_data.get("vendorItemId") or 0
    spid = item.get("sellerProductId") or os_data.get("sellerProductId")
    sp_name = item.get("sellerProductName") or os_data.get("sellerProductName", "")
    order_price = _extract_price(item.get("orderPrice"))
    sales_price = _extract_price(item.get("salesPrice"))
    shipping_price = _extract_price(os_data.get("shippingPrice"))
    remote_price = _extract_price(os_data.get("remotePrice"))

    orderer = os_data.get("orderer") or {}
    receiver = os_data.get("receiver") or {}
    overseas = os_data.get("overseaShippingInfoDto") or {}
    addr1 = receiver.get("addr1", "") or ""
    addr2 = receiver.get("addr2", "") or ""

    ordered_at = _parse_dt(os_data.get("orderedAt"))
    delivered_date = _parse_dt(os_data.get("deliveredDate"))

    # 등록옵션명 / 최초등록옵션명 (orderItems 레벨)
    seller_item_name = item.get("sellerProductItemName") or ""
    first_item_name = item.get("firstSellerProductItemName") or ""

    return {
        "계정": acct_name,
        "묶음배송번호": os_data.get("shipmentBoxId"),
        "주문번호": os_data.get("orderId"),
        "상품명": sp_name,
        "옵션명": item.get("vendorItemName") or "",
        "수량": int(item.get("shippingCount", 0) or 0),
        "결제금액": order_price,
        "주문일": ordered_at[:10] if ordered_at else "",
        "수취인": receiver.get("name", ""),
        "상태": status,
        "택배사": os_data.get("deliveryCompanyName", ""),
        "운송장번호": os_data.get("invoiceNumber", ""),
        "배송완료일": delivered_date[:10] if delivered_date else "",
        "취소": bool(item.get("canceled", False)),
        "_account_id": acct_id,
        "_vendor_item_id": int(vid) if vid else 0,
        "_seller_product_id": int(spid) if spid else None,
        "_order_price_raw": order_price,
        "주문일시": ordered_at or "",
        "구매자": orderer.get("name", ""),
        "구매자전화번호": orderer.get("safeNumber") or orderer.get("ordererNumber") or "",
        "수취인전화번호": receiver.get("safeNumber") or receiver.get("receiverNumber") or "",
        "우편번호": receiver.get("postCode", ""),
        "수취인주소": f"{addr1} {addr2}".strip(),
        "배송메세지": os_data.get("parcelPrintMessage") or "",
        "배송비": shipping_price,
        "도서산간추가배송비": remote_price,
        "결제위치": os_data.get("refer", ""),
        "분리배송가능": bool(os_data.get("ableSplitShipping", False)),
        "주문시출고예정일": item.get("estimatedShippingDate") or "",
        "배송비구분": item.get("deliveryChargeTypeName") or "",
        "판매단가": sales_price,
        "최초등록상품옵션명": f"{sp_name},{first_item_name}" if first_item_name else "",
        "업체상품코드": item.get("externalVendorSkuCode") or "",
        "개인통관번호": overseas.get("personalCustomsClearanceCode") or "",
        "통관용전화번호": overseas.get("ordererPhoneNumber") or "",
        "등록옵션명": seller_item_name,
    }


def load_all_orders_live(accounts_df):
    """ACCEPT/INSTRUCT → API 실시간, 나머지 → DB. 세션 캐시 30초."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    _cache_key = "_orders_live_cache"
    _ts_key = "_orders_live_ts"
    now = time.time()
    if now - st.session_state.get(_ts_key, 0) < 30 and _cache_key in st.session_state:
        return st.session_state[_cache_key]

    _today = date.today()
    _from = (_today - timedelta(days=7)).isoformat()
    _to = _today.isoformat()

    # ── 1) ACCEPT/INSTRUCT: API 실시간 ──
    acct_clients = []
    for _, acct in accounts_df.iterrows():
        client = create_wing_client(acct)
        if client:
            acct_clients.append((acct, client))

    api_rows = []
    if acct_clients:
        def _fetch(acct, client, status):
            try:
                return acct, status, client.get_all_ordersheets(_from, _to, status=status)
            except Exception as e:
                logger.warning(f"[{acct['account_name']}] {status} 조회 실패: {e}")
                return acct, status, []

        max_workers = min(len(acct_clients) * len(_API_STATUSES), 15) or 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [
                pool.submit(_fetch, acct, client, status)
                for acct, client in acct_clients
                for status in _API_STATUSES
            ]
            for f in as_completed(futures):
                acct, status, ordersheets = f.result()
                acct_name = acct["account_name"]
                acct_id = int(acct["id"])

                if ordersheets:
                    _save_ordersheets_to_db(acct, ordersheets, status)

                for os_data in ordersheets:
                    if not os_data.get("shipmentBoxId") or not os_data.get("orderId"):
                        continue
                    items = os_data.get("orderItems") or [os_data]
                    for item in items:
                        api_rows.append(_api_row(acct_name, acct_id, status, os_data, item))

    api_df = pd.DataFrame(api_rows) if api_rows else pd.DataFrame()

    # ── 2) 나머지 상태: DB 조회 (빠름) ──
    _from_30d = (_today - timedelta(days=30)).isoformat()
    db_df = query_df("""
        SELECT a.account_name AS "계정",
               o.shipment_box_id AS "묶음배송번호",
               o.order_id AS "주문번호",
               o.seller_product_name AS "상품명",
               o.vendor_item_name AS "옵션명",
               o.shipping_count AS "수량",
               o.order_price AS "결제금액",
               to_char(o.ordered_at, 'YYYY-MM-DD') AS "주문일",
               o.receiver_name AS "수취인",
               o.status AS "상태",
               o.delivery_company_name AS "택배사",
               o.invoice_number AS "운송장번호",
               to_char(o.delivered_date, 'YYYY-MM-DD') AS "배송완료일",
               COALESCE(o.canceled, false) AS "취소",
               o.account_id AS "_account_id",
               o.vendor_item_id AS "_vendor_item_id",
               o.seller_product_id AS "_seller_product_id",
               o.order_price AS "_order_price_raw",
               to_char(o.ordered_at, 'YYYY-MM-DD HH24:MI:SS') AS "주문일시",
               o.orderer_name AS "구매자",
               '' AS "구매자전화번호",
               '' AS "수취인전화번호",
               o.receiver_post_code AS "우편번호",
               o.receiver_addr AS "수취인주소",
               '' AS "배송메세지",
               COALESCE(o.shipping_price, 0) AS "배송비",
               0 AS "도서산간추가배송비",
               COALESCE(o.refer, '') AS "결제위치",
               false AS "분리배송가능",
               '' AS "주문시출고예정일",
               '' AS "배송비구분",
               COALESCE(o.sales_price, 0) AS "판매단가",
               '' AS "최초등록상품옵션명",
               '' AS "업체상품코드",
               '' AS "개인통관번호",
               '' AS "통관용전화번호"
        FROM orders o
        JOIN accounts a ON o.account_id = a.id
        WHERE (o.status IN ('DEPARTURE','DELIVERING','NONE_TRACKING'))
           OR (o.status = 'FINAL_DELIVERY' AND o.ordered_at >= :date_from)
           OR (o.status IN ('ACCEPT','INSTRUCT') AND o.ordered_at < :api_cutoff AND o.canceled = false)
        ORDER BY "주문일시" DESC
    """, {"date_from": _from_30d, "api_cutoff": _from})

    # ── 3) 합치기 (API 우선, DB 보완 — 중복 제거) ──
    if not api_df.empty:
        api_df["_src"] = "api"
    if not db_df.empty:
        db_df["_src"] = "db"
    frames = [df for df in [api_df, db_df] if not df.empty]
    if frames:
        result = pd.concat(frames, ignore_index=True)
        # API 행 우선: 같은 (묶음배송번호, _vendor_item_id) 중복 시 API 유지
        result = result.sort_values("_src").drop_duplicates(
            subset=["묶음배송번호", "_vendor_item_id", "_account_id"], keep="first"
        )
        result = result.drop(columns=["_src"]).sort_values("주문일시", ascending=False).reset_index(drop=True)
    else:
        result = pd.DataFrame()

    st.session_state["order_last_synced"] = (datetime.utcnow() + timedelta(hours=9)).strftime("%H:%M:%S")
    st.session_state[_cache_key] = result
    st.session_state[_ts_key] = now
    return result


def get_instruct_orders(all_orders):
    """INSTRUCT 상태 필터링 (취소 제외)"""
    if all_orders.empty:
        return pd.DataFrame()
    instruct = all_orders[all_orders["상태"] == "INSTRUCT"].copy()
    if instruct.empty:
        return pd.DataFrame()
    return instruct[~instruct["취소"]].copy()


def get_instruct_by_box(instruct_all):
    """묶음배송 단위 그룹핑 (주소 포함)"""
    if instruct_all.empty:
        return pd.DataFrame()
    # 수취인주소가 없는 경우 빈 문자열로 채움
    df = instruct_all.copy()
    if "수취인주소" not in df.columns:
        df["수취인주소"] = ""
    else:
        df["수취인주소"] = df["수취인주소"].fillna("")
    return df.groupby(["계정", "묶음배송번호", "주문번호", "주문일", "수취인", "수취인주소"]).agg(
        상품명=("상품명", lambda x: " / ".join(x.unique())),
        수량=("수량", "sum"),
        결제금액=("_order_price_raw", "sum"),
    ).reset_index()


def clear_order_caches():
    """세션 캐시 초기화 → 다음 호출 시 API 재조회"""
    st.session_state.pop("_orders_live_cache", None)
    st.session_state.pop("_orders_live_ts", None)
    st.cache_data.clear()


def fmt_krw_short(val):
    """금액 축약 포맷"""
    val = int(val)
    if abs(val) >= 100_000_000:
        return f"{val / 100_000_000:.1f}억"
    elif abs(val) >= 10_000:
        return f"{val / 10_000:.0f}만"
    else:
        return f"{val:,}"


def sync_live_orders(accounts_df):
    """수동 새로고침 — 캐시 클리어 후 재조회"""
    clear_order_caches()
    result = load_all_orders_live(accounts_df)
    return len(result) if not result.empty else 0


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 공유 DeliveryList 생성 로직
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def _get_recent_delivery_receivers():
    """미등록 배송리스트 + 최근 3일 등록 건의 수취인 정보 반환.

    한진 N-Focus는 이름만 같아도, 주소만 같아도 합배송할 수 있으므로,
    미등록(아직 한진에 올라가있는) 건은 시간 무관하게 전부 체크하고,
    최근 등록 완료 건도 3일간 체크하여 합배송을 방지한다.

    Returns:
        (name_to_boxes, addr_to_boxes): 이름→[box_ids], 주소→[box_ids]
    """
    from sqlalchemy import text
    from core.database import engine

    name_to_boxes = {}
    addr_to_boxes = {}
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT d.shipment_box_id, o.receiver_name, o.receiver_addr
                FROM delivery_list_logs d
                JOIN orders o ON d.shipment_box_id = o.shipment_box_id
                WHERE d.registered = FALSE
                   OR d.downloaded_at >= NOW() - INTERVAL '3 days'
            """)).fetchall()
            for r in rows:
                name = (r[1] or "").strip()
                addr = (r[2] or "").strip()
                box_id = int(r[0])
                if name:
                    if name not in name_to_boxes:
                        name_to_boxes[name] = []
                    if box_id not in name_to_boxes[name]:
                        name_to_boxes[name].append(box_id)
                if addr:
                    if addr not in addr_to_boxes:
                        addr_to_boxes[addr] = []
                    if box_id not in addr_to_boxes[addr]:
                        addr_to_boxes[addr].append(box_id)
    except Exception:
        pass  # DB 접근 실패 시 기존 로직(현재 배치만)으로 동작
    return name_to_boxes, addr_to_boxes


def _build_receiver_suffix_map(orders_df):
    """동일 배송주소(primary) 또는 동일 수취인이름(secondary)이 서로 다른 묶음배송번호에 있으면 구분자 부여.

    한진 N-Focus는 배송 주소가 같으면 자동 합배송하므로 주소를 1차 기준으로 감지하고,
    이름이 같은 경우도 보조적으로 구분자를 부여한다.
    (다른 계정에서 같은 주소로 주문한 건도 감지)

    미등록 배송리스트 전체 + 최근 3일 등록 건도 함께 고려하여,
    배치 간(cross-batch) 합배송을 방지한다.

    Returns:
        dict: {묶음배송번호: {"name": str, "addr": str}}
    """
    # 1) 미등록 + 최근 3일 등록 건 로드 (배치 간 합배송 방지)
    name_to_boxes, addr_to_boxes = _get_recent_delivery_receivers()

    # 2) 현재 배치 주문 추가
    current_box_ids = set()
    _names = orders_df.get("수취인", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
    _addrs = orders_df.get("수취인주소", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
    _boxes = orders_df["묶음배송번호"].astype(int)
    for name, addr, box_id in zip(_names, _addrs, _boxes):
        current_box_ids.add(box_id)
        if name:
            if name not in name_to_boxes:
                name_to_boxes[name] = []
            if box_id not in name_to_boxes[name]:
                name_to_boxes[name].append(box_id)
        if addr:
            if addr not in addr_to_boxes:
                addr_to_boxes[addr] = []
            if box_id not in addr_to_boxes[addr]:
                addr_to_boxes[addr].append(box_id)

    suffix_map = {}  # 묶음배송번호 → {"name": str, "addr": str}

    # 주소 기준 구분자 (PRIMARY — 한진은 주소가 같으면 합배송)
    for addr, box_ids in addr_to_boxes.items():
        if len(box_ids) <= 1:
            continue
        for i, box_id in enumerate(box_ids):
            if box_id in current_box_ids:
                tag = f" ({i + 1})"
                if box_id not in suffix_map:
                    suffix_map[box_id] = {"name": "", "addr": ""}
                suffix_map[box_id]["name"] = tag
                suffix_map[box_id]["addr"] = tag

    # 이름 기준 구분자 (SECONDARY — 주소 구분자가 없는 건만 보완)
    for name, box_ids in name_to_boxes.items():
        if len(box_ids) <= 1:
            continue
        for i, box_id in enumerate(box_ids):
            if box_id in current_box_ids:
                tag = f" ({i + 1})"
                if box_id not in suffix_map:
                    suffix_map[box_id] = {"name": "", "addr": ""}
                if not suffix_map[box_id]["name"]:
                    suffix_map[box_id]["name"] = tag

    return suffix_map


def build_delivery_rows(orders_df):
    """주문 DataFrame → 쿠팡 DeliveryList 40컬럼 행 목록 변환.

    Returns:
        list[dict]: 각 행이 DeliveryList 한 줄에 해당하는 딕셔너리 목록.
        _account_id, _vendor_item_id 내부 컬럼 포함.
    """
    suffix_map = _build_receiver_suffix_map(orders_df)

    rows = []
    for idx, (_i, row) in enumerate(orders_df.iterrows(), 1):
        box_id = int(row["묶음배송번호"])
        receiver_name = str(row.get("수취인", ""))
        receiver_addr = str(row.get("수취인주소", ""))
        suffixes = suffix_map.get(box_id, {"name": "", "addr": ""})
        # 하위호환: 기존 문자열 suffix도 처리
        if isinstance(suffixes, str):
            suffixes = {"name": suffixes, "addr": ""}

        rows.append({
            "번호": idx,
            "묶음배송번호": box_id,
            "주문번호": int(row["주문번호"]),
            "택배사": "한진택배",
            "운송장번호": "",
            "분리배송 Y/N": "분리배송가능" if row.get("분리배송가능") else "분리배송불가",
            "분리배송 출고예정일": "",
            "주문시 출고예정일": row.get("주문시출고예정일", ""),
            "출고일(발송일)": "",
            "주문일": row.get("주문일시", row.get("주문일", "")),
            "등록상품명": str(row.get("상품명") or ""),
            "등록옵션명": row.get("등록옵션명") or row.get("옵션명", ""),
            "노출상품명(옵션명)": f"{row.get('상품명', '')}, {row.get('옵션명', '')}",
            "노출상품ID": str(row.get("_seller_product_id", "")),
            "옵션ID": str(row.get("_vendor_item_id", "")),
            "최초등록등록상품명/옵션명": row.get("최초등록상품옵션명", ""),
            "업체상품코드": row.get("업체상품코드", ""),
            "바코드": "",
            "결제액": int(row.get("결제금액", 0)),
            "배송비구분": row.get("배송비구분", ""),
            "배송비": row.get("배송비", 0),
            "도서산간 추가배송비": int(row.get("도서산간추가배송비", 0)),
            "구매수(수량)": int(row.get("수량", 0)),
            "옵션판매가(판매단가)": int(row.get("판매단가", 0) or row.get("결제금액", 0)),
            "구매자": row.get("구매자", ""),
            "구매자전화번호": row.get("구매자전화번호", ""),
            "수취인이름": receiver_name + suffixes["name"],
            "수취인전화번호": row.get("수취인전화번호", ""),
            "우편번호": row.get("우편번호", ""),
            "수취인 주소": receiver_addr + suffixes["addr"],
            "배송메세지": row.get("배송메세지", ""),
            "상품별 추가메시지": "",
            "주문자 추가메시지": "",
            "배송완료일": "",
            "구매확정일자": "",
            "개인통관번호(PCCC)": row.get("개인통관번호", ""),
            "통관용수취인전화번호": row.get("통관용전화번호", ""),
            "기타": f"BOX:{box_id}",
            "결제위치": row.get("결제위치", ""),
            "배송유형": "판매자 배송",
            "_account_id": int(row.get("_account_id", 0)),
            "_vendor_item_id": int(row.get("_vendor_item_id", 0)),
        })
    return rows


def build_delivery_excel_bytes(orders_df, *, include_internal_cols=False, sort_and_color=True):
    """주문 DataFrame → DeliveryList 엑셀 bytes.

    Args:
        orders_df: INSTRUCT 주문 DataFrame
        include_internal_cols: True면 _account_id, _vendor_item_id 포함
        sort_and_color: True면 책별 정렬 + 색상 + 픽킹리스트 시트 포함

    Returns:
        (bytes, DataFrame): 엑셀 바이트, 내부 컬럼 포함된 DataFrame (세션 저장용)
    """
    import io
    from datetime import date

    dl_rows = build_delivery_rows(orders_df)
    dl_df = pd.DataFrame(dl_rows)
    dl_df["등록상품명"] = dl_df["등록상품명"].fillna("").astype(str)

    if sort_and_color:
        # 묶음배송 구분: 단건 먼저, 묶음 뒤로 + 같은 책끼리 그룹핑
        box_counts = dl_df.groupby("묶음배송번호")["묶음배송번호"].transform("count")
        dl_df["_is_bundle"] = (box_counts > 1).astype(int)
        dl_df["_bundle_first_book"] = dl_df.groupby("묶음배송번호")["등록상품명"].transform("first")
        dl_df["_is_single_qty"] = (dl_df["구매수(수량)"] <= 1).astype(int)
        dl_df = dl_df.sort_values(
            ["_is_bundle", "_bundle_first_book", "등록상품명", "_is_single_qty", "묶음배송번호"]
        ).reset_index(drop=True)
        dl_df = dl_df.drop(columns=["_is_bundle", "_bundle_first_book", "_is_single_qty"])
        dl_df["번호"] = range(1, len(dl_df) + 1)

    # 엑셀 출력용 (내부 컬럼 제외)
    excel_cols = [c for c in dl_df.columns if not c.startswith("_")]
    dl_excel = dl_df[excel_cols].copy()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        dl_excel.to_excel(writer, sheet_name="Delivery", index=False)
        ws = writer.sheets["Delivery"]
        from openpyxl.utils import get_column_letter

        # 텍스트 포맷 (지수 표기 방지)
        for col_name in ["묶음배송번호", "주문번호", "노출상품ID", "옵션ID"]:
            if col_name in dl_excel.columns:
                col_idx = dl_excel.columns.get_loc(col_name)
                col_letter = get_column_letter(col_idx + 1)
                for row_idx in range(2, len(dl_excel) + 2):
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.value = str(int(cell.value)) if cell.value is not None else ""
                    cell.number_format = "@"

        if sort_and_color:
            from openpyxl.styles import PatternFill
            # 같은 책 첫 행에 색상 표시
            prev_book = None
            fill = PatternFill(start_color="D9E8FB", end_color="D9E8FB", fill_type="solid")
            for row_idx, book in enumerate(dl_excel["등록상품명"], start=2):
                if book != prev_book:
                    for c in range(1, len(dl_excel.columns) + 1):
                        ws.cell(row=row_idx, column=c).fill = fill
                    prev_book = book

            # 픽킹리스트 시트
            pick_summary = (
                dl_df.groupby("등록상품명")
                .agg(건수=("묶음배송번호", "count"), 총수량=("구매수(수량)", "sum"))
                .sort_index()
                .reset_index()
            )
            pick_summary.columns = ["도서명", "주문건수", "총수량"]
            pick_summary.to_excel(writer, sheet_name="픽킹리스트", index=False)
            ws2 = writer.sheets["픽킹리스트"]
            ws2.column_dimensions["A"].width = 60
            ws2.column_dimensions["B"].width = 12
            ws2.column_dimensions["C"].width = 12

    buf.seek(0)
    return buf.getvalue(), dl_df

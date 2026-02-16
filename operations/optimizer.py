"""상품명/검색어 자동 최적화 + Wing Excel 내보내기"""

import os
import re
from copy import copy
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from core.config import AnalysisConfig

# ── 카테고리 파싱 ──

# 카테고리에서 학년/과목 추출용 매핑
GRADE_MAP = {
    "초등학습": "초등",
    "초등참고서": "초등",
    "어린이": "초등",
    "중학교": "중등",
    "중고등참고서": "",  # 하위에서 판단
    "고등학교": "고등",
    "수험서/자격증": "자격증",
}

SUBJECT_FROM_CAT = {
    "영어": "영어", "수학": "수학", "국어": "국어",
    "과학": "과학", "사회": "사회",
    "물리": "물리학", "화학": "화학", "생명": "생명과학", "지구": "지구과학",
    "윤리": "윤리", "한국사": "한국사", "역사": "역사",
    "독해": "영어 독해", "문법": "문법", "작문": "작문",
}


def parse_category(cat_str: str) -> dict:
    """카테고리 문자열에서 학년/과목/세부 정보 추출"""
    info = {"grade": "", "grade_detail": "", "subject": "", "book_type": "", "raw": cat_str}
    if not cat_str:
        return info

    # "[76236] 도서>국내도서>중고등참고서>고등학교>문제집" 형식
    parts = [p.strip() for p in cat_str.split(">")]

    for part in parts:
        # 학년 추출
        for key, grade in GRADE_MAP.items():
            if key in part and not info["grade"]:
                info["grade"] = grade

        # 중1, 중2, 중3, 고1 등
        m = re.search(r"(중|고)(\d)", part)
        if m:
            info["grade"] = "중등" if m.group(1) == "중" else "고등"
            info["grade_detail"] = f"{m.group(1)}{m.group(2)}"

        # 과목 추출
        for key, subject in SUBJECT_FROM_CAT.items():
            if key in part and not info["subject"]:
                info["subject"] = subject

        # 문제집, 기본서 등
        if "문제집" in part:
            info["book_type"] = "문제집"
        elif "기본서" in part or "개념" in part:
            info["book_type"] = "기본서"
        elif "기출" in part:
            info["book_type"] = "기출문제집"

    return info


# ── 상품명 파싱 ──

def parse_product_name(name: str) -> dict:
    """기존 상품명에서 구성 요소 추출"""
    info = {
        "year": "", "series": "", "title": name,
        "subject": "", "grade": "", "set_info": "",
        "edition": "", "extras": [],
    }

    # 연도 추출
    m = re.search(r"(20\d{2})(?:년?(?:용)?)?", name)
    if m:
        info["year"] = m.group(1)

    # 세트 정보
    m = re.search(r"(전\s?\d+권\s*세트|세트\s*(?:전\s?\d+권)?|\d+권\s*세트|전\s?\d+권)", name)
    if m:
        info["set_info"] = m.group(1).strip()

    # 학기 정보 (1-1, 2-2 등)
    m = re.search(r"(\d-[12])", name)
    if m:
        info["grade"] = m.group(1)

    # 학년/대상
    grade_patterns = [
        (r"고등\s*\d?\s*학년", ""), (r"중등\s*\d?\s*학년", ""),
        (r"고[123]", ""), (r"중[123]", ""),
        (r"초등\s*\d?\s*학년", ""),
        (r"고등학생", "고등"), (r"중학생", "중등"), (r"초등학생", "초등"),
    ]
    for pat, _ in grade_patterns:
        m = re.search(pat, name)
        if m and not info["grade"]:
            info["grade"] = m.group(0)

    # 과목
    subject_pats = [
        "통합과학", "통합사회", "공통수학", "수학", "영어", "국어",
        "물리학", "화학", "생명과학", "지구과학",
        "사회·문화", "사회문화", "생활과 윤리", "윤리와 사상",
        "한국사", "세계사", "경제", "정치와 법",
        "확률과 통계", "미적분", "기하",
        "독서", "문학", "화법과 작문", "언어와 매체",
    ]
    for subj in subject_pats:
        if subj in name and not info["subject"]:
            info["subject"] = subj

    # 개정 정보
    m = re.search(r"(20\d{2}\s*개정|22개정|개정\s*교육과정|개정판)", name)
    if m:
        info["edition"] = m.group(1)

    return info


# ── 상품명 최적화 ──

def optimize_name(name: str, brand: str, maker: str, category: str) -> str:
    """상품명 최적화 — 원본 유지 원칙, 연도만 보정.

    분석 결과:
    - 연도가 뒤(50%+)에 있을 때 순위 가장 좋음 (35.8)
    - 키워드는 포함 여부가 중요, 위치는 무관
    - 상위권 패턴: [브랜드] [핵심제목] [학년] (연도)
    → 원본 상품명 구조 유지, 연도만 추가/갱신
    """
    if not name or not name.strip():
        return name

    result = name.strip()
    name_info = parse_product_name(result)

    # 연도 결정
    if any(k in result for k in ["수능", "수특", "모의고사", "수능대비"]):
        target_year = "2027"
    else:
        target_year = "2026"

    if name_info["year"]:
        old_year = name_info["year"]
        old_year_int = int(old_year)
        # 올해(2026) 이전 연도만 갱신, 이미 최신이면 그대로
        if old_year_int < 2026:
            # 기존 연도를 새 연도로 교체 (형식 유지)
            result = re.sub(
                rf"({old_year})(년?\s*(?:용)?(?:\))?)",
                rf"{target_year}\2",
                result,
                count=1,
            )
    else:
        # 연도 없으면 뒤에 추가
        # 이미 괄호로 끝나면 앞에, 아니면 (연도) 형태로
        result = result.rstrip()
        if result.endswith(")"):
            # 마지막 괄호 앞에 연도 삽입은 복잡하므로 뒤에 추가
            result = f"{result} ({target_year}년)"
        else:
            result = f"{result} ({target_year}년)"

    return result


# ── 검색어 자동 생성 ──

# 과목 관련 확장 키워드
SUBJECT_EXPANSIONS = {
    "수학": ["수학문제집", "수학교재", "수학기출"],
    "영어": ["영어문제집", "영어교재", "영어기출"],
    "국어": ["국어문제집", "국어교재", "국어기출"],
    "과학": ["과학문제집", "과학교재"],
    "사회": ["사회문제집", "사회교재"],
    "물리학": ["물리학1", "물리"],
    "화학": ["화학1", "화학2"],
    "생명과학": ["생명과학1", "생과"],
    "지구과학": ["지구과학1", "지과"],
}

# 학년 관련 확장
GRADE_EXPANSIONS = {
    "고등": ["고등학생", "고등학교"],
    "고1": ["고등 1학년", "고1"],
    "고2": ["고등 2학년", "고2"],
    "고3": ["고등 3학년", "고3", "수능"],
    "중등": ["중학생", "중학교"],
    "중1": ["중학 1학년", "중1"],
    "중2": ["중학 2학년", "중2"],
    "중3": ["중학 3학년", "중3"],
    "초등": ["초등학생", "초등학교"],
}


def generate_keywords(name: str, brand: str, maker: str, category: str,
                      existing_keywords: str = "", max_count: int = 20) -> str:
    """상품 정보에서 검색어를 자동 생성. 쉼표 구분 문자열 반환."""
    keywords = set()

    name_info = parse_product_name(name)
    cat_info = parse_category(category)

    brand_clean = re.sub(r"\(.*?\)", "", brand).strip() if brand else ""
    maker_clean = re.sub(r"\(.*?\)", "", maker).strip() if maker else ""

    # 1) 기존 검색어 유지
    if existing_keywords and existing_keywords != "None":
        for kw in existing_keywords.split(","):
            kw = kw.strip()
            if kw:
                keywords.add(kw)

    # 2) 브랜드/제조사
    if brand_clean:
        keywords.add(brand_clean)
    if maker_clean and maker_clean != brand_clean:
        keywords.add(maker_clean)
    # 원래 괄호 포함 브랜드도
    if brand and brand != brand_clean:
        keywords.add(brand)

    # 3) 상품명에서 의미 있는 토큰 추출
    name_tokens = re.findall(r"[가-힣]{2,}|[A-Za-z]{2,}|\d{4}", name)
    # 불용어 제거
    stopwords = {"전권", "세트", "전2권", "전3권", "전4권", "학년", "고등학생",
                 "중학생", "모든", "단품", "바로배송", "오늘출발", "개정",
                 "교육과정", "년용", "권세트"}
    for token in name_tokens:
        if token not in stopwords and len(token) >= 2:
            keywords.add(token)

    # 4) 과목 확장
    subject = name_info["subject"] or cat_info["subject"]
    if subject:
        keywords.add(subject)
        for exp in SUBJECT_EXPANSIONS.get(subject, []):
            keywords.add(exp)

    # 5) 학년 확장
    grade = name_info["grade"] or cat_info["grade_detail"] or cat_info["grade"]
    if grade:
        keywords.add(grade)
        for exp in GRADE_EXPANSIONS.get(grade, []):
            keywords.add(exp)

    # 6) 카테고리에서 키워드
    if cat_info["grade"]:
        keywords.add(cat_info["grade"])
    if cat_info["book_type"]:
        keywords.add(cat_info["book_type"])

    # 7) 연도
    year = name_info["year"]
    if year:
        keywords.add(year)
    keywords.add("2026")
    if any(k in name for k in ["수능", "수특", "모의고사", "수능대비"]):
        keywords.add("2027")
        keywords.add("수능대비")
        keywords.add("2027수능")

    # 8) 교재 공통 키워드
    if cat_info["grade"] in ("중등", "고등", "초등"):
        keywords.add("내신")
        keywords.add("내신대비")
    if "기출" in name:
        keywords.add("기출문제집")
        keywords.add("기출문제")
    if "문제집" in name or cat_info["book_type"] == "문제집":
        keywords.add("문제집")

    # 9) 학기 정보
    m = re.search(r"(\d)-([12])", name)
    if m:
        semester = f"{m.group(1)}-{m.group(2)}"
        keywords.add(semester)
        keywords.add(f"{m.group(2)}학기")

    # 정리: 1글자 제거, 정렬, 상한
    keywords = {k for k in keywords if len(k) >= 2}
    keyword_list = sorted(keywords)[:max_count]

    return ",".join(keyword_list)


# ── 배치 최적화 ──

def batch_optimize(account_code: str, config: AnalysisConfig = None,
                   preview: int = 0) -> pd.DataFrame:
    """계정의 Wing Excel을 읽어 상품명/검색어 최적화.

    Args:
        account_code: 계정 코드
        config: 설정
        preview: > 0이면 해당 개수만 미리보기

    Returns:
        최적화된 DataFrame (등록상품ID, 원래상품명, 최적화상품명, 원래검색어, 최적화검색어)
    """
    config = config or AnalysisConfig()
    excel_path = os.path.join(config.base_dir, "엑셀", f"{account_code}.xlsx")

    if not os.path.exists(excel_path):
        print(f"  Excel 파일을 찾을 수 없습니다: {excel_path}")
        return pd.DataFrame()

    print(f"  파일 로딩: {excel_path}")
    df = pd.read_excel(excel_path, sheet_name="Template", header=3, dtype=str)
    df = df.dropna(subset=["등록상품ID"])

    if preview > 0:
        df = df.head(preview)

    print(f"  대상 상품: {len(df)}개")

    results = []
    for _, row in df.iterrows():
        pid = str(row["등록상품ID"])
        orig_name = str(row.get("쿠팡 노출상품명", "")) if pd.notna(row.get("쿠팡 노출상품명")) else ""
        brand = str(row.get("브랜드", "")) if pd.notna(row.get("브랜드")) else ""
        maker = str(row.get("제조사", "")) if pd.notna(row.get("제조사")) else ""
        category = str(row.get("카테고리", "")) if pd.notna(row.get("카테고리")) else ""
        orig_kw = str(row.get("검색어", "")) if pd.notna(row.get("검색어")) and str(row.get("검색어")) != "None" else ""
        status = str(row.get("판매상태", "")) if pd.notna(row.get("판매상태")) else ""

        new_name = optimize_name(orig_name, brand, maker, category)
        new_kw = generate_keywords(orig_name, brand, maker, category, orig_kw)

        results.append({
            "등록상품ID": pid,
            "판매상태": status,
            "원래상품명": orig_name,
            "최적화상품명": new_name,
            "이름변경": "O" if new_name != orig_name else "",
            "원래검색어": orig_kw,
            "최적화검색어": new_kw,
            "검색어수": len(new_kw.split(",")) if new_kw else 0,
        })

    result_df = pd.DataFrame(results)
    changed = result_df[result_df["이름변경"] == "O"]
    print(f"  상품명 변경: {len(changed)}/{len(result_df)}개")
    print(f"  검색어 평균: {result_df['검색어수'].mean():.1f}개")

    return result_df


def export_wing_excel(account_code: str, output_path: str = "",
                      config: AnalysisConfig = None) -> str:
    """최적화된 데이터를 Wing Excel 형식으로 내보내기.
    원본 Excel을 복사하고 노출상품명/검색어만 수정."""
    config = config or AnalysisConfig()
    excel_path = os.path.join(config.base_dir, "엑셀", f"{account_code}.xlsx")

    if not os.path.exists(excel_path):
        print(f"  원본 Excel을 찾을 수 없습니다: {excel_path}")
        return ""

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(config.base_dir, "엑셀", "optimized")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{account_code}_optimized_{timestamp}.xlsx")

    # 원본 로드 (openpyxl)
    print(f"  원본 로드: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb["Template"]

    # 헤더 행 (4번째 = index 3, openpyxl은 1-based이므로 row 4)
    header = [str(ws.cell(row=4, column=c).value or "") for c in range(1, ws.max_column + 1)]

    name_col = header.index("쿠팡 노출상품명") + 1  # 1-based
    search_col = header.index("검색어") + 1
    brand_col = header.index("브랜드") + 1
    maker_col = header.index("제조사") + 1
    cat_col = header.index("카테고리") + 1

    # 데이터 행은 row 5부터
    modified = 0
    for row_num in range(5, ws.max_row + 1):
        pid = ws.cell(row=row_num, column=1).value
        if not pid:
            continue

        orig_name = str(ws.cell(row=row_num, column=name_col).value or "")
        brand = str(ws.cell(row=row_num, column=brand_col).value or "")
        maker = str(ws.cell(row=row_num, column=maker_col).value or "")
        category = str(ws.cell(row=row_num, column=cat_col).value or "")
        orig_kw = str(ws.cell(row=row_num, column=search_col).value or "")
        if orig_kw == "None":
            orig_kw = ""

        new_name = optimize_name(orig_name, brand, maker, category)
        new_kw = generate_keywords(orig_name, brand, maker, category, orig_kw)

        ws.cell(row=row_num, column=name_col).value = new_name
        ws.cell(row=row_num, column=search_col).value = new_kw
        modified += 1

    wb.save(output_path)
    wb.close()

    print(f"  {modified}개 상품 최적화 완료")
    print(f"  출력: {output_path}")
    return output_path


# ── CLI 핸들러 ──

def cmd_optimize(args, config: AnalysisConfig):
    """최적화 명령어 핸들러"""
    if not hasattr(args, "optimize_action") or not args.optimize_action:
        print("\n사용법: python main.py optimize {preview|run|export} -a 계정코드")
        return

    if args.optimize_action == "preview":
        n = args.count if hasattr(args, "count") else 10
        print(f"\n[미리보기] 계정: {args.account}, {n}개")
        result = batch_optimize(args.account, config, preview=n)
        if result.empty:
            return

        for _, r in result.iterrows():
            print(f"\n  [{r['등록상품ID']}]")
            if r["이름변경"]:
                print(f"  전: {r['원래상품명'][:60]}")
                print(f"  후: {r['최적화상품명'][:60]}")
            else:
                print(f"  이름: {r['원래상품명'][:60]} (변경 없음)")
            print(f"  검색어({r['검색어수']}): {r['최적화검색어'][:80]}")

    elif args.optimize_action == "run":
        print(f"\n[전체 최적화] 계정: {args.account}")
        result = batch_optimize(args.account, config)
        if result.empty:
            return

        # 결과 요약
        changed = result[result["이름변경"] == "O"]
        print(f"\n  결과 요약:")
        print(f"    상품명 변경: {len(changed)}/{len(result)}개")
        print(f"    검색어 평균: {result['검색어수'].mean():.1f}개")
        print(f"\n  Wing Excel로 내보내려면: python main.py optimize export -a {args.account}")

    elif args.optimize_action == "export":
        print(f"\n[Wing Excel 내보내기] 계정: {args.account}")
        path = export_wing_excel(args.account, config=config)
        if path:
            print(f"\n  이 파일을 Wing에 업로드하세요: {path}")

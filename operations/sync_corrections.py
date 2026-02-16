"""계정 간 상품 수정사항 동기화 도구.

007-ez 계정에서 수정한 상품명/가격/이미지를
다른 계정(002-bm, 007-bm, 007-book, big6ceo)에 일괄 반영.

3단계 매핑: 바코드(ISBN) → 등록상품명 → 퍼지매칭
"""

import os
import sys
import re
import difflib
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from core.constants import calc_original_price, ORIGINAL_PRICE_RATIO

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


# ─── 헬퍼 ────────────────────────────────────────────

def _safe_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "none" else s


def _safe_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


# ─── 데이터 로딩 ─────────────────────────────────────

def load_detailinfo(path: str) -> pd.DataFrame:
    """Wing detailinfo Excel 로드.

    Sheet='Template', header=Row4 (0-indexed=3).
    주요 컬럼: C1=등록상품ID, C2=등록상품명, C3=쿠팡 노출상품명,
              C4=카테고리, C5=제조사, C6=브랜드, C7=검색어, C231=바코드
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")

    # Template 시트, 헤더가 4번째 행 (0-indexed=3)
    try:
        df = pd.read_excel(path, sheet_name="Template", header=3)
    except ValueError:
        df = pd.read_excel(path, header=3)

    # 컬럼명 정리 (공백 제거)
    df.columns = [_safe_str(c) for c in df.columns]

    # 빈 행 제거
    id_col = _find_column(df, ["등록상품ID", "업체상품ID", "업체상품 ID"])
    if id_col:
        df = df.dropna(subset=[id_col])
        df = df[df[id_col].apply(lambda x: _safe_str(x) != "")]

    return df


def load_price_inventory(path: str) -> pd.DataFrame:
    """Wing price_inventory Excel 로드.

    Sheet='data', header=Row3 (0-indexed=2).
    주요 컬럼: C1=업체상품 ID, C2=Product ID, C3=옵션 ID,
              C5=바코드, C7=쿠팡 노출 상품명, C8=업체 등록 상품명, C10=판매가격
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")

    try:
        df = pd.read_excel(path, sheet_name="data", header=2)
    except ValueError:
        df = pd.read_excel(path, header=2)

    df.columns = [_safe_str(c) for c in df.columns]

    id_col = _find_column(df, ["업체상품 ID", "업체상품ID"])
    if id_col:
        df = df.dropna(subset=[id_col])
        df = df[df[id_col].apply(lambda x: _safe_str(x) != "")]

    return df


def _find_column(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """DataFrame에서 후보 컬럼명 중 존재하는 것 반환."""
    for c in candidates:
        if c in df.columns:
            return c
        # 부분 매칭
        for col in df.columns:
            if c in col:
                return col
    return None


# ─── 매핑 ─────────────────────────────────────────────

def build_mapping(source_detail: pd.DataFrame,
                  source_price: pd.DataFrame,
                  target_detail: pd.DataFrame,
                  target_price: pd.DataFrame) -> dict:
    """3단계 매핑 구축.

    Returns: {
        "matched": [{
            "source_spid": str, "target_spid": str,
            "match_key": "barcode"|"registered_name"|"fuzzy",
            "source_name": str, "target_name": str,
            "source_price": int, "target_price": int,
            "name_changed": bool, "price_changed": bool,
            "fuzzy_score": float (퍼지 매칭 시),
        }],
        "unmatched_source": [...],  # 007-ez에만 있는 상품
        "unmatched_target": [...],  # 타겟에만 있는 상품
        "stats": {"barcode": N, "registered_name": N, "fuzzy": N,
                  "unmatched_source": N, "unmatched_target": N}
    }
    """
    # 소스/타겟 상품 데이터 통합
    source_products = _merge_detail_price(source_detail, source_price)
    target_products = _merge_detail_price(target_detail, target_price)

    print(f"  소스 상품: {len(source_products)}개")
    print(f"  타겟 상품: {len(target_products)}개")

    matched = []
    stats = {"barcode": 0, "registered_name": 0, "fuzzy": 0,
             "unmatched_source": 0, "unmatched_target": 0}

    # 이미 매핑된 타겟 SPID 추적
    matched_source_spids = set()
    matched_target_spids = set()

    # ── Pass 1: 바코드(ISBN) 매칭 ──
    target_by_barcode: dict[str, dict] = {}
    for tp in target_products:
        bc = _safe_str(tp.get("barcode", ""))
        if bc and len(bc) >= 8:  # 최소 8자리 이상만 바코드로 인정
            target_by_barcode[bc] = tp

    for sp in source_products:
        bc = _safe_str(sp.get("barcode", ""))
        if bc and bc in target_by_barcode:
            tp = target_by_barcode[bc]
            if tp["spid"] not in matched_target_spids:
                entry = _build_match_entry(sp, tp, "barcode")
                matched.append(entry)
                matched_source_spids.add(sp["spid"])
                matched_target_spids.add(tp["spid"])
                stats["barcode"] += 1

    print(f"  Pass 1 (바코드): {stats['barcode']}개 매핑")

    # ── Pass 2: 등록상품명 매칭 ──
    target_by_regname: dict[str, dict] = {}
    for tp in target_products:
        if tp["spid"] in matched_target_spids:
            continue
        regname = _safe_str(tp.get("registered_name", ""))
        if regname:
            target_by_regname[regname] = tp

    for sp in source_products:
        if sp["spid"] in matched_source_spids:
            continue
        regname = _safe_str(sp.get("registered_name", ""))
        if regname and regname in target_by_regname:
            tp = target_by_regname[regname]
            entry = _build_match_entry(sp, tp, "registered_name")
            matched.append(entry)
            matched_source_spids.add(sp["spid"])
            matched_target_spids.add(tp["spid"])
            stats["registered_name"] += 1

    print(f"  Pass 2 (등록상품명): {stats['registered_name']}개 매핑")

    # ── Pass 3: 퍼지 매칭 (80%+ 만 자동) ──
    unmatched_source_list = [sp for sp in source_products
                             if sp["spid"] not in matched_source_spids]
    unmatched_target_list = [tp for tp in target_products
                             if tp["spid"] not in matched_target_spids]

    if unmatched_source_list and unmatched_target_list:
        # 타겟 핵심 이름 → 타겟 상품 매핑
        target_core_map: dict[str, list[dict]] = {}
        for tp in unmatched_target_list:
            core = _extract_core_name(_safe_str(tp.get("display_name", "")))
            if core:
                target_core_map.setdefault(core, []).append(tp)

        for sp in unmatched_source_list:
            if sp["spid"] in matched_source_spids:
                continue
            source_core = _extract_core_name(_safe_str(sp.get("display_name", "")))
            if not source_core:
                continue

            best_score = 0.0
            best_target = None

            for target_core, tps in target_core_map.items():
                score = _fuzzy_score(source_core, target_core)
                if score > best_score:
                    best_score = score
                    # 가장 첫 번째 미매핑 타겟
                    for tp in tps:
                        if tp["spid"] not in matched_target_spids:
                            best_target = tp
                            break

            if best_score >= 0.93 and best_target:
                entry = _build_match_entry(sp, best_target, "fuzzy")
                entry["fuzzy_score"] = round(best_score, 3)
                matched.append(entry)
                matched_source_spids.add(sp["spid"])
                matched_target_spids.add(best_target["spid"])
                stats["fuzzy"] += 1

    print(f"  Pass 3 (퍼지): {stats['fuzzy']}개 매핑")

    # ── 미매핑 집계 ──
    unmatched_source = []
    for sp in source_products:
        if sp["spid"] not in matched_source_spids:
            unmatched_source.append({
                "spid": sp["spid"],
                "display_name": sp.get("display_name", ""),
                "registered_name": sp.get("registered_name", ""),
                "barcode": sp.get("barcode", ""),
                "price": sp.get("price"),
            })

    unmatched_target = []
    for tp in target_products:
        if tp["spid"] not in matched_target_spids:
            unmatched_target.append({
                "spid": tp["spid"],
                "display_name": tp.get("display_name", ""),
                "registered_name": tp.get("registered_name", ""),
                "barcode": tp.get("barcode", ""),
                "price": tp.get("price"),
            })

    stats["unmatched_source"] = len(unmatched_source)
    stats["unmatched_target"] = len(unmatched_target)

    total_matched = stats["barcode"] + stats["registered_name"] + stats["fuzzy"]
    print(f"\n  총 매핑: {total_matched}개")
    print(f"  미매핑 (소스): {stats['unmatched_source']}개")
    print(f"  미매핑 (타겟): {stats['unmatched_target']}개")

    return {
        "matched": matched,
        "unmatched_source": unmatched_source,
        "unmatched_target": unmatched_target,
        "stats": stats,
    }


def _merge_detail_price(detail_df: pd.DataFrame,
                        price_df: pd.DataFrame) -> list[dict]:
    """detailinfo + price_inventory를 SPID 기준으로 통합."""
    products = []

    # detailinfo 컬럼 찾기 (Row4 헤더)
    d_spid_col = _find_column(detail_df, ["등록상품ID", "업체상품ID", "업체상품 ID"])
    d_display_col = _find_column(detail_df, ["쿠팡 노출상품명", "노출상품명"])
    d_reg_col = _find_column(detail_df, ["등록상품명"])
    d_barcode_col = _find_column(detail_df, ["바코드"])
    d_category_col = _find_column(detail_df, ["카테고리"])
    d_pid_col = _find_column(detail_df, ["노출상품ID", "Product ID"])
    d_option_col = _find_column(detail_df, ["옵션 ID", "옵션ID"])

    # price_inventory 컬럼 찾기 (Row3 헤더)
    p_spid_col = _find_column(price_df, ["업체상품 ID", "업체상품ID"])
    p_price_col = _find_column(price_df, ["판매가격"])
    p_barcode_col = _find_column(price_df, ["바코드"])
    p_name_col = _find_column(price_df, ["쿠팡 노출 상품명", "상품명"])
    p_reg_col = _find_column(price_df, ["업체 등록 상품명"])

    # price를 SPID → 가격/바코드/등록상품명 dict로 변환
    price_map: dict[str, int] = {}
    price_barcode_map: dict[str, str] = {}
    price_regname_map: dict[str, str] = {}
    if p_spid_col and p_price_col:
        for _, row in price_df.iterrows():
            spid = _safe_str(row.get(p_spid_col, ""))
            if spid:
                price = _safe_int(row.get(p_price_col))
                if price is not None:
                    price_map[spid] = price
                bc = _safe_str(row.get(p_barcode_col, "")) if p_barcode_col else ""
                if bc:
                    price_barcode_map[spid] = bc
                rn = _safe_str(row.get(p_reg_col, "")) if p_reg_col else ""
                if rn:
                    price_regname_map[spid] = rn

    # detailinfo 기준으로 상품 목록 구성
    if not d_spid_col:
        return products

    seen_spids = set()
    for _, row in detail_df.iterrows():
        spid = _safe_str(row.get(d_spid_col, ""))
        if not spid or spid in seen_spids:
            continue
        seen_spids.add(spid)

        display = _safe_str(row.get(d_display_col, "")) if d_display_col else ""
        registered = _safe_str(row.get(d_reg_col, "")) if d_reg_col else ""
        barcode = _safe_str(row.get(d_barcode_col, "")) if d_barcode_col else ""
        category = _safe_str(row.get(d_category_col, "")) if d_category_col else ""
        pid = _safe_str(row.get(d_pid_col, "")) if d_pid_col else ""
        option_id = _safe_str(row.get(d_option_col, "")) if d_option_col else ""

        # 바코드가 detailinfo에 없으면 price_inventory에서 가져오기
        if not barcode and spid in price_barcode_map:
            barcode = price_barcode_map[spid]

        # 등록상품명: detailinfo의 등록상품명이 노출상품명과 같으면
        # price_inventory의 업체 등록 상품명을 우선 사용 (원본일 가능성 높음)
        if registered == display and spid in price_regname_map:
            registered = price_regname_map[spid]
        elif not registered and spid in price_regname_map:
            registered = price_regname_map[spid]

        price = price_map.get(spid)

        products.append({
            "spid": spid,
            "display_name": display,
            "registered_name": registered,
            "barcode": barcode,
            "category": category,
            "pid": pid,
            "option_id": option_id,
            "price": price,
        })

    return products


def _build_match_entry(source: dict, target: dict, match_key: str) -> dict:
    """매칭 결과 엔트리 생성."""
    s_name = source.get("display_name", "")
    t_name = target.get("display_name", "")
    s_price = source.get("price")
    t_price = target.get("price")

    return {
        "source_spid": source["spid"],
        "target_spid": target["spid"],
        "match_key": match_key,
        "source_name": s_name,
        "target_name": t_name,
        "source_registered_name": source.get("registered_name", ""),
        "target_registered_name": target.get("registered_name", ""),
        "source_barcode": source.get("barcode", ""),
        "source_price": s_price,
        "target_price": t_price,
        "name_changed": s_name != t_name and bool(s_name),
        "price_changed": s_price != t_price and s_price is not None,
    }


def _extract_core_name(name: str) -> str:
    """상품명에서 핵심 키워드 추출 (세트/묶음/권수 등 제거).

    예: '2025 EBS 수능특강 국어 세트(2권)' → 'EBS 수능특강 국어'
    """
    if not name:
        return ""

    # 연도 제거
    s = re.sub(r'20\d{2}\s*', '', name)
    # 괄호 내용 제거
    s = re.sub(r'\([^)]*\)', '', s)
    s = re.sub(r'\[[^\]]*\]', '', s)
    # 세트/묶음 관련 키워드 제거
    s = re.sub(r'세트|묶음|패키지|전\s*\d+\s*권|총\s*\d+\s*권|\d+권|\d+\s*세트', '', s)
    # 특수문자/여분 공백 정리
    s = re.sub(r'[^\w\s가-힣a-zA-Z]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()

    return s


def _fuzzy_score(name_a: str, name_b: str) -> float:
    """두 상품명의 유사도 점수 (0~1). difflib.SequenceMatcher 사용."""
    if not name_a or not name_b:
        return 0.0
    return difflib.SequenceMatcher(None, name_a, name_b).ratio()


# ─── 수정 Excel 생성 ──────────────────────────────────

def generate_corrected_detailinfo(mapping: dict,
                                  source_detail: pd.DataFrame,
                                  target_detailinfo_path: str,
                                  output_path: str,
                                  include_fuzzy: bool = False) -> str:
    """타겟 detailinfo Excel에 007-ez의 수정된 노출상품명 반영.

    openpyxl로 원본 로드 → 매핑된 행의 노출상품명 컬럼만 수정 → 저장.
    include_fuzzy=False(기본)이면 퍼지 매칭은 제외 (오매핑 위험).
    """
    if not os.path.exists(target_detailinfo_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {target_detailinfo_path}")

    # 매핑에서 이름이 변경된 것만 필터
    name_changes: dict[str, str] = {}  # target_spid → new_name
    for m in mapping["matched"]:
        if m["name_changed"]:
            if m["match_key"] == "fuzzy" and not include_fuzzy:
                continue
            name_changes[m["target_spid"]] = m["source_name"]

    if not name_changes:
        print("  상품명 변경 사항 없음")
        return ""

    # openpyxl로 원본 로드
    wb = load_workbook(target_detailinfo_path)
    try:
        ws = wb["Template"]
    except KeyError:
        ws = wb.active

    # 헤더 찾기 (4번째 행, openpyxl 1-based = row 4)
    header = [_safe_str(ws.cell(row=4, column=c).value)
              for c in range(1, ws.max_column + 1)]

    # 컬럼 인덱스 찾기 (1-based)
    spid_col = _find_col_index(header, ["등록상품ID", "업체상품ID", "업체상품 ID"])
    name_col = _find_col_index(header, ["쿠팡 노출상품명", "노출상품명"])

    if not spid_col or not name_col:
        raise ValueError(f"필수 컬럼을 찾을 수 없습니다. 헤더: {header[:15]}")

    # 데이터 행은 5번째부터 (헤더가 4행)
    modified = 0
    for row_num in range(5, ws.max_row + 1):
        spid = _safe_str(ws.cell(row=row_num, column=spid_col).value)
        if spid in name_changes:
            old_name = _safe_str(ws.cell(row=row_num, column=name_col).value)
            new_name = name_changes[spid]
            ws.cell(row=row_num, column=name_col).value = new_name
            modified += 1

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()

    print(f"  detailinfo 수정 완료: {modified}개 상품명 변경 → {output_path}")
    return output_path


def generate_corrected_price(mapping: dict,
                             source_price: pd.DataFrame,
                             target_price_path: str,
                             output_path: str,
                             include_fuzzy: bool = False) -> str:
    """타겟 price_inventory Excel에 007-ez의 수정된 가격 반영.

    openpyxl로 원본 로드 → 매핑된 행의 판매가격 컬럼만 수정 → 저장.
    include_fuzzy=False(기본)이면 퍼지 매칭은 제외.
    """
    if not os.path.exists(target_price_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {target_price_path}")

    # 매핑에서 가격이 변경된 것만 필터
    price_changes: dict[str, int] = {}  # target_spid → new_price
    for m in mapping["matched"]:
        if m["price_changed"] and m["source_price"] is not None:
            if m["match_key"] == "fuzzy" and not include_fuzzy:
                continue
            price_changes[m["target_spid"]] = m["source_price"]

    if not price_changes:
        print("  가격 변경 사항 없음")
        return ""

    # openpyxl로 원본 로드
    wb = load_workbook(target_price_path)
    try:
        ws = wb["data"]
    except KeyError:
        ws = wb.active

    # 헤더 찾기 (3번째 행, openpyxl 1-based = row 3)
    header = [_safe_str(ws.cell(row=3, column=c).value)
              for c in range(1, ws.max_column + 1)]

    spid_col = _find_col_index(header, ["업체상품 ID", "업체상품ID"])
    # 수정 요청용 판매가격 컬럼 (O열=16, "변경/수정 요청" 섹션)
    # Row3 헤더에서 두 번째 "판매가격" 찾기 (첫 번째는 C10=조회용, 두 번째는 C16=수정용)
    price_col = None
    for i, h in enumerate(header):
        if "판매가격" in h and i >= 15:  # C16 이후 (0-indexed=15)
            price_col = i + 1
            break
    if not price_col:
        # 못 찾으면 C16 직접 사용 (수정 요청 판매가격)
        price_col = 16

    if not spid_col:
        raise ValueError(f"SPID 컬럼을 찾을 수 없습니다. 헤더: {header[:15]}")

    # 데이터 행은 4번째부터
    modified = 0
    for row_num in range(4, ws.max_row + 1):
        spid = _safe_str(ws.cell(row=row_num, column=spid_col).value)
        if spid in price_changes:
            ws.cell(row=row_num, column=price_col).value = price_changes[spid]
            modified += 1

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()

    print(f"  price_inventory 수정 완료: {modified}개 가격 변경 → {output_path}")
    return output_path


def _find_col_index(header: list[str], candidates: list[str]) -> Optional[int]:
    """헤더 리스트에서 후보 컬럼명의 1-based 인덱스 반환."""
    for c in candidates:
        for i, h in enumerate(header):
            if c == h or c in h:
                return i + 1  # 1-based
    return None


# ─── 리포트 ───────────────────────────────────────────

def generate_mapping_report(mapping: dict, output_path: str) -> str:
    """매핑 결과 Excel 리포트 생성.

    시트1: 매핑된 상품 (source<->target, 변경 내용)
    시트2: 미매핑 소스 상품 (갭 상품)
    시트3: 미매핑 타겟 상품
    시트4: 통계 요약
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 시트1: 매핑된 상품
        if mapping["matched"]:
            matched_data = []
            for m in mapping["matched"]:
                matched_data.append({
                    "소스_SPID": m["source_spid"],
                    "타겟_SPID": m["target_spid"],
                    "매칭방법": m["match_key"],
                    "소스_상품명": m["source_name"],
                    "타겟_상품명": m["target_name"],
                    "소스_가격": m.get("source_price"),
                    "타겟_가격": m.get("target_price"),
                    "상품명변경": "O" if m["name_changed"] else "",
                    "가격변경": "O" if m["price_changed"] else "",
                    "유사도": m.get("fuzzy_score", ""),
                    "소스_바코드": m.get("source_barcode", ""),
                })
            pd.DataFrame(matched_data).to_excel(
                writer, sheet_name="매핑결과", index=False)

        # 시트2: 미매핑 소스 (갭 상품)
        if mapping["unmatched_source"]:
            pd.DataFrame(mapping["unmatched_source"]).to_excel(
                writer, sheet_name="갭_소스에만", index=False)

        # 시트3: 미매핑 타겟
        if mapping["unmatched_target"]:
            pd.DataFrame(mapping["unmatched_target"]).to_excel(
                writer, sheet_name="미매핑_타겟에만", index=False)

        # 시트4: 통계
        stats = mapping["stats"]
        name_changed = sum(1 for m in mapping["matched"] if m["name_changed"])
        price_changed = sum(1 for m in mapping["matched"] if m["price_changed"])

        stats_data = [
            {"항목": "바코드 매칭", "건수": stats["barcode"]},
            {"항목": "등록상품명 매칭", "건수": stats["registered_name"]},
            {"항목": "퍼지 매칭", "건수": stats["fuzzy"]},
            {"항목": "총 매핑", "건수": stats["barcode"] + stats["registered_name"] + stats["fuzzy"]},
            {"항목": "", "건수": ""},
            {"항목": "상품명 변경 대상", "건수": name_changed},
            {"항목": "가격 변경 대상", "건수": price_changed},
            {"항목": "", "건수": ""},
            {"항목": "미매핑 (소스=갭)", "건수": stats["unmatched_source"]},
            {"항목": "미매핑 (타겟)", "건수": stats["unmatched_target"]},
        ]
        pd.DataFrame(stats_data).to_excel(
            writer, sheet_name="통계", index=False)

    print(f"  매핑 리포트 생성: {output_path}")
    return output_path


# ─── 갭 채우기 (Phase B): 미등록 상품 신규 등록 ──────

def register_gap_products(source_account: str, target_account: str,
                          mapping: dict, *,
                          dry_run: bool = True,
                          test_limit: int = 0) -> dict:
    """007-ez에만 있고 타겟에 없는 상품을 API로 신규 등록.

    1. mapping["unmatched_source"]에서 미등록 상품 목록 추출
    2. 007-ez API (get_product_by_id)로 수정된 상품 상세 조회
    3. 타겟 계정 API (create_product)로 등록

    Returns: {"gap": N, "created": N, "skipped": N, "error": N, "errors": [...]}
    """
    from operations.product_api import _get_client, ACCOUNTS

    gap_items = mapping.get("unmatched_source", [])
    result = {
        "gap": len(gap_items),
        "created": 0,
        "skipped": 0,
        "error": 0,
        "errors": [],
    }

    if not gap_items:
        print("  갭 상품 없음")
        return result

    if test_limit > 0:
        gap_items = gap_items[:test_limit]

    source_client = _get_client(source_account)
    target_client = _get_client(target_account)
    target_vendor_id = ACCOUNTS[target_account]["vendor_id"]

    mode = "[미리보기]" if dry_run else "[등록]"
    print(f"\n{mode} 갭 상품 {len(gap_items)}개 처리...")

    for i, item in enumerate(gap_items):
        spid = item["spid"]
        name = item.get("display_name", "")

        # 소스 API에서 상품 상세 조회
        try:
            detail = source_client.get_product_by_id(str(spid))
            if detail.get("code") == "ERROR":
                result["errors"].append(f"SPID {spid}: 조회 실패 - {detail.get('message', '')[:60]}")
                result["error"] += 1
                continue

            data = detail.get("data", detail)
            body = _clone_product_body(data, target_vendor_id, target_account)

        except Exception as e:
            result["errors"].append(f"SPID {spid}: {str(e)[:60]}")
            result["error"] += 1
            continue

        prod_name = body.get("sellerProductName", name)

        if dry_run:
            if result["created"] < 10:
                items_data = body.get("items", [{}])
                price = items_data[0].get("salePrice", 0) if items_data else 0
                print(f"  [{i+1}] {prod_name[:60]}")
                print(f"    {source_account} → {target_account}  가격: {price:,}원")
            result["created"] += 1
            continue

        try:
            resp = target_client.create_product(body)
            code = resp.get("code", "")
            if code == "ERROR":
                msg = resp.get("message", "")
                result["errors"].append(f"{prod_name[:40]}: {msg[:60]}")
                result["error"] += 1
            else:
                result["created"] += 1
                new_spid = resp.get("data", "")
                if (i + 1) <= 3 or (i + 1) % 20 == 0:
                    print(f"  [{i+1}] 등록: {prod_name[:50]} (SPID: {new_spid})")
        except Exception as e:
            result["errors"].append(f"{prod_name[:40]}: {str(e)[:60]}")
            result["error"] += 1

    return result


# 계정별 배송/반품 설정 (API 조회 결과 캐시)
ACCOUNT_SHIPPING_INFO: dict[str, dict] = {}


def _get_account_shipping_info(account: str) -> dict:
    """계정의 배송/반품 설정을 API에서 조회 (캐시)."""
    if account in ACCOUNT_SHIPPING_INFO:
        return ACCOUNT_SHIPPING_INFO[account]

    from operations.product_api import _get_client
    client = _get_client(account)
    resp = client.get_seller_products(max_per_page=1)
    products = resp.get("data", [])
    if not products:
        return {}

    spid = products[0].get("sellerProductId", "")
    detail = client.get_product_by_id(str(spid))
    d = detail.get("data", detail)

    info = {
        "vendorUserId": d.get("vendorUserId", ""),
        "returnCenterCode": d.get("returnCenterCode", ""),
        "outboundShippingPlaceCode": d.get("outboundShippingPlaceCode", ""),
        "returnChargeName": d.get("returnChargeName", ""),
        "returnZipCode": d.get("returnZipCode", ""),
        "returnAddress": d.get("returnAddress", ""),
        "returnAddressDetail": d.get("returnAddressDetail", ""),
        "companyContactNumber": d.get("companyContactNumber", ""),
    }
    ACCOUNT_SHIPPING_INFO[account] = info
    return info


def _clone_product_body(source_detail: dict, target_vendor_id: str,
                        target_account: str = "") -> dict:
    """007-ez API 응답에서 타겟용 create_product body 구성.

    소스 상품 데이터를 전체 복제하되:
    - vendorId → 타겟 계정으로 교체
    - vendorUserId, returnCenterCode, outboundShippingPlaceCode → 타겟 것으로 교체
    - 소스 고유 ID (sellerProductId, productId, vendorItemId 등) 제거
    - 읽기전용 필드 (status, statusName, mdId 등) 제거
    """
    import copy
    body = copy.deepcopy(source_detail)

    # 타겟 vendorId로 교체
    body["vendorId"] = target_vendor_id

    # 타겟 계정의 배송/반품 설정 교체
    if target_account:
        shipping = _get_account_shipping_info(target_account)
        if shipping:
            for k, v in shipping.items():
                if v:
                    body[k] = v

    # 소스 고유 ID / 읽기전용 필드 제거
    for key in [
        "sellerProductId", "productId", "categoryId", "trackingId",
        "displayProductName", "generalProductName",
        "mdId", "mdName", "statusName", "status",
        "contributorType", "requested",
        "requiredDocuments", "extraInfoMessage",
        "roleCode", "multiShippingInfos", "multiReturnInfos",
    ]:
        body.pop(key, None)

    # items에서 소스 고유 ID 제거
    for item in body.get("items", []):
        for key in [
            "sellerProductItemId", "vendorItemId", "itemId",
            "supplyPrice", "saleAgentCommission",
            "isAutoGenerated", "freePriceType",
            "bestPriceGuaranteed3P",
        ]:
            item.pop(key, None)

    return body


# ─── API 일괄 적용 (이름+가격+갭+이미지 통합) ────────

def apply_corrections(source_account: str, target_account: str,
                      mapping: dict, *,
                      dry_run: bool = True,
                      test_limit: int = 0,
                      include_fuzzy: bool = False,
                      skip_images: bool = False) -> dict:
    """매핑 결과를 API로 일괄 적용.

    1단계: 기존 상품 가격 수정 (PATCH - 승인 유지)
    2단계: 기존 상품 이름+이미지 수정 (PUT - 임시저장 → 재승인 필요)
    3단계: 갭 상품 신규 등록 (POST - 007-ez에서 복제)

    Returns: {
        "price_updated": N, "name_updated": N, "image_updated": N,
        "gap_created": N, "skipped": N, "error": N, "errors": [...]
    }
    """
    from operations.product_api import _get_client, ACCOUNTS
    import time

    source_client = _get_client(source_account)
    target_client = _get_client(target_account)
    target_vendor_id = ACCOUNTS[target_account]["vendor_id"]

    result = {
        "price_updated": 0, "name_updated": 0, "image_updated": 0,
        "gap_created": 0, "skipped": 0, "error": 0, "errors": [],
    }

    safe_keys = {"barcode", "registered_name"}
    if include_fuzzy:
        safe_keys.add("fuzzy")

    matched = [m for m in mapping.get("matched", [])
               if m["match_key"] in safe_keys]
    gap_items = mapping.get("unmatched_source", [])

    # 변경 대상 분류
    price_targets = [m for m in matched if m["price_changed"]]
    name_targets = [m for m in matched if m["name_changed"]]

    total_work = len(price_targets) + len(name_targets) + len(gap_items)
    if test_limit > 0:
        price_targets = price_targets[:test_limit]
        name_targets = name_targets[:test_limit]
        gap_items = gap_items[:test_limit]
        total_work = len(price_targets) + len(name_targets) + len(gap_items)

    mode = "[미리보기]" if dry_run else "[적용]"
    print(f"\n{mode} {source_account} → {target_account}")
    print(f"  가격 수정: {len(price_targets)}개 (PATCH, 승인 유지)")
    print(f"  이름 수정: {len(name_targets)}개 (PUT, 재승인 필요)")
    print(f"  갭 등록:   {len(gap_items)}개 (POST, 신규)")
    if not include_fuzzy:
        fuzzy_skipped = sum(1 for m in mapping.get("matched", [])
                           if m["match_key"] == "fuzzy"
                           and (m["price_changed"] or m["name_changed"]))
        if fuzzy_skipped:
            print(f"  퍼지 제외: {fuzzy_skipped}개 (--include-fuzzy로 포함)")

    # ── 1단계: 가격 PATCH (승인 상태 유지) ──
    if price_targets:
        print(f"\n  ── 1단계: 가격 수정 ({len(price_targets)}개) ──")
        for i, m in enumerate(price_targets):
            t_spid = m["target_spid"]
            new_price = m["source_price"]
            name = m.get("target_name", "")[:45]

            if dry_run:
                if i < 10:
                    old_p = m.get("target_price", 0) or 0
                    print(f"    [{i+1}] {name}  {old_p:,} → {new_price:,}원")
                result["price_updated"] += 1
                continue

            try:
                detail = target_client.get_product_by_id(str(t_spid))
                items = detail.get("data", {}).get("items", [])
                if not items:
                    result["errors"].append(f"SPID {t_spid}: items 없음")
                    result["error"] += 1
                    continue

                # [세트물 안전 주의] 모든 items에 동일 가격을 적용합니다.
                # 세트물의 경우 옵션별 가격이 다를 수 있으므로,
                # 동기화 대상이 세트물인지 반드시 확인하세요.
                # BUG FIX: 1.11 하드코딩 → calc_original_price() 사용
                patch_items = []
                for item in items:
                    patch_items.append({
                        "vendorItemId": item["vendorItemId"],
                        "salePrice": new_price,
                        "originalPrice": calc_original_price(new_price),
                    })

                body = {"sellerProductId": int(t_spid), "items": patch_items}
                resp = target_client.patch_product(str(t_spid), body)

                if resp.get("code") == "ERROR":
                    result["errors"].append(f"SPID {t_spid}: {resp.get('message', '')[:60]}")
                    result["error"] += 1
                else:
                    result["price_updated"] += 1
                    if (i + 1) <= 3 or (i + 1) % 50 == 0:
                        print(f"    [{i+1}] 가격: {name}  → {new_price:,}원")

                time.sleep(0.3)  # API rate limit
            except Exception as e:
                result["errors"].append(f"SPID {t_spid}: {str(e)[:60]}")
                result["error"] += 1

    # ── 2단계: 이름(+이미지) PUT (재승인 필요) ──
    if name_targets:
        print(f"\n  ── 2단계: 이름 수정 ({len(name_targets)}개) ──")
        for i, m in enumerate(name_targets):
            s_spid = m["source_spid"]
            t_spid = m["target_spid"]
            new_name = m["source_name"]
            old_name = m.get("target_name", "")[:40]

            if dry_run:
                if i < 10:
                    print(f"    [{i+1}] {old_name}")
                    print(f"         → {new_name[:55]}")
                result["name_updated"] += 1
                continue

            try:
                # 타겟 상품 상세 조회
                t_resp = target_client.get_product_by_id(str(t_spid))
                t_data = t_resp.get("data", t_resp)

                # 이름 변경
                t_data["sellerProductName"] = new_name
                # [세트물 안전 주의] 모든 items의 itemName을 동일하게 변경합니다.
                # 세트물의 경우 옵션별 itemName이 다를 수 있으므로 주의하세요.
                for item in t_data.get("items", []):
                    item["itemName"] = new_name

                # 이미지도 동기화 (skip_images가 아니면)
                if not skip_images:
                    try:
                        s_resp = source_client.get_product_by_id(str(s_spid))
                        s_data = s_resp.get("data", s_resp)
                        s_items = s_data.get("items", [])
                        t_items = t_data.get("items", [])

                        s_urls = _extract_image_urls(
                            s_items[0].get("images", []) if s_items else [])
                        t_urls = _extract_image_urls(
                            t_items[0].get("images", []) if t_items else [])

                        if s_urls != t_urls:
                            for j, t_item in enumerate(t_items):
                                if j < len(s_items):
                                    t_item["images"] = s_items[j].get("images", [])
                            result["image_updated"] += 1
                    except Exception:
                        pass  # 이미지 실패해도 이름은 진행

                # PUT body에서 읽기전용 필드 제거 (sellerProductId는 유지!)
                for key in [
                    "productId", "categoryId", "trackingId",
                    "displayProductName", "generalProductName",
                    "mdId", "mdName", "statusName",
                    "contributorType", "requested",
                    "requiredDocuments", "extraInfoMessage",
                    "roleCode", "multiShippingInfos", "multiReturnInfos",
                ]:
                    t_data.pop(key, None)
                for item in t_data.get("items", []):
                    for key in [
                        "supplyPrice", "saleAgentCommission",
                        "isAutoGenerated", "freePriceType",
                        "bestPriceGuaranteed3P",
                    ]:
                        item.pop(key, None)

                resp = target_client.update_product(str(t_spid), t_data)
                if resp.get("code") == "ERROR":
                    result["errors"].append(f"SPID {t_spid}: {resp.get('message', '')[:60]}")
                    result["error"] += 1
                else:
                    result["name_updated"] += 1
                    if (i + 1) <= 3 or (i + 1) % 50 == 0:
                        print(f"    [{i+1}] 이름+이미지: {new_name[:50]}")

                time.sleep(0.5)  # PUT은 더 느리게
            except Exception as e:
                result["errors"].append(f"SPID {t_spid}: {str(e)[:60]}")
                result["error"] += 1

    # ── 3단계: 갭 상품 등록 ──
    if gap_items:
        print(f"\n  ── 3단계: 갭 상품 등록 ({len(gap_items)}개) ──")
        for i, item in enumerate(gap_items):
            spid = item["spid"]
            name = item.get("display_name", "")

            try:
                detail = source_client.get_product_by_id(str(spid))
                if detail.get("code") == "ERROR":
                    result["errors"].append(f"SPID {spid}: 조회실패")
                    result["error"] += 1
                    continue

                data = detail.get("data", detail)
                body = _clone_product_body(data, target_vendor_id, target_account)
                prod_name = body.get("sellerProductName", name)

                if dry_run:
                    if i < 10:
                        items_data = body.get("items", [{}])
                        price = items_data[0].get("salePrice", 0) if items_data else 0
                        print(f"    [{i+1}] {prod_name[:55]}  {price:,}원")
                    result["gap_created"] += 1
                    continue

                resp = target_client.create_product(body)
                if resp.get("code") == "ERROR":
                    msg = resp.get("message", "")
                    result["errors"].append(f"{prod_name[:35]}: {msg[:50]}")
                    result["error"] += 1
                else:
                    result["gap_created"] += 1
                    if (i + 1) <= 3 or (i + 1) % 50 == 0:
                        new_spid = resp.get("data", "")
                        print(f"    [{i+1}] 등록: {prod_name[:45]} (SPID:{new_spid})")

                time.sleep(0.5)
            except Exception as e:
                result["errors"].append(f"SPID {spid}: {str(e)[:60]}")
                result["error"] += 1

    return result


# ─── 이미지 동기화 (Phase C) ──────────────────────────

def sync_images(source_account: str, target_account: str,
                mapping: dict, *, dry_run: bool = True,
                test_limit: int = 0) -> dict:
    """매핑된 기존 상품의 이미지를 API로 동기화.

    1. 007-ez API에서 상품 상세 → 이미지 URL 추출
    2. 타겟 API에서 상품 상세 조회
    3. 이미지가 다르면 PUT으로 교체
    ※ PUT 사용 → 상태가 '임시저장'으로 변경됨 → Wing에서 일괄 재승인 필요

    Returns: {"total": N, "changed": N, "same": N, "error": N, "errors": [...]}
    """
    from operations.product_api import _get_client

    result = {
        "total": 0,
        "changed": 0,
        "same": 0,
        "error": 0,
        "errors": [],
    }

    matched = mapping.get("matched", [])
    if not matched:
        print("  매핑된 상품 없음")
        return result

    # 매핑된 상품만 대상
    targets = matched
    if test_limit > 0:
        targets = targets[:test_limit]

    result["total"] = len(targets)

    source_client = _get_client(source_account)
    target_client = _get_client(target_account)

    mode = "[미리보기]" if dry_run else "[동기화]"
    print(f"\n{mode} 이미지 동기화 {len(targets)}개 상품...")

    for i, m in enumerate(targets):
        s_spid = m["source_spid"]
        t_spid = m["target_spid"]

        try:
            # 소스 상품 상세
            s_resp = source_client.get_product_by_id(str(s_spid))
            s_data = s_resp.get("data", s_resp)
            s_items = s_data.get("items", [])
            s_images = s_items[0].get("images", []) if s_items else []

            # 타겟 상품 상세
            t_resp = target_client.get_product_by_id(str(t_spid))
            t_data = t_resp.get("data", t_resp)
            t_items = t_data.get("items", [])
            t_images = t_items[0].get("images", []) if t_items else []

            # 이미지 URL 비교
            s_urls = _extract_image_urls(s_images)
            t_urls = _extract_image_urls(t_images)

            if s_urls == t_urls:
                result["same"] += 1
                continue

            name = m.get("source_name", "")[:50]

            if dry_run:
                if result["changed"] < 10:
                    print(f"  [{i+1}] {name}")
                    print(f"    소스 이미지: {len(s_images)}장 → 타겟 이미지: {len(t_images)}장")
                result["changed"] += 1
                continue

            # 타겟 상품의 이미지를 소스 이미지로 교체 (PUT = 전체 수정)
            for j, t_item in enumerate(t_items):
                if j < len(s_items):
                    t_item["images"] = s_items[j].get("images", [])

            resp = target_client.update_product(str(t_spid), t_data)
            code = resp.get("code", "")
            if code == "ERROR":
                msg = resp.get("message", "")
                result["errors"].append(f"SPID {t_spid}: {msg[:60]}")
                result["error"] += 1
            else:
                result["changed"] += 1
                if (i + 1) <= 3 or (i + 1) % 20 == 0:
                    print(f"  [{i+1}] 이미지 동기화: {name}")

        except Exception as e:
            result["errors"].append(f"SPID {t_spid}: {str(e)[:60]}")
            result["error"] += 1

    return result


def _extract_image_urls(images: list) -> list[str]:
    """이미지 리스트에서 URL만 추출하여 정렬."""
    urls = []
    for img in images:
        if isinstance(img, dict):
            url = img.get("imageUrl", img.get("cdnPath", ""))
        elif isinstance(img, str):
            url = img
        else:
            continue
        if url:
            urls.append(url)
    return sorted(urls)


# ─── CLI 핸들러 ───────────────────────────────────────

def cmd_sync(args, config):
    """메인 CLI 핸들러"""
    action = getattr(args, "sync_action", None)

    if not action:
        print("\nsync 하위 명령을 지정하세요:")
        print("  map     매핑 리포트 생성 (먼저 실행)")
        print("  apply   API로 일괄 적용 (가격PATCH + 이름PUT + 갭POST)")
        print("  fix     수정 Excel 생성 (Wing 수동 업로드용)")
        print("  gap     미등록 상품 신규 등록 (API)")
        print("  images  이미지 동기화 (API)")
        return

    if action == "map":
        _cmd_map(args, config)
    elif action == "apply":
        _cmd_apply(args, config)
    elif action == "fix":
        _cmd_fix(args, config)
    elif action == "gap":
        _cmd_gap(args, config)
    elif action == "images":
        _cmd_images(args, config)


def _cmd_map(args, config):
    """매핑 리포트 생성"""
    sd_path = args.source_detail
    sp_path = args.source_price
    td_path = args.target_detail
    tp_path = args.target_price

    print(f"\n[매핑] 소스 detailinfo: {os.path.basename(sd_path)}")
    print(f"       소스 price:      {os.path.basename(sp_path)}")
    print(f"       타겟 detailinfo: {os.path.basename(td_path)}")
    print(f"       타겟 price:      {os.path.basename(tp_path)}")

    # 데이터 로드
    source_detail = load_detailinfo(sd_path)
    source_price = load_price_inventory(sp_path)
    target_detail = load_detailinfo(td_path)
    target_price = load_price_inventory(tp_path)

    # 매핑 구축
    print(f"\n  3단계 매핑 시작...")
    mapping = build_mapping(source_detail, source_price, target_detail, target_price)

    # 리포트 생성
    output = getattr(args, "output", "") or ""
    if not output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = os.path.join("reports", f"mapping_report_{timestamp}.xlsx")

    generate_mapping_report(mapping, output)

    # 변경 요약
    name_cnt = sum(1 for m in mapping["matched"] if m["name_changed"])
    price_cnt = sum(1 for m in mapping["matched"] if m["price_changed"])
    print(f"\n  [변경 요약]")
    print(f"    상품명 변경: {name_cnt}개")
    print(f"    가격 변경:   {price_cnt}개")
    print(f"    갭 상품:     {mapping['stats']['unmatched_source']}개")


def _cmd_apply(args, config):
    """API로 일괄 적용 (가격+이름+이미지+갭)"""
    source = args.source
    target = args.target
    report_path = args.mapping_report
    dry_run = getattr(args, "dry_run", False)
    test = getattr(args, "test", 0)
    include_fuzzy = getattr(args, "include_fuzzy", False)
    skip_images = getattr(args, "skip_images", False)

    if not os.path.exists(report_path):
        print(f"\n  매핑 리포트를 찾을 수 없습니다: {report_path}")
        print("  먼저 sync map을 실행하세요.")
        return

    # 매핑 리포트에서 데이터 읽기
    matched_items = []
    try:
        matched_df = pd.read_excel(report_path, sheet_name="매핑결과")
        for _, row in matched_df.iterrows():
            matched_items.append({
                "source_spid": _safe_str(row.get("소스_SPID", "")),
                "target_spid": _safe_str(row.get("타겟_SPID", "")),
                "match_key": _safe_str(row.get("매칭방법", "")),
                "source_name": _safe_str(row.get("소스_상품명", "")),
                "target_name": _safe_str(row.get("타겟_상품명", "")),
                "source_price": _safe_int(row.get("소스_가격")),
                "target_price": _safe_int(row.get("타겟_가격")),
                "name_changed": row.get("상품명변경") == "O",
                "price_changed": row.get("가격변경") == "O",
            })
    except Exception as e:
        print(f"\n  매핑결과 시트 읽기 실패: {e}")
        return

    gap_items = []
    try:
        gap_df = pd.read_excel(report_path, sheet_name="갭_소스에만")
        for _, row in gap_df.iterrows():
            gap_items.append({
                "spid": _safe_str(row.get("spid", "")),
                "display_name": _safe_str(row.get("display_name", "")),
                "price": _safe_int(row.get("price")),
            })
    except Exception:
        pass  # 갭 시트가 없을 수 있음

    mapping = {"matched": matched_items, "unmatched_source": gap_items}

    result = apply_corrections(
        source, target, mapping,
        dry_run=dry_run, test_limit=test,
        include_fuzzy=include_fuzzy,
        skip_images=skip_images,
    )

    # 결과 요약
    print(f"\n  {'=' * 50}")
    print(f"  [결과] {source} → {target}")
    print(f"    가격 수정: {result['price_updated']}개")
    print(f"    이름 수정: {result['name_updated']}개")
    if result['image_updated']:
        print(f"    이미지 동기화: {result['image_updated']}개")
    print(f"    갭 등록:   {result['gap_created']}개")
    if result['error']:
        print(f"    오류:      {result['error']}개")
    print(f"  {'=' * 50}")

    if result["name_updated"] > 0 and not dry_run:
        print(f"\n  ※ 이름 수정된 {result['name_updated']}개 상품은 '임시저장' 상태")
        print(f"  ※ Wing에서 일괄 재승인 필요")

    if result["errors"]:
        print(f"\n  [오류 목록]")
        for err in result["errors"][:10]:
            print(f"    - {err}")
        if len(result["errors"]) > 10:
            print(f"    ... 외 {len(result['errors']) - 10}건")


def _cmd_fix(args, config):
    """수정 Excel 생성 (Phase A)"""
    sd_path = args.source_detail
    sp_path = args.source_price
    td_path = args.target_detail
    tp_path = args.target_price
    output_dir = getattr(args, "output", "") or "corrected"
    include_fuzzy = getattr(args, "include_fuzzy", False)

    print(f"\n[수정] Excel 생성...")
    if not include_fuzzy:
        print("  ※ 퍼지 매칭은 제외 (오매핑 위험). --include-fuzzy로 포함 가능")

    # 데이터 로드 & 매핑
    source_detail = load_detailinfo(sd_path)
    source_price = load_price_inventory(sp_path)
    target_detail = load_detailinfo(td_path)
    target_price = load_price_inventory(tp_path)

    mapping = build_mapping(source_detail, source_price, target_detail, target_price)

    # 수정된 Excel 생성
    os.makedirs(output_dir, exist_ok=True)

    td_base = os.path.splitext(os.path.basename(td_path))[0]
    tp_base = os.path.splitext(os.path.basename(tp_path))[0]

    detail_out = os.path.join(output_dir, f"{td_base}_corrected.xlsx")
    price_out = os.path.join(output_dir, f"{tp_base}_corrected.xlsx")
    report_out = os.path.join(output_dir, "mapping_report.xlsx")

    generate_corrected_detailinfo(mapping, source_detail, td_path, detail_out,
                                  include_fuzzy=include_fuzzy)
    generate_corrected_price(mapping, source_price, tp_path, price_out,
                             include_fuzzy=include_fuzzy)
    generate_mapping_report(mapping, report_out)

    # 변경 요약
    safe_keys = {"barcode", "registered_name"}
    if include_fuzzy:
        safe_keys.add("fuzzy")
    name_cnt = sum(1 for m in mapping["matched"]
                   if m["name_changed"] and m["match_key"] in safe_keys)
    price_cnt = sum(1 for m in mapping["matched"]
                    if m["price_changed"] and m["match_key"] in safe_keys)
    print(f"\n  [적용된 변경]")
    print(f"    상품명 수정: {name_cnt}개")
    print(f"    가격 수정:   {price_cnt}개")
    print(f"\n  [완료] 출력 폴더: {output_dir}")


def _cmd_gap(args, config):
    """미등록 상품 신규 등록 (Phase B)"""
    source = args.source
    target = args.target
    report_path = args.mapping_report
    dry_run = getattr(args, "dry_run", False)
    test = getattr(args, "test", 0)

    mode = "[미리보기]" if dry_run else "[갭 채우기]"
    print(f"\n{mode} {source} → {target}")

    # 매핑 리포트에서 갭 상품 읽기
    if not os.path.exists(report_path):
        raise FileNotFoundError(f"매핑 리포트를 찾을 수 없습니다: {report_path}")

    gap_df = pd.read_excel(report_path, sheet_name="갭_소스에만")

    gap_items = []
    for _, row in gap_df.iterrows():
        gap_items.append({
            "spid": _safe_str(row.get("spid", "")),
            "display_name": _safe_str(row.get("display_name", "")),
            "registered_name": _safe_str(row.get("registered_name", "")),
            "barcode": _safe_str(row.get("barcode", "")),
            "price": _safe_int(row.get("price")),
        })

    mapping_for_gap = {"unmatched_source": gap_items, "matched": []}

    result = register_gap_products(
        source, target, mapping_for_gap,
        dry_run=dry_run, test_limit=test,
    )

    print(f"\n  갭 상품: {result['gap']}개")
    print(f"  등록: {result['created']}, 건너뜀: {result['skipped']}, 오류: {result['error']}")
    if result["errors"]:
        print(f"\n  [오류 목록]")
        for err in result["errors"][:5]:
            print(f"    - {err}")


def _cmd_images(args, config):
    """이미지 동기화 (Phase C)"""
    source = args.source
    target = args.target
    report_path = args.mapping_report
    dry_run = getattr(args, "dry_run", False)
    test = getattr(args, "test", 0)

    mode = "[미리보기]" if dry_run else "[이미지 동기화]"
    print(f"\n{mode} {source} → {target}")

    if not os.path.exists(report_path):
        raise FileNotFoundError(f"매핑 리포트를 찾을 수 없습니다: {report_path}")

    matched_df = pd.read_excel(report_path, sheet_name="매핑결과")

    matched_items = []
    for _, row in matched_df.iterrows():
        matched_items.append({
            "source_spid": _safe_str(row.get("소스_SPID", "")),
            "target_spid": _safe_str(row.get("타겟_SPID", "")),
            "match_key": _safe_str(row.get("매칭방법", "")),
            "source_name": _safe_str(row.get("소스_상품명", "")),
            "target_name": _safe_str(row.get("타겟_상품명", "")),
            "name_changed": row.get("상품명변경") == "O",
            "price_changed": row.get("가격변경") == "O",
        })

    mapping_for_images = {"matched": matched_items, "unmatched_source": []}

    result = sync_images(
        source, target, mapping_for_images,
        dry_run=dry_run, test_limit=test,
    )

    print(f"\n  전체: {result['total']}개")
    print(f"  변경: {result['changed']}, 동일: {result['same']}, 오류: {result['error']}")
    if result["errors"]:
        print(f"\n  [오류 목록]")
        for err in result["errors"][:5]:
            print(f"    - {err}")
    if result["changed"] > 0 and not dry_run:
        print(f"\n  ※ PUT 사용으로 상품 상태가 '임시저장'으로 변경됨")
        print(f"  ※ Wing에서 일괄 재승인이 필요합니다")

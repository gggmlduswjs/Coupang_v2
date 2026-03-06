"""Excel 교정 — 매핑 결과를 기반으로 수정된 Excel 생성."""

import os
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from .mapper import _safe_str


def generate_corrected_detailinfo(mapping: dict,
                                  source_detail: pd.DataFrame,
                                  target_detailinfo_path: str,
                                  output_path: str,
                                  include_fuzzy: bool = False) -> str:
    """타겟 detailinfo Excel에 007-ez의 수정된 노출상품명 반영.

    openpyxl로 원본 로드 → 매핑된 행의 노출상품명 컬럼만 수정 → 저장.
    include_fuzzy=False(기본)이면 퍼지 매칭은 제외 (오매핑 위험).
    """
    if not os.path.exists(target_detailinfo_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {target_detailinfo_path}")

    # 매핑에서 이름이 변경된 것만 필터
    name_changes: dict[str, str] = {}  # target_spid → new_name
    for m in mapping["matched"]:
        if m["name_changed"]:
            if m["match_key"] == "fuzzy" and not include_fuzzy:
                continue
            name_changes[m["target_spid"]] = m["source_name"]

    if not name_changes:
        print("  상품명 변경 사항 없음")
        return ""

    # openpyxl로 원본 로드
    wb = load_workbook(target_detailinfo_path)
    try:
        ws = wb["Template"]
    except KeyError:
        ws = wb.active

    # 헤더 찾기 (4번째 행, openpyxl 1-based = row 4)
    header = [_safe_str(ws.cell(row=4, column=c).value)
              for c in range(1, ws.max_column + 1)]

    # 컬럼 인덱스 찾기 (1-based)
    spid_col = _find_col_index(header, ["등록상품ID", "업체상품ID", "업체상품 ID"])
    name_col = _find_col_index(header, ["쿠팡 노출상품명", "노출상품명"])

    if not spid_col or not name_col:
        raise ValueError(f"필수 컬럼을 찾을 수 없습니다. 헤더: {header[:15]}")

    # 데이터 행은 5번째부터 (헤더가 4행)
    modified = 0
    for row_num in range(5, ws.max_row + 1):
        spid = _safe_str(ws.cell(row=row_num, column=spid_col).value)
        if spid in name_changes:
            old_name = _safe_str(ws.cell(row=row_num, column=name_col).value)
            new_name = name_changes[spid]
            ws.cell(row=row_num, column=name_col).value = new_name
            modified += 1

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()

    print(f"  detailinfo 수정 완료: {modified}개 상품명 변경 → {output_path}")
    return output_path


def generate_corrected_price(mapping: dict,
                             source_price: pd.DataFrame,
                             target_price_path: str,
                             output_path: str,
                             include_fuzzy: bool = False) -> str:
    """타겟 price_inventory Excel에 007-ez의 수정된 가격 반영.

    openpyxl로 원본 로드 → 매핑된 행의 판매가격 컬럼만 수정 → 저장.
    include_fuzzy=False(기본)이면 퍼지 매칭은 제외.
    """
    if not os.path.exists(target_price_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {target_price_path}")

    # 매핑에서 가격이 변경된 것만 필터
    price_changes: dict[str, int] = {}  # target_spid → new_price
    for m in mapping["matched"]:
        if m["price_changed"] and m["source_price"] is not None:
            if m["match_key"] == "fuzzy" and not include_fuzzy:
                continue
            price_changes[m["target_spid"]] = m["source_price"]

    if not price_changes:
        print("  가격 변경 사항 없음")
        return ""

    # openpyxl로 원본 로드
    wb = load_workbook(target_price_path)
    try:
        ws = wb["data"]
    except KeyError:
        ws = wb.active

    # 헤더 찾기 (3번째 행, openpyxl 1-based = row 3)
    header = [_safe_str(ws.cell(row=3, column=c).value)
              for c in range(1, ws.max_column + 1)]

    spid_col = _find_col_index(header, ["업체상품 ID", "업체상품ID"])
    # 수정 요청용 판매가격 컬럼 (O열=16, "변경/수정 요청" 섹션)
    # Row3 헤더에서 두 번째 "판매가격" 찾기 (첫 번째는 C10=조회용, 두 번째는 C16=수정용)
    price_col = None
    for i, h in enumerate(header):
        if "판매가격" in h and i >= 15:  # C16 이후 (0-indexed=15)
            price_col = i + 1
            break
    if not price_col:
        # 못 찾으면 C16 직접 사용 (수정 요청 판매가격)
        price_col = 16

    if not spid_col:
        raise ValueError(f"SPID 컬럼을 찾을 수 없습니다. 헤더: {header[:15]}")

    # 데이터 행은 4번째부터
    modified = 0
    for row_num in range(4, ws.max_row + 1):
        spid = _safe_str(ws.cell(row=row_num, column=spid_col).value)
        if spid in price_changes:
            ws.cell(row=row_num, column=price_col).value = price_changes[spid]
            modified += 1

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()

    print(f"  price_inventory 수정 완료: {modified}개 가격 변경 → {output_path}")
    return output_path


def generate_mapping_report(mapping: dict, output_path: str) -> str:
    """매핑 결과 Excel 리포트 생성.

    시트1: 매핑된 상품 (source<->target, 변경 내용)
    시트2: 미매핑 소스 상품 (갭 상품)
    시트3: 미매핑 타겟 상품
    시트4: 통계 요약
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 시트1: 매핑된 상품
        if mapping["matched"]:
            matched_data = []
            for m in mapping["matched"]:
                matched_data.append({
                    "소스_SPID": m["source_spid"],
                    "타겟_SPID": m["target_spid"],
                    "매칭방법": m["match_key"],
                    "소스_상품명": m["source_name"],
                    "타겟_상품명": m["target_name"],
                    "소스_가격": m.get("source_price"),
                    "타겟_가격": m.get("target_price"),
                    "상품명변경": "O" if m["name_changed"] else "",
                    "가격변경": "O" if m["price_changed"] else "",
                    "유사도": m.get("fuzzy_score", ""),
                    "소스_바코드": m.get("source_barcode", ""),
                })
            pd.DataFrame(matched_data).to_excel(
                writer, sheet_name="매핑결과", index=False)

        # 시트2: 미매핑 소스 (갭 상품)
        if mapping["unmatched_source"]:
            pd.DataFrame(mapping["unmatched_source"]).to_excel(
                writer, sheet_name="갭_소스에만", index=False)

        # 시트3: 미매핑 타겟
        if mapping["unmatched_target"]:
            pd.DataFrame(mapping["unmatched_target"]).to_excel(
                writer, sheet_name="미매핑_타겟에만", index=False)

        # 시트4: 통계
        stats = mapping["stats"]
        name_changed = sum(1 for m in mapping["matched"] if m["name_changed"])
        price_changed = sum(1 for m in mapping["matched"] if m["price_changed"])

        stats_data = [
            {"항목": "바코드 매칭", "건수": stats["barcode"]},
            {"항목": "등록상품명 매칭", "건수": stats["registered_name"]},
            {"항목": "퍼지 매칭", "건수": stats["fuzzy"]},
            {"항목": "총 매핑", "건수": stats["barcode"] + stats["registered_name"] + stats["fuzzy"]},
            {"항목": "", "건수": ""},
            {"항목": "상품명 변경 대상", "건수": name_changed},
            {"항목": "가격 변경 대상", "건수": price_changed},
            {"항목": "", "건수": ""},
            {"항목": "미매핑 (소스=갭)", "건수": stats["unmatched_source"]},
            {"항목": "미매핑 (타겟)", "건수": stats["unmatched_target"]},
        ]
        pd.DataFrame(stats_data).to_excel(
            writer, sheet_name="통계", index=False)

    print(f"  매핑 리포트 생성: {output_path}")
    return output_path


def _find_col_index(header: list[str], candidates: list[str]) -> Optional[int]:
    """헤더 리스트에서 후보 컬럼명의 1-based 인덱스 반환."""
    for c in candidates:
        for i, h in enumerate(header):
            if c == h or c in h:
                return i + 1  # 1-based
    return None
